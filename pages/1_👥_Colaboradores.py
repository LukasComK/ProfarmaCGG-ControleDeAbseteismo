"""
Página de Processamento de Colaboradores
Streamlit multi-page application
"""

import streamlit as st
import pandas as pd
import io
import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from funcoes_processamento_csv import processar_csv_colaboradores, validar_csv, extrair_tabela_supervisores

# Mapa de cores do app.py
MAPA_CORES = {
    'P': 'FF90EE90',      # Verde claro
    'FI': 'FFFF0000',     # Vermelho puro
    'FA': 'FFFFFF00',     # Amarelo puro
    'Afastamento': 'FFC0C0C0',  # Cinza
    'FERIADO': 'FF000000',      # Preto (com texto branco)
    'FÉRIAS-BH': 'FF000000',    # Preto (com texto branco)
    'DESLIGADO': 'FF800080',   # Roxo
    'DESCANSO': 'FFC0C0C0'  # Cinza
}

st.set_page_config(page_title="Processamento de Colaboradores", layout="wide")

st.header("👥 Processamento de CSV de Colaboradores")
st.write("Extraia dados de colaboradores, calcule turnos e supervisores")

st.divider()

# Upload do CSV
uploaded_file = st.file_uploader(
    "📤 Faça upload do CSV com dados de colaboradores",
    type=["csv", "xlsx", "xlsm"],
    key="colaboradores_uploader"
)

if uploaded_file is not None:
    try:
        # Lê o arquivo
        if uploaded_file.name.endswith('.xlsx'):
            df = pd.read_excel(uploaded_file)
        else:
            # Tenta diferentes encodings e separadores
            import csv as csv_module
            import io as io_module
            
            encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'utf-16']
            separators = [',', ';', '\t', '|']
            df = None
            encoding_usado = None
            separador_usado = None
            skiprows_usado = None
            
            # Lê o arquivo como bytes uma única vez
            uploaded_file.seek(0)
            file_bytes = uploaded_file.read()
            
            # Detecta se há linha de título "Colaboradores"
            primeira_linha = file_bytes[:200].decode('latin-1', errors='ignore')
            tem_titulo = 'colaborador' in primeira_linha.lower()[:100]
            
            st.write(f"🔍 **Debug:** Primeiros caracteres: `{primeira_linha[:100]}`")
            st.write(f"🔍 **Debug:** Título detectado: {'Sim' if tem_titulo else 'Não'}")
            
            # MÉTODO ÚNICO: Leitura direta com os parâmetros corretos para este arquivo
            st.write("🔄 **Lendo arquivo CSV...**")
            try:
                uploaded_file.seek(0)
                # Este arquivo usa: encoding latin-1, separador ;, e tem linha de título "Colaboradores"
                df = pd.read_csv(uploaded_file, sep=';', encoding='latin-1', skiprows=1,
                                engine='python', on_bad_lines='skip')
                
                if len(df.columns) > 3:
                    encoding_usado = 'latin-1'
                    separador_usado = ';'
                    skiprows_usado = 1
                    st.success(f"✅ Arquivo lido com sucesso!")
                    st.write(f"   - Encoding: `latin-1`")
                    st.write(f"   - Separador: `;`")
                    st.write(f"   - Skip rows: `1`")
                    st.write(f"   - Colunas encontradas: `{len(df.columns)}`")
                else:
                    st.error(f"❌ Arquivo lido mas com apenas {len(df.columns)} colunas (esperado > 3)")
                    df = None
            except Exception as e:
                st.error(f"❌ Erro ao ler arquivo: {str(e)}")
                import traceback
                st.write(traceback.format_exc())
                df = None
        
        st.success(f"✅ Arquivo carregado: {uploaded_file.name}")
        st.write(f"Total de linhas: {len(df)}")
        st.write(f"Colunas encontradas: {len(df.columns)}")
        
        # Preview do arquivo carregado
        with st.expander("📋 Preview do Arquivo Carregado (primeiras 10 linhas)", expanded=False):
            st.dataframe(df.head(10), use_container_width=True)
            st.write(f"**Dimensões:** {df.shape[0]} linhas × {df.shape[1]} colunas")
            st.write(f"**Colunas:** {list(df.columns)}")
        
        # Valida o CSV
        é_válido, erros, colunas = validar_csv(df)
        
        if not é_válido:
            st.error("❌ CSV inválido! Erros encontrados:")
            for erro in erros:
                st.write(f"  • {erro}")
        else:
            st.success("✅ CSV validado com sucesso!")
            
            # DEBUG: Mostra a tabela de ENCARREGADOS e seus gestores
            st.divider()
            st.subheader("🔧 DEBUG: Tabela de Supervisores Extraída")
            
            try:
                tabela_supervisores = extrair_tabela_supervisores(df, colunas)
                
                if tabela_supervisores:
                    # Converte para DataFrame para melhor visualização
                    df_debug = pd.DataFrame([
                        {"Encarregado": k, "Gestor": v}
                        for k, v in tabela_supervisores.items()
                    ])
                    
                    st.write(f"**Total de Encarregados Encontrados:** {len(df_debug)}")
                    st.dataframe(df_debug, use_container_width=True, hide_index=True)
                else:
                    st.warning("⚠️ Nenhum encarregado encontrado no CSV!")
            except Exception as e:
                st.error(f"❌ Erro ao extrair supervisores: {str(e)}")
                import traceback
                st.write(traceback.format_exc())
            
            st.divider()
            with st.expander("🔍 Cargos a filtrar", expanded=True):
                cargos_padrao = [
                    "AUXILIAR DEPOSITO I",
                    "AUXILIAR DEPOSITO II",
                    "AUXILIAR DEPOSITO III",
                    "OPERADOR EMPILHADEIRA"
                ]
                
                st.write("Cargos selecionados para filtro:")
                for cargo in cargos_padrao:
                    st.write(f"  ✓ {cargo}")
            
            # PROCESSA AUTOMATICAMENTE (sem botão)
            with st.spinner("⏳ Processando colaboradores..."):
                try:
                    # Processa o DataFrame
                    df_resultado, info = processar_csv_colaboradores(df)
                    
                    # Exibe informações de processamento
                    col_info1, col_info2, col_info3 = st.columns(3)
                    
                    with col_info1:
                        st.metric("📝 Total Original", info["total_linhas_original"])
                    
                    with col_info2:
                        st.metric("✅ Processado", info["total_linhas_processado"])
                    
                    with col_info3:
                        st.metric("❌ Filtrado", info["linhas_filtradas"])
                    
                    st.divider()
                    
                    # Exibe preview dos dados
                    st.subheader("📋 Preview dos Dados Processados")
                    st.dataframe(df_resultado, use_container_width=True)
                    
                    st.divider()
                    
                    # Download do arquivo de colaboradores
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_resultado.to_excel(writer, index=False, sheet_name='Colaboradores')
                        
                        # Formata o Excel
                        ws = writer.sheets['Colaboradores']
                        
                        # Header
                        header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
                        header_font = Font(bold=True, color='FFFFFFFF', size=11)
                        
                        for col_idx in range(1, len(df_resultado.columns) + 1):
                            cell = ws.cell(row=1, column=col_idx)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        # Ajusta largura das colunas
                        for col_idx, col_name in enumerate(df_resultado.columns, 1):
                            max_len = df_resultado[col_name].astype(str).str.len().max()
                            adjusted_width = min(max(max_len + 2, 10), 50)
                            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
                    
                    output.seek(0)
                    
                    # Botão de download
                    st.download_button(
                        label="📥 Baixar XLSX (Colaboradores)",
                        data=output.getvalue(),
                        file_name=f"colaboradores_processado_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_colaboradores"
                    )
                    
                    st.divider()
                    
                    # SEÇÃO: Criar Planilha Mestra de Controle de Absenteísmo
                    st.subheader("📅 Criar Planilha Mestra de Controle de Absenteísmo")
                    
                    # Selectbox para escolher mês/ano
                    col_mes, col_ano = st.columns(2)
                    
                    meses_pt = {
                        1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
                        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
                        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
                    }
                    
                    with col_mes:
                        mes_selecionado = st.selectbox(
                            "Selecione o Mês:",
                            options=list(meses_pt.keys()),
                            format_func=lambda x: meses_pt[x],
                            key="selectbox_mes_abs"
                        )
                    
                    with col_ano:
                        ano_selecionado = st.number_input(
                            "Selecione o Ano:",
                            min_value=2020,
                            max_value=2100,
                            value=datetime.datetime.now().year,
                            key="input_ano_abs"
                        )
                    
                    # Botão para gerar planilha mestra
                    if st.button("🎯 Gerar Planilha Mestra", key="btn_gerar_mestra"):
                        with st.spinner("⏳ Gerando planilha mestra..."):
                            try:
                                import calendar
                                
                                # Gera lista de dias do mês
                                num_dias = calendar.monthrange(ano_selecionado, mes_selecionado)[1]
                                dias_mes = list(range(1, num_dias + 1))
                                
                                # Cria DataFrame com TODAS as colunas do df_resultado (como no app.py)
                                df_mestra = df_resultado.copy()
                                
                                # Renomeia as colunas para ficar IGUAL ao app.py
                                rename_dict = {
                                    "Colaborador": "NOME",
                                    "Cargo": "FUNÇÃO",
                                    "Descrição Situação": "SITUAÇÃO",
                                    "Descrição CC": "AREA",
                                    "Nome Gestor": "GESTOR",
                                    "Supervisor": "SUPERVISOR",
                                    "Descrição da Unidade Organizacional": "SETOR",
                                    "Turno": "TURNO",
                                    "Jornada": "HORARIO"
                                }
                                
                                df_mestra = df_mestra.rename(columns=rename_dict)
                                
                                # Adiciona colunas para cada dia do mês (formato: 01/11, 02/11, etc - como no app.py)
                                for dia in dias_mes:
                                    data_obj = datetime.date(ano_selecionado, mes_selecionado, dia)
                                    # Formato com mês: "01/11", "02/11", etc
                                    nome_coluna = f"{dia:02d}/{mes_selecionado:02d}"
                                    df_mestra[nome_coluna] = ""
                                
                                # Download da planilha mestra
                                output_mestra = io.BytesIO()
                                with pd.ExcelWriter(output_mestra, engine='openpyxl') as writer:
                                    df_mestra.to_excel(writer, index=False, sheet_name='Dados')
                                    
                                    # Formata o Excel EXATAMENTE COMO NO APP.PY
                                    ws_mestra = writer.sheets['Dados']
                                    
                                    # Função para calcular largura baseada no maior valor da coluna
                                    def calc_width(df, col_name, min_width=10, max_width=50):
                                        if col_name not in df.columns:
                                            return min_width
                                        max_len = df[col_name].astype(str).str.len().max()
                                        header_len = len(str(col_name))
                                        largest = max(max_len, header_len)
                                        width = min(max(largest + 2, min_width), max_width)
                                        return width
                                    
                                    # Header - Cores baseadas no app.py
                                    header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
                                    header_font = Font(bold=True, color='FFFFFFFF', size=11)
                                    
                                    # Formata header
                                    for col_idx in range(1, len(df_mestra.columns) + 1):
                                        cell = ws_mestra.cell(row=1, column=col_idx)
                                        cell.fill = header_fill
                                        cell.font = header_font
                                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                    
                                    # Aplica cores e largura para CADA coluna (como no app.py)
                                    col_names = df_mestra.columns.tolist()
                                    
                                    for col_idx, col_name in enumerate(col_names, 1):
                                        col_letter = get_column_letter(col_idx)
                                        
                                        # Define preenchimento e largura para cada tipo de coluna
                                        if col_name == 'NOME':
                                            col_fill = PatternFill(start_color='FFCCE5FF', end_color='FFCCE5FF', fill_type='solid')  # Azul claro suave
                                            width = calc_width(df_mestra, col_name, min_width=15, max_width=40)
                                            ws_mestra.column_dimensions[col_letter].width = width
                                        elif col_name == 'AREA':
                                            col_fill = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')  # Verde claro suave
                                            ws_mestra.column_dimensions[col_letter].width = 25
                                        elif col_name == 'GESTOR':
                                            col_fill = PatternFill(start_color='FFffbf5e', end_color='FFffbf5e', fill_type='solid')  # Laranja #ffbf5e
                                            width = calc_width(df_mestra, col_name, min_width=15, max_width=40)
                                            ws_mestra.column_dimensions[col_letter].width = width
                                        else:
                                            col_fill = None
                                            # Verifica se é coluna de data (formato DD/MM) ou outra coluna padrão
                                            try:
                                                datetime.datetime.strptime(str(col_name), '%d/%m')
                                                # É uma coluna de data - auto-fit reduzido
                                                ws_mestra.column_dimensions[col_letter].width = 7
                                            except:
                                                # Outras colunas - auto-fit normal
                                                width = calc_width(df_mestra, col_name, min_width=10, max_width=25)
                                                ws_mestra.column_dimensions[col_letter].width = width
                                        
                                        # Aplica a cor a todas as linhas de dados desta coluna
                                        if col_fill is not None:
                                            for row_idx in range(2, ws_mestra.max_row + 1):
                                                cell = ws_mestra.cell(row=row_idx, column=col_idx)
                                                cell.fill = col_fill
                                    
                                    # Formata células de dias com branco e centrado
                                    white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
                                    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                    
                                    for col_idx, col_name in enumerate(col_names, 1):
                                        # Verifica se é coluna de data
                                        try:
                                            datetime.datetime.strptime(str(col_name), '%d/%m')
                                            # É uma coluna de data
                                            for row_idx in range(2, ws_mestra.max_row + 1):
                                                cell = ws_mestra.cell(row=row_idx, column=col_idx)
                                                cell.fill = white_fill
                                                cell.alignment = center_alignment
                                        except:
                                            pass
                                
                                output_mestra.seek(0)
                                
                                # Exibe preview
                                st.success(f"✅ Planilha mestra gerada para {meses_pt[mes_selecionado]}/{ano_selecionado}")
                                st.dataframe(df_mestra, use_container_width=True)
                                
                                # Botão de download
                                st.download_button(
                                    label="📥 Baixar Planilha Mestra XLSX",
                                    data=output_mestra.getvalue(),
                                    file_name=f"controle_abs_mestra_{ano_selecionado}{mes_selecionado:02d}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="download_mestra"
                                )
                                
                            except Exception as e:
                                st.error(f"❌ Erro ao gerar planilha mestra: {str(e)}")
                                import traceback
                                st.write(traceback.format_exc())
                    
                except Exception as e:
                    st.error(f"❌ Erro ao processar: {str(e)}")
                    import traceback
                    st.write(traceback.format_exc())
    
    except Exception as e:
        st.error(f"❌ Erro ao ler arquivo: {str(e)}")
else:
    st.info("👆 Selecione um arquivo CSV ou XLSX para começar")
