"""
Página: Banco de Horas
Descrição: Visualização e geração de relatório de banco de horas com consolidação e TOP 15
Recebe: Arquivo XLSX com banco de horas (coluna E: CentroDeCustos, coluna G: Nomes, coluna R: SaldoFinal)
"""

import streamlit as st
import pandas as pd
import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from funcoes_processamento_csv import determinar_turno

st.set_page_config(page_title="Banco de Horas", layout="wide")

st.title("🏦 Banco de Horas")
st.write("Relatório consolidado de banco de horas com TOP 15 ofensores")

st.divider()

# Upload dos arquivos em duas colunas
st.subheader("📥 Carregue os arquivos necessários")
col1, col2 = st.columns(2)

with col1:
    st.write("**Arquivo de Banco de Horas**")
    file_banco_horas = st.file_uploader(
        "Arquivo XLSX com banco de horas",
        type=["xlsx"],
        key="banco_horas"
    )

with col2:
    st.write("**BASE CSV de Colaboradores**")
    file_csv_colaboradores = st.file_uploader(
        "Arquivo CSV com dados de colaboradores (colunas: Colaborador, Nome Gestor)",
        type=["csv"],
        key="csv_colaboradores"
    )

if file_banco_horas and file_csv_colaboradores:
    try:
        # Carrega o arquivo
        df = pd.read_excel(file_banco_horas)
        
        st.success("✅ Arquivo carregado com sucesso!")
        
        st.divider()
        
        # Checkbox para Dia de Fechamento
        eh_dia_fechamento = st.checkbox("📅 Dia de Fechamento (usar Pagamentos e Descontos)", value=False)
        
        if eh_dia_fechamento:
            st.info("ℹ️ Usando colunas **Pagamentos** (horas positivas) e **Descontos** (horas negativas)")
        else:
            st.info("ℹ️ Usando coluna **SaldoFinal** para análise de banco de horas")
        
        st.divider()
        
        # Botão para gerar relatório
        if st.button("📊 Gerar Relatório", use_container_width=True):
            try:
                # Verifica se as colunas esperadas existem
                if 'CentroDeCustos' not in df.columns:
                    st.error("❌ Coluna 'CentroDeCustos' não encontrada!")
                    st.write(f"Colunas disponíveis: {list(df.columns)}")
                elif eh_dia_fechamento and 'Pagamentos' not in df.columns:
                    st.error("❌ Coluna 'Pagamentos' não encontrada!")
                    st.write(f"Colunas disponíveis: {list(df.columns)}")
                elif eh_dia_fechamento and 'Descontos' not in df.columns:
                    st.error("❌ Coluna 'Descontos' não encontrada!")
                    st.write(f"Colunas disponíveis: {list(df.columns)}")
                elif not eh_dia_fechamento and 'SaldoFinal' not in df.columns:
                    st.error("❌ Coluna 'SaldoFinal' não encontrada!")
                    st.write(f"Colunas disponíveis: {list(df.columns)}")
                else:
                    # ===== CARREGAMENTO E PROCESSAMENTO DOS DADOS =====
                    df_processado = df[['CentroDeCustos', df.columns[6]]].copy()
                    df_processado.columns = ['CentroDeCustos', 'Colaborador']
                    
                    # Remove linhas onde centro de custo está vazio
                    df_processado = df_processado[df_processado['CentroDeCustos'].notna()]
                    df_processado = df_processado[df_processado['CentroDeCustos'].astype(str).str.strip() != '']
                    df_processado = df_processado[df_processado['CentroDeCustos'] != 'P']
                    df_processado = df_processado[df_processado['CentroDeCustos'].str.strip().str.upper() != 'RECURSOS HUMANOS']
                    
                    # Cria barra de progresso
                    progress_bar = st.progress(0, text="⏳ Processando dados...")
                    status_text = st.empty()
                    
                    # Define a coluna de origem
                    if eh_dia_fechamento:
                        # Modo Dia de Fechamento
                        # Copia dados do DataFrame original ANTES de qualquer filtro
                        original_indices = df_processado.index.copy()
                        df_processado['Pagamentos'] = df.loc[original_indices, 'Pagamentos'].values
                        df_processado['Descontos'] = df.loc[original_indices, 'Descontos'].values
                        
                        # Função auxiliar para verificar se valor é considerado "vazio"
                        def is_empty_value(x):
                            if pd.isna(x):
                                return True
                            if x == '' or x == 0:
                                return True
                            # Para valores datetime.time, nunca são vazios se não forem NaN
                            if hasattr(x, 'hour'):
                                return False
                            # Para strings, verifica se é "00:00:00" ou similar
                            str_val = str(x).strip()
                            if str_val in ['0', 'nan', '0:00:00', '00:00:00']:
                                return True
                            return False
                        
                        # Marca quais valores são vazios (será usado para filtro)
                        df_processado['Pagamentos_tem_valor'] = df_processado['Pagamentos'].apply(lambda x: not is_empty_value(x))
                        df_processado['Descontos_tem_valor'] = df_processado['Descontos'].apply(lambda x: not is_empty_value(x))
                        
                        # Filtra registros que têm pelo menos Pagamentos OU Descontos com valor
                        df_processado = df_processado[(df_processado['Pagamentos_tem_valor']) | (df_processado['Descontos_tem_valor'])].reset_index(drop=True)
                        
                        # Agora sim, cria as versões "limpas" para conversão
                        df_processado['Pagamentos_limpo'] = df_processado['Pagamentos'].apply(lambda x: x if not is_empty_value(x) else 0)
                        df_processado['Descontos_limpo'] = df_processado['Descontos'].apply(lambda x: x if not is_empty_value(x) else 0)
                        
                        coluna_processamento = None
                        tipo_mode = "Fechamento"
                    else:
                        # Modo SaldoFinal
                        original_indices = df_processado.index.copy()
                        df_processado['SaldoFinal'] = df.loc[original_indices, 'SaldoFinal'].values
                        coluna_processamento = 'SaldoFinal'
                        tipo_mode = "SaldoFinal"
                    
                    # Função para converter tempo para horas decimais
                    def tempo_para_horas(valor):
                        """Converte tempo em formato HH:MM:SS para horas decimais, detectando sinal"""
                        if pd.isna(valor) or valor == '' or valor == 0:
                            return 0.0, 0.0
                        
                        try:
                            if hasattr(valor, 'hour'):
                                horas_dec = valor.hour + valor.minute/60 + valor.second/3600
                                return horas_dec, 0.0
                            
                            valor_str = str(valor).strip()
                            if not valor_str or valor_str == 'nan' or valor_str == '0':
                                return 0.0, 0.0
                            
                            eh_negativo = valor_str.startswith('-')
                            valor_limpo = valor_str.lstrip('-').strip()
                            
                            if ':' in valor_limpo:
                                partes = valor_limpo.split(':')
                                if len(partes) >= 2:
                                    try:
                                        horas = float(partes[0])
                                        minutos = float(partes[1])
                                        segundos = float(partes[2]) if len(partes) > 2 else 0.0
                                        total = horas + minutos/60 + segundos/3600
                                        return (0.0, total) if eh_negativo else (total, 0.0)
                                    except ValueError:
                                        return 0.0, 0.0
                            else:
                                try:
                                    num = float(valor_limpo)
                                    return (0.0, num) if eh_negativo else (num, 0.0)
                                except ValueError:
                                    return 0.0, 0.0
                        except Exception as e:
                            st.warning(f"Erro ao processar valor: {valor} - {e}")
                            return 0.0, 0.0
                        
                        return 0.0, 0.0
                    
                    # Função para converter tempo para horas decimais (apenas positivo)
                    def tempo_para_horas_simples(valor):
                        """Converte tempo em formato HH:MM:SS para horas decimais (valores positivos - remove sinal negativo)"""
                        if pd.isna(valor) or valor == '' or valor == 0:
                            return 0.0
                        
                        try:
                            if hasattr(valor, 'hour'):
                                result = valor.hour + valor.minute/60 + valor.second/3600
                                return abs(result)  # Retorna valor absoluto
                            
                            valor_str = str(valor).strip()
                            if not valor_str or valor_str == 'nan' or valor_str == '0':
                                return 0.0
                            
                            # Remove sinal negativo se existir
                            valor_str = valor_str.lstrip('-').strip()
                            
                            if ':' in valor_str:
                                partes = valor_str.split(':')
                                if len(partes) >= 2:
                                    try:
                                        horas = float(partes[0])
                                        minutos = float(partes[1])
                                        segundos = float(partes[2]) if len(partes) > 2 else 0.0
                                        result = horas + minutos/60 + segundos/3600
                                        return result  # Já é positivo após lstrip
                                    except ValueError:
                                        return 0.0
                            else:
                                try:
                                    return abs(float(valor_str))  # Valor absoluto para números
                                except ValueError:
                                    return 0.0
                        except Exception as e:
                            st.warning(f"Erro em tempo_para_horas_simples: {valor} - {e}")
                            return 0.0
                        
                        return 0.0
                    
                    # Processa os dados baseado no modo
                    if eh_dia_fechamento:
                        df_processado[['POSITIVO_num', 'NEGATIVO_num']] = df_processado.apply(
                            lambda row: pd.Series([
                                tempo_para_horas_simples(row['Pagamentos_limpo']),
                                tempo_para_horas_simples(row['Descontos_limpo'])
                            ]),
                            axis=1
                        )
                    else:
                        df_processado[['POSITIVO_num', 'NEGATIVO_num']] = df_processado['SaldoFinal'].apply(
                            lambda x: pd.Series(tempo_para_horas(x))
                        )
                    
                    progress_bar.progress(25, text="⏳ Processando dados... (25%)")
                    status_text.text("Convertendo horas...")
                    
                    # Função para converter horas decimais de volta para HH:MM:SS
                    def horas_para_tempo(horas):
                        if pd.isna(horas) or horas == 0:
                            return "00:00:00"
                        
                        total_segundos = int(abs(horas) * 3600)
                        h = total_segundos // 3600
                        m = (total_segundos % 3600) // 60
                        s = total_segundos % 60
                        return f"{h:02d}:{m:02d}:{s:02d}"
                    
                    # ===== PREPARAÇÃO DOS DATAFRAMES =====
                    df_resumo = df_processado.groupby('CentroDeCustos')[['POSITIVO_num', 'NEGATIVO_num']].sum().reset_index()
                    df_resumo.columns = ['Centro de Custo', 'POSITIVO_num', 'NEGATIVO_num']
                    df_resumo['POSITIVO'] = df_resumo['POSITIVO_num'].apply(horas_para_tempo)
                    df_resumo['NEGATIVO'] = df_resumo['NEGATIVO_num'].apply(horas_para_tempo)
                    
                    df_top15_pos = df_processado.nlargest(15, 'POSITIVO_num')[['Colaborador', 'CentroDeCustos', 'POSITIVO_num']].copy()
                    df_top15_pos['POSITIVO'] = df_top15_pos['POSITIVO_num'].apply(horas_para_tempo)
                    df_top15_pos = df_top15_pos.reset_index(drop=True)
                    df_top15_pos.index = df_top15_pos.index + 1
                    
                    df_top15_neg = df_processado.nlargest(15, 'NEGATIVO_num')[['Colaborador', 'CentroDeCustos', 'NEGATIVO_num']].copy()
                    df_top15_neg['NEGATIVO'] = df_top15_neg['NEGATIVO_num'].apply(horas_para_tempo)
                    df_top15_neg = df_top15_neg.reset_index(drop=True)
                    df_top15_neg.index = df_top15_neg.index + 1
                    
                    progress_bar.progress(75, text="⏳ Gerando arquivo Excel... (75%)")
                    status_text.text("Criando sheets...")
                    
                    # Cria arquivo Excel para download
                    wb = Workbook()
                    
                    # ===== SHEET 1: CONSOLIDAÇÃO =====
                    ws1 = wb.active
                    ws1.title = "Consolidação"
                    
                    # Define estilos Profarma
                    header_fill_principal = PatternFill(start_color="FF275316", end_color="FF275316", fill_type="solid")
                    header_fill_horas = PatternFill(start_color="FFC0E6F5", end_color="FFC0E6F5", fill_type="solid")
                    header_font_principal = Font(bold=True, color="FFFFFFFF", size=12, name="Calibri")
                    header_font_horas = Font(bold=True, color="FF000000", size=12, name="Calibri")
                    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    border_style = Side(style="thin", color="000000")
                    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                    
                    center_alignment = Alignment(horizontal="center", vertical="center")
                    left_alignment = Alignment(horizontal="left", vertical="center")
                    
                    # Calcula totais
                    total_positivo = df_resumo['POSITIVO_num'].sum()
                    total_negativo = df_resumo['NEGATIVO_num'].sum()
                    
                    # ===== SEÇÃO 1: RÓTULOS E SOMA DE SALDO (Colunas B-C) =====
                    # Linha 2: Headers
                    ws1.cell(row=2, column=2, value="ROTULO DE LINHA")
                    ws1.cell(row=2, column=2).fill = header_fill_principal
                    ws1.cell(row=2, column=2).font = header_font_principal
                    ws1.cell(row=2, column=2).alignment = header_alignment
                    ws1.cell(row=2, column=2).border = border
                    
                    ws1.cell(row=2, column=3, value="SOMA DE SALDO")
                    ws1.cell(row=2, column=3).fill = header_fill_horas
                    ws1.cell(row=2, column=3).font = header_font_horas
                    ws1.cell(row=2, column=3).alignment = header_alignment
                    ws1.cell(row=2, column=3).border = border
                    
                    # Define cores branco e preto para dados
                    white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
                    black_font = Font(bold=False, size=11, color="FF000000", name="Calibri")
                    
                    # Linha 3: POSITIVO
                    ws1.cell(row=3, column=2, value="POSITIVO")
                    ws1.cell(row=3, column=2).fill = white_fill
                    ws1.cell(row=3, column=2).font = black_font
                    ws1.cell(row=3, column=2).alignment = left_alignment
                    ws1.cell(row=3, column=2).border = border
                    
                    ws1.cell(row=3, column=3, value=horas_para_tempo(total_positivo))
                    ws1.cell(row=3, column=3).fill = white_fill
                    ws1.cell(row=3, column=3).font = black_font
                    ws1.cell(row=3, column=3).alignment = center_alignment
                    ws1.cell(row=3, column=3).border = border
                    
                    # Linha 4: NEGATIVO
                    ws1.cell(row=4, column=2, value="NEGATIVO")
                    ws1.cell(row=4, column=2).fill = white_fill
                    ws1.cell(row=4, column=2).font = black_font
                    ws1.cell(row=4, column=2).alignment = left_alignment
                    ws1.cell(row=4, column=2).border = border
                    
                    ws1.cell(row=4, column=3, value=horas_para_tempo(total_negativo))
                    ws1.cell(row=4, column=3).fill = white_fill
                    ws1.cell(row=4, column=3).font = black_font
                    ws1.cell(row=4, column=3).alignment = center_alignment
                    ws1.cell(row=4, column=3).border = border
                    
                    # Linha 5: Total Geral
                    total_geral = total_positivo + abs(total_negativo)
                    ws1.cell(row=5, column=2, value="Total Geral")
                    ws1.cell(row=5, column=2).fill = header_fill_principal
                    ws1.cell(row=5, column=2).font = header_font_principal
                    ws1.cell(row=5, column=2).alignment = left_alignment
                    ws1.cell(row=5, column=2).border = border
                    
                    ws1.cell(row=5, column=3, value=horas_para_tempo(total_geral))
                    ws1.cell(row=5, column=3).fill = header_fill_horas
                    ws1.cell(row=5, column=3).font = header_font_horas
                    ws1.cell(row=5, column=3).alignment = center_alignment
                    ws1.cell(row=5, column=3).border = border
                    
                    # Headers
                    headers = ['Centro de Custo', 'POSITIVO', 'NEGATIVO']
                    for col_idx, header in enumerate(headers, 5):
                        cell = ws1.cell(row=1, column=col_idx, value=header)
                        if col_idx == 5:
                            cell.fill = header_fill_principal
                            cell.font = header_font_principal
                        else:
                            cell.fill = header_fill_horas
                            cell.font = header_font_horas
                        cell.alignment = header_alignment
                        cell.border = border
                    
                    ws1.row_dimensions[1].height = 25
                    
                    # Dados consolidação
                    for row_idx, (_, row) in enumerate(df_resumo.iterrows(), 2):
                        ws1.cell(row=row_idx, column=5, value=row['Centro de Custo'])
                        ws1.cell(row=row_idx, column=6, value=row['POSITIVO'])
                        
                        if row['NEGATIVO_num'] > 0:
                            ws1.cell(row=row_idx, column=7, value=row['NEGATIVO'])
                        else:
                            ws1.cell(row=row_idx, column=7, value='')
                        
                        for col in [5, 6, 7]:
                            cell = ws1.cell(row=row_idx, column=col)
                            cell.border = border
                            if col == 5:
                                cell.alignment = left_alignment
                            else:
                                cell.alignment = center_alignment
                    
                    # Linha de total (lado direito nas colunas E-G)
                    total_row = len(df_resumo) + 2
                    
                    total_fill_principal = PatternFill(start_color="FF275316", end_color="FF275316", fill_type="solid")
                    total_fill_horas = PatternFill(start_color="FFC0E6F5", end_color="FFC0E6F5", fill_type="solid")
                    total_font_principal = Font(bold=True, size=11, color="FFFFFFFF", name="Calibri")
                    total_font_horas = Font(bold=True, size=11, color="FF000000", name="Calibri")
                    
                    # Coluna E: Total Geral (label)
                    ws1.cell(row=total_row, column=5, value="Total Geral")
                    ws1.cell(row=total_row, column=5).fill = total_fill_principal
                    ws1.cell(row=total_row, column=5).font = total_font_principal
                    ws1.cell(row=total_row, column=5).alignment = left_alignment
                    ws1.cell(row=total_row, column=5).border = border
                    
                    # Coluna F: Horas Positivas Totais
                    ws1.cell(row=total_row, column=6, value=horas_para_tempo(total_positivo))
                    ws1.cell(row=total_row, column=6).fill = total_fill_horas
                    ws1.cell(row=total_row, column=6).font = total_font_horas
                    ws1.cell(row=total_row, column=6).alignment = center_alignment
                    ws1.cell(row=total_row, column=6).border = border
                    
                    # Coluna G: Horas Negativas Totais
                    ws1.cell(row=total_row, column=7, value=horas_para_tempo(total_negativo))
                    ws1.cell(row=total_row, column=7).fill = total_fill_horas
                    ws1.cell(row=total_row, column=7).font = total_font_horas
                    ws1.cell(row=total_row, column=7).alignment = center_alignment
                    ws1.cell(row=total_row, column=7).border = border
                    
                    ws1.column_dimensions['B'].width = 42
                    ws1.column_dimensions['C'].width = 18
                    ws1.column_dimensions['E'].width = 50
                    ws1.column_dimensions['F'].width = 18
                    ws1.column_dimensions['G'].width = 18
                    
                    # Remove grid lines da sheet CONSOLIDAÇÃO
                    ws1.sheet_view.showGridLines = False
                    
                    # Carrega dados do CSV para lookup de gestores (tenta múltiplos encodings)
                    df_gestores = None
                    try:
                        # Tenta diferentes encodings com delimiter correto (;)
                        encodings = ['latin-1', 'iso-8859-1', 'cp1252', 'utf-8']
                        for encoding in encodings:
                            try:
                                df_gestores = pd.read_csv(
                                    file_csv_colaboradores, 
                                    encoding=encoding,
                                    sep=';',  # Delimiter correto
                                    skiprows=1  # Pula a primeira linha "Colaboradores"
                                )
                                break
                            except Exception:
                                continue
                        
                        if df_gestores is None:
                            st.warning("⚠️ Não foi possível carregar o CSV com os encodings disponíveis")
                            df_gestores = None
                    except Exception as e:
                        st.warning(f"⚠️ Erro ao carregar CSV: {e}")
                        df_gestores = None
                    
                    # Função para fazer lookup do gestor (PROCV)
                    def buscar_gestor(nome_colaborador, df_csv):
                        """
                        Busca o gestor de um colaborador no DataFrame CSV
                        Retorna o nome do gestor ou 'N/A' se não encontrar
                        """
                        if df_csv is None or df_csv.empty:
                            return "N/A"
                        
                        try:
                            # Limpa o nome para comparação
                            nome_limpo = str(nome_colaborador).strip().upper()
                            
                            # Procura na coluna 'Colaborador'
                            if 'Colaborador' in df_csv.columns:
                                linha = df_csv[df_csv['Colaborador'].astype(str).str.strip().str.upper() == nome_limpo]
                                if not linha.empty and 'Nome Gestor' in df_csv.columns:
                                    return str(linha.iloc[0]['Nome Gestor']).strip()
                        except Exception as e:
                            pass
                        
                        return "N/A"
                    
                    # ===== SHEET 2: OFENSORES (Positivos + Negativos) =====
                    ws2 = wb.create_sheet("Ofensores")
                    
                    # Define estilos para OFENSORES
                    header_ofensores_fill = PatternFill(start_color="FF265216", end_color="FF265216", fill_type="solid")
                    header_ofensores_font = Font(bold=True, color="FFFFFFFF", size=11, name="Calibri")
                    
                    data_fill = PatternFill(start_color="FFDBF2D0", end_color="FFDBF2D0", fill_type="solid")
                    data_font = Font(color="FF000000", name="Calibri", size=11)
                    
                    status_pos_fill = PatternFill(start_color="FF8ED973", end_color="FF8ED973", fill_type="solid")
                    status_pos_font = Font(bold=True, color="FFFFFFFF", name="Calibri", size=11)
                    
                    status_neg_fill = PatternFill(start_color="FFFF0101", end_color="FFFF0101", fill_type="solid")
                    status_neg_font = Font(bold=True, color="FFFFFFFF", name="Calibri", size=11)
                    
                    # Cores para coluna SALDO ATUAL por posição (1-15)
                    saldo_colors = {
                        1: "FFF8696B",
                        2: "FFFCA477",
                        3: "FFFCB37A",
                        4: "FFFDC07C",
                        5: "FFFED17F",
                        6: "FFFFE483",
                        7: "FFFFEB84",
                        8: "FFFEEB85",
                        9: "FFFCEA83",
                        10: "FFF1E783",
                        11: "FFE5E382",
                        12: "FFB1D47F",
                        13: "FF8CCA7D",
                        14: "FF71C37A",
                        15: "FF62BF7B"
                    }
                    
                    border_normal = Border(
                        left=Side(style="thin", color="000000"),
                        right=Side(style="thin", color="000000"),
                        top=Side(style="thin", color="000000"),
                        bottom=Side(style="thin", color="000000")
                    )
                    
                    # Headers das colunas
                    headers_ofensores = ['FUNCIONÁRIO', 'SETOR', 'SALDO ATUAL', 'STATUS', 'GESTOR']
                    
                    # ===== TOP 15 POSITIVOS =====
                    row_idx = 1
                    
                    # Headers POSITIVOS
                    for col_idx, header in enumerate(headers_ofensores, 1):
                        cell = ws2.cell(row=row_idx, column=col_idx, value=header)
                        cell.fill = header_ofensores_fill
                        cell.font = header_ofensores_font
                        cell.alignment = center_alignment
                        cell.border = border_normal
                    
                    ws2.row_dimensions[row_idx].height = 20
                    row_idx += 1
                    
                    # Dados POSITIVOS
                    for idx, (_, row) in enumerate(df_top15_pos.iterrows(), 1):
                        nome_colab = row['Colaborador']
                        ws2.cell(row=row_idx, column=1, value=nome_colab)
                        ws2.cell(row=row_idx, column=2, value=row['CentroDeCustos'])
                        ws2.cell(row=row_idx, column=3, value=row['POSITIVO'])
                        ws2.cell(row=row_idx, column=4, value="POSITIVO")
                        ws2.cell(row=row_idx, column=5, value=buscar_gestor(nome_colab, df_gestores))
                        
                        for col in range(1, 6):
                            cell = ws2.cell(row=row_idx, column=col)
                            cell.border = border_normal
                            cell.font = data_font
                            
                            if col == 4:  # STATUS
                                cell.fill = status_pos_fill
                                cell.font = status_pos_font
                            elif col == 3:  # SALDO ATUAL com cor por posição
                                color_key = idx if idx in saldo_colors else 15
                                cell.fill = PatternFill(start_color=saldo_colors[color_key], end_color=saldo_colors[color_key], fill_type="solid")
                            else:
                                cell.fill = data_fill
                            
                            if col == 3 or col == 4:
                                cell.alignment = center_alignment
                            else:
                                cell.alignment = left_alignment
                        
                        row_idx += 1
                    
                    # Linha de intervalo (quebra de linha)
                    row_idx += 1
                    
                    # Headers NEGATIVOS
                    for col_idx, header in enumerate(headers_ofensores, 1):
                        cell = ws2.cell(row=row_idx, column=col_idx, value=header)
                        cell.fill = header_ofensores_fill
                        cell.font = header_ofensores_font
                        cell.alignment = center_alignment
                        cell.border = border_normal
                    
                    ws2.row_dimensions[row_idx].height = 20
                    row_idx += 1
                    
                    # Dados NEGATIVOS
                    for idx, (_, row) in enumerate(df_top15_neg.iterrows(), 1):
                        nome_colab = row['Colaborador']
                        ws2.cell(row=row_idx, column=1, value=nome_colab)
                        ws2.cell(row=row_idx, column=2, value=row['CentroDeCustos'])
                        ws2.cell(row=row_idx, column=3, value=row['NEGATIVO'])
                        ws2.cell(row=row_idx, column=4, value="NEGATIVO")
                        ws2.cell(row=row_idx, column=5, value=buscar_gestor(nome_colab, df_gestores))
                        
                        for col in range(1, 6):
                            cell = ws2.cell(row=row_idx, column=col)
                            cell.border = border_normal
                            cell.font = data_font
                            
                            if col == 4:  # STATUS
                                cell.fill = status_neg_fill
                                cell.font = status_neg_font
                            elif col == 3:  # SALDO ATUAL com cor por posição
                                color_key = idx if idx in saldo_colors else 15
                                cell.fill = PatternFill(start_color=saldo_colors[color_key], end_color=saldo_colors[color_key], fill_type="solid")
                            else:
                                cell.fill = data_fill
                            
                            if col == 3 or col == 4:
                                cell.alignment = center_alignment
                            else:
                                cell.alignment = left_alignment
                        
                        row_idx += 1
                    
                    # Remove grid lines da sheet OFENSORES
                    ws2.sheet_view.showGridLines = False
                    
                    ws2.column_dimensions['A'].width = 42
                    ws2.column_dimensions['B'].width = 45
                    ws2.column_dimensions['C'].width = 13
                    ws2.column_dimensions['D'].width = 13
                    ws2.column_dimensions['E'].width = 45
                    
                    # ===== SHEET 3 E 4: DETALHES MOVIMENTACAO E ARMAZENAGEM (POSITIVOS E NEGATIVOS) =====
                    # Filtra apenas MOVIMENTACAO E ARMAZENAGEM
                    df_mov_arm = df_processado[
                        df_processado['CentroDeCustos'].str.strip().str.upper() == 'MOVIMENTACAO E ARMAZENAGEM'
                    ].copy()
                    
                    if len(df_mov_arm) > 0:
                        # Define estilos para detalhes
                        header_detail_fill = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
                        header_detail_font = Font(bold=True, color="FFFFFFFF", size=11, name="Calibri")
                        
                        detail_data_fill = PatternFill(start_color="FFE7E6E6", end_color="FFE7E6E6", fill_type="solid")
                        detail_data_font = Font(color="FF000000", name="Calibri", size=10)
                        
                        detail_total_fill = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
                        detail_total_font = Font(bold=True, color="FFFFFFFF", name="Calibri", size=11)
                        
                        border_detail = Border(
                            left=Side(style="thin", color="000000"),
                            right=Side(style="thin", color="000000"),
                            top=Side(style="thin", color="000000"),
                            bottom=Side(style="thin", color="000000")
                        )
                        

                        
                    # ===== SHEETS DE OFENSORES POR TURNO =====
                    # Função auxiliar para buscar turno de um colaborador
                    def buscar_turno_colaborador(nome_colaborador, df_csv):
                        """
                        Busca o turno de um colaborador no DataFrame CSV
                        Retorna o turno (TURNO 1, 2, 3 ou "Indeterminado")
                        """
                        if df_csv is None or df_csv.empty:
                            return "Indeterminado"
                        
                        try:
                            nome_limpo = str(nome_colaborador).strip().upper()
                            
                            # Procura na coluna 'Colaborador'
                            if 'Colaborador' in df_csv.columns:
                                linha = df_csv[df_csv['Colaborador'].astype(str).str.strip().str.upper() == nome_limpo]
                                if not linha.empty and 'Jornada' in df_csv.columns:
                                    jornada = str(linha.iloc[0]['Jornada']).strip()
                                    turno = determinar_turno(jornada)
                                    return turno
                        except Exception:
                            pass
                        
                        return "Indeterminado"
                    
                    # ===== SHEETS DE OFENSORES POR TURNO =====
                    # Adiciona turno para TODOS os colaboradores (não apenas TOP 15)
                    df_com_todos_turnos = df_processado[['Colaborador', 'CentroDeCustos', 'POSITIVO_num', 'NEGATIVO_num']].copy()
                    df_com_todos_turnos['Turno'] = df_com_todos_turnos['Colaborador'].apply(
                        lambda x: buscar_turno_colaborador(x, df_gestores)
                    )
                    
                    # Cria sheets para cada turno
                    for num_turno in [1, 2, 3]:
                        turno_label = f"TURNO {num_turno}"
                        
                        # Filtra apenas colaboradores deste turno
                        df_turno_todos = df_com_todos_turnos[df_com_todos_turnos['Turno'] == turno_label].copy()
                        
                        # Pega TOP 15 positivos e negativos deste turno
                        df_turno_pos = df_turno_todos.nlargest(15, 'POSITIVO_num')[['Colaborador', 'CentroDeCustos', 'POSITIVO_num']].copy()
                        df_turno_pos['POSITIVO'] = df_turno_pos['POSITIVO_num'].apply(horas_para_tempo)
                        
                        df_turno_neg = df_turno_todos.nlargest(15, 'NEGATIVO_num')[['Colaborador', 'CentroDeCustos', 'NEGATIVO_num']].copy()
                        df_turno_neg['NEGATIVO'] = df_turno_neg['NEGATIVO_num'].apply(horas_para_tempo)
                        
                        # Cria sheet apenas se houver dados
                        if len(df_turno_pos) > 0 or len(df_turno_neg) > 0:
                            ws_turno = wb.create_sheet(f"Ofensores Turno {num_turno}")
                            
                            row_idx = 1
                            
                            # ===== POSITIVOS DO TURNO =====
                            if len(df_turno_pos) > 0:
                                # Headers POSITIVOS
                                for col_idx, header in enumerate(headers_ofensores, 1):
                                    cell = ws_turno.cell(row=row_idx, column=col_idx, value=header)
                                    cell.fill = header_ofensores_fill
                                    cell.font = header_ofensores_font
                                    cell.alignment = center_alignment
                                    cell.border = border_normal
                                
                                ws_turno.row_dimensions[row_idx].height = 20
                                row_idx += 1
                                
                                # Dados POSITIVOS
                                for idx, (_, row) in enumerate(df_turno_pos.iterrows(), 1):
                                    nome_colab = row['Colaborador']
                                    ws_turno.cell(row=row_idx, column=1, value=nome_colab)
                                    ws_turno.cell(row=row_idx, column=2, value=row['CentroDeCustos'])
                                    ws_turno.cell(row=row_idx, column=3, value=row['POSITIVO'])
                                    ws_turno.cell(row=row_idx, column=4, value="POSITIVO")
                                    ws_turno.cell(row=row_idx, column=5, value=buscar_gestor(nome_colab, df_gestores))
                                    
                                    for col in range(1, 6):
                                        cell = ws_turno.cell(row=row_idx, column=col)
                                        cell.border = border_normal
                                        cell.font = data_font
                                        
                                        if col == 4:  # STATUS
                                            cell.fill = status_pos_fill
                                            cell.font = status_pos_font
                                        elif col == 3:  # SALDO ATUAL com cor por posição
                                            color_key = idx if idx in saldo_colors else 15
                                            cell.fill = PatternFill(start_color=saldo_colors[color_key], end_color=saldo_colors[color_key], fill_type="solid")
                                        else:
                                            cell.fill = data_fill
                                        
                                        if col == 3 or col == 4:
                                            cell.alignment = center_alignment
                                        else:
                                            cell.alignment = left_alignment
                                    
                                    row_idx += 1
                            
                            # Blank line entre seções
                            row_idx += 1
                            
                            # ===== NEGATIVOS DO TURNO =====
                            if len(df_turno_neg) > 0:
                                # Headers NEGATIVOS
                                for col_idx, header in enumerate(headers_ofensores, 1):
                                    cell = ws_turno.cell(row=row_idx, column=col_idx, value=header)
                                    cell.fill = header_ofensores_fill
                                    cell.font = header_ofensores_font
                                    cell.alignment = center_alignment
                                    cell.border = border_normal
                                
                                ws_turno.row_dimensions[row_idx].height = 20
                                row_idx += 1
                                
                                # Dados NEGATIVOS
                                for idx, (_, row) in enumerate(df_turno_neg.iterrows(), 1):
                                    nome_colab = row['Colaborador']
                                    ws_turno.cell(row=row_idx, column=1, value=nome_colab)
                                    ws_turno.cell(row=row_idx, column=2, value=row['CentroDeCustos'])
                                    ws_turno.cell(row=row_idx, column=3, value=row['NEGATIVO'])
                                    ws_turno.cell(row=row_idx, column=4, value="NEGATIVO")
                                    ws_turno.cell(row=row_idx, column=5, value=buscar_gestor(nome_colab, df_gestores))
                                    
                                    for col in range(1, 6):
                                        cell = ws_turno.cell(row=row_idx, column=col)
                                        cell.border = border_normal
                                        cell.font = data_font
                                        
                                        if col == 4:  # STATUS
                                            cell.fill = status_neg_fill
                                            cell.font = status_neg_font
                                        elif col == 3:  # SALDO ATUAL com cor por posição
                                            color_key = idx if idx in saldo_colors else 15
                                            cell.fill = PatternFill(start_color=saldo_colors[color_key], end_color=saldo_colors[color_key], fill_type="solid")
                                        else:
                                            cell.fill = data_fill
                                        
                                        if col == 3 or col == 4:
                                            cell.alignment = center_alignment
                                        else:
                                            cell.alignment = left_alignment
                                    
                                    row_idx += 1
                            
                            # Remove grid lines
                            ws_turno.sheet_view.showGridLines = False
                            
                            ws_turno.column_dimensions['A'].width = 42
                            ws_turno.column_dimensions['B'].width = 45
                            ws_turno.column_dimensions['C'].width = 13
                            ws_turno.column_dimensions['D'].width = 13
                            ws_turno.column_dimensions['E'].width = 45
                    
                    # ===== SHEET BASE =====
                    # Cria uma view consolidada de todos os colaboradores com status
                    df_base = df_processado.copy()
                    
                    # Determina o status (POSITIVO ou NEGATIVO) e o saldo
                    # Status sempre será POSITIVO (para Pagamentos), NEGATIVO (para Descontos) ou ZERO
                    df_base['STATUS'] = df_base.apply(
                        lambda row: 'POSITIVO' if row['POSITIVO_num'] > 0 else ('NEGATIVO' if row['NEGATIVO_num'] > 0 else 'ZERO'),
                        axis=1
                    )
                    
                    df_base['SALDO'] = df_base.apply(
                        lambda row: horas_para_tempo(row['POSITIVO_num']) if row['POSITIVO_num'] > 0 else horas_para_tempo(row['NEGATIVO_num']),
                        axis=1
                    )
                    
                    ws_base = wb.create_sheet("BASE")
                    
                    # Headers
                    headers_base = ['COLABORADOR', 'SALDO', 'STATUS', 'CENTRO DE CUSTOS']
                    for col_idx, header in enumerate(headers_base, 1):
                        cell = ws_base.cell(row=1, column=col_idx, value=header)
                        cell.fill = header_detail_fill
                        cell.font = header_detail_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border_detail
                    
                    ws_base.row_dimensions[1].height = 20
                    
                    # Ordena por colaborador
                    df_base_sorted = df_base.sort_values('Colaborador').reset_index(drop=True)
                    
                    # Dados
                    row_idx = 2
                    for _, row in df_base_sorted.iterrows():
                        ws_base.cell(row=row_idx, column=1, value=row['Colaborador'])
                        ws_base.cell(row=row_idx, column=2, value=row['SALDO'])
                        ws_base.cell(row=row_idx, column=3, value=row['STATUS'])
                        ws_base.cell(row=row_idx, column=4, value=row['CentroDeCustos'])
                        
                        for col in range(1, 5):
                            cell = ws_base.cell(row=row_idx, column=col)
                            cell.fill = detail_data_fill
                            cell.font = detail_data_font
                            cell.border = border_detail
                            
                            if col == 1 or col == 4:
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                            else:
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        row_idx += 1
                    
                    # Formata colunas
                    ws_base.column_dimensions['A'].width = 50
                    ws_base.column_dimensions['B'].width = 20
                    ws_base.column_dimensions['C'].width = 15
                    ws_base.column_dimensions['D'].width = 45
                    
                    # Remove grid lines
                    ws_base.sheet_view.showGridLines = False
                    
                    # Salva em memória
                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)
                    
                    progress_bar.progress(100, text="✅ Concluído! (100%)")
                    status_text.text("Relatório pronto para download!")
                    
                    st.divider()
                    
                    # Botão de download
                    st.download_button(
                        label="⬇️ Baixar Relatório (XLSX)",
                        data=output.getvalue(),
                        file_name="Banco de Horas.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            
            except Exception as e:
                st.error(f"❌ Erro ao gerar relatório: {str(e)}")
                import traceback
                st.write(traceback.format_exc())
    
    except Exception as e:
        st.error(f"❌ Erro ao carregar arquivo: {str(e)}")
        st.write(f"Detalhes: {e}")

else:
    st.info("""
    ### 📌 Como usar:
    
    1. **Upload**: Carregue o arquivo XLSX com banco de horas
    2. **Estrutura esperada**:
       - Coluna E: `CentroDeCustos` (identificação do centro de custo)
       - Coluna R: `SaldoFinal` (horas com sinal + positivo ou - negativo)
    3. **Gerar**: Clique em "Gerar Relatório"
    4. **Download**: Baixe o arquivo XLSX com o resumo consolidado
    
    ✅ O sistema detecta automaticamente valores negativos (com - ou parenteses)
    """)

