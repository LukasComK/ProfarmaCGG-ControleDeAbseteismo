"""
Página: Validação de ABS vs PONTO
Descrição: Compara a planilha mestra de absenteísmo (gerada pelos encarregados)
com a planilha de PONTO (verdade absoluta) e gera:
1. Planilha MESTRA CORRIGIDA baseada no PONTO (com formatação idêntica à do Controle_de_Absenteismo.py)
2. Relatório de Divergências para apresentar aos encarregados
Filtro: Apenas cargos AUXILIAR DEPOSITO I, II, III
Tratamento: Fins de semana (sábado sem 6x1 = D, domingo = D)
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, date, timedelta
import io
import re
import zipfile
from typing import Dict, List, Tuple
from unidecode import unidecode
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

st.set_page_config(page_title="Validação ABS vs PONTO", layout="wide")

st.title("✅ Validação de Absenteísmo vs Ponto")
st.markdown("Compara a planilha MESTRA de ABS (gerada pelos encarregados) com a(s) planilha(s) de PONTO (verdade absoluta)")
st.info("🔍 **Filtro ativo:** Apenas colaboradores com cargo **AUXILIAR DEPOSITO I, II ou III**")

# ===== SESSÃO PARA ARMAZENAR DADOS CONSOLIDADOS =====
if 'marcacoes_ponto_consolidadas' not in st.session_state:
    st.session_state.marcacoes_ponto_consolidadas = {}
if 'total_pontos_processados' not in st.session_state:
    st.session_state.total_pontos_processados = 0
if 'ultimos_resultados' not in st.session_state:
    st.session_state.ultimos_resultados = None

# ============================================================
# DICIONÁRIO DE MAPEAMENTO: OCORRÊNCIA+JUSTIFICATIVA DO PONTO -> MARCAÇÃO NA MESTRA
# ============================================================
JUSTIFICATIVAS_ABONADAS = [
    'OBITO DE FAMILIAR', 'OBITO', 'FALECIMENTO',
    'AMAMENTACAO', 'AMAMENTAÇÃO',
    'SERVICO EXTERNO', 'SERVIÇO EXTERNO',
    'FOLGA', 'FOLGA OURO DA CASA',
    'ANIVERSARIO - DIA LIVRE', 'ANIVERSÁRIO - DIA LIVRE',
    'INTEGRACAO', 'INTEGRAÇÃO',
    'CURSO DE APRENDIZAGEM',
    'LIBERACAO DA EMPRESA - DIA', 'LIBERAÇÃO DA EMPRESA - DIA',
    'BANCO DE HORAS', 'DECLARACAO DE HORAS', 'DECLARAÇÃO DE HORAS',
    'PARTE OU TESTEMUNHA DE PROCESSO JUDICIAL',
    'LIBERACAO EMPRESA - HORAS', 'LIBERAÇÃO EMPRESA - HORAS',
]

OCORRENCIAS_FA_15 = [
    'AFAST DOENCA <= 15 DIAS', 'AFAST DOENÇA <= 15 DIAS',
    'AFAST ACID TRAB <= 15 DIAS',
    'OUTROS TIPOS DE AFASTAMENTO',
]

OCORRENCIAS_IGNORAR = [
    'AFAST DOENCA > 15 DIAS', 'AFAST DOENÇA > 15 DIAS',
    'AFAST ACID TRAB > 15 DIAS',
]

OCORRENCIAS_FERIAS = [
    'FERIAS NORMAIS', 'FÉRIAS NORMAIS',
    'AFAST LICENCA MATERNIDADE', 'AFAST LICENÇA MATERNIDADE',
]

OCORRENCIAS_SEM_MARCACAO_ENTRADA = [
    'SEM MARCAÇÃO DE ENTRADA', 'SEM MARCACAO DE ENTRADA',
]
OCORRENCIAS_SEM_MARCACAO_SAIDA = [
    'SEM MARCAÇÃO DE SAÍDA', 'SEM MARCACAO DE SAIDA',
]

OCORRENCIAS_ATRASO = [
    'ENTRADA EM ATRASO',
]

# Cargos permitidos (apenas Auxiliar Deposito)
CARGOS_PERMITIDOS = [
    'AUXILIAR DEPOSITO I', 'AUXILIAR DEPOSITO II', 'AUXILIAR DEPOSITO III',
    'AUXILIAR DE DEPOSITO I', 'AUXILIAR DE DEPOSITO II', 'AUXILIAR DE DEPOSITO III',
]

# Cores do MAPA_CORES (mesmo do Controle_de_Absenteismo.py)
MAPA_CORES = {
    'P': 'FF90EE90',      # Verde claro
    'FI': 'FFFF0000',     # Vermelho puro
    'FA': 'FFFFFF00',     # Amarelo puro
    'Afastamento': 'FFC0C0C0',  # Cinza
    'FERIADO': 'FF000000',      # Preto (com texto branco)
    'FÉRIAS-BH': 'FF000000',    # Preto (com texto branco)
    'DESLIGADO': 'FF800080',   # Roxo
    'DESCANSO': 'FFC0C0C0',  # Cinza
    'FERIAS-BH': 'FF000000',   # Preto (com texto branco) - sem acento
}


def normalizar(texto) -> str:
    if pd.isna(texto):
        return ''
    return unidecode(str(texto)).upper().strip()


def limpar_nome(nome) -> str:
    if pd.isna(nome) or str(nome).strip().upper() == "NAN":
        return ""
    nome_limpo = str(nome).replace('"', '').replace("'", "")
    return " ".join(nome_limpo.strip().upper().split())


def cargo_permitido(cargo) -> bool:
    """Verifica se o cargo é um dos permitidos (AUXILIAR DEPOSITO I, II, III)"""
    cargo_norm = normalizar(str(cargo)) if pd.notna(cargo) else ''
    for cargo_permitido in CARGOS_PERMITIDOS:
        if cargo_permitido in cargo_norm or cargo_norm in cargo_permitido:
            return True
    return False


def jornada_trabalha_sabado(jornada) -> bool:
    """
    Verifica se a jornada de trabalho do colaborador inclui trabalho aos sábados.
    Retorna True se contiver '6x1' ou similar que indique trabalho no sábado.
    """
    if pd.isna(jornada) or not str(jornada).strip():
        return False
    jornada_str = normalizar(str(jornada))
    # "6X1" é o principal indicador de trabalho aos sábados
    if '6X1' in jornada_str or '6X2' in jornada_str:
        return True
    return False


def determinar_marcacao_por_ponto(ocorrencia: str, justificativa: str) -> str:
    occ_norm = normalizar(ocorrencia)
    just_norm = normalizar(justificativa)
    
    for occ_ignorar in OCORRENCIAS_IGNORAR:
        if occ_ignorar in occ_norm or occ_norm in occ_ignorar:
            return None
    
    for occ_fa in OCORRENCIAS_FA_15:
        if occ_fa in occ_norm or occ_norm in occ_fa:
            return 'FA'
    
    for occ_ferias in OCORRENCIAS_FERIAS:
        if occ_ferias in occ_norm or occ_norm in occ_ferias:
            return 'FERIAS-BH'
    
    for occ_sem_entrada in OCORRENCIAS_SEM_MARCACAO_ENTRADA:
        if occ_sem_entrada in occ_norm or occ_norm in occ_sem_entrada:
            return 'FI'
    
    for occ_sem_saida in OCORRENCIAS_SEM_MARCACAO_SAIDA:
        if occ_sem_saida in occ_norm or occ_norm in occ_sem_saida:
            return 'P'
    
    for occ_atraso in OCORRENCIAS_ATRASO:
        if occ_atraso in occ_norm or occ_norm in occ_atraso:
            return 'P'
    
    if 'FALTA' in occ_norm:
        for just_abonada in JUSTIFICATIVAS_ABONADAS:
            if just_abonada in just_norm or just_norm in just_abonada:
                return 'FERIAS-BH'
        return 'FI'
    
    return 'P'


def processar_ponto_para_marcacoes(df_ponto, ano: int, mes: int) -> Dict[Tuple[str, str], str]:
    """
    Processa a planilha de PONTO e retorna um dicionário:
    {(nome_colaborador, data_str): marcacao}
    
    Considera fins de semana:
    - Domingo (weekday=6): sempre 'D'
    - Sábado (weekday=5): depende da jornada (6x1 = trabalha, senão = 'D')
    """
    col_nome = df_ponto.columns[3]
    col_cargo = df_ponto.columns[7]
    col_escala = df_ponto.columns[11]  # Coluna L - Escala/Jornada (ex: "06:00 14:20 - 6x1")
    col_ocorrencia = df_ponto.columns[25]
    col_justificativa = df_ponto.columns[27]
    col_data = df_ponto.columns[38]
    
    marcacoes = {}
    
    for idx, row in df_ponto.iterrows():
        nome = limpar_nome(row[col_nome])
        if not nome:
            continue
        
        # Filtra apenas cargos permitidos
        cargo = row[col_cargo] if pd.notna(row[col_cargo]) else ''
        if not cargo_permitido(cargo):
            continue
        
        ocorrencia = str(row[col_ocorrencia]) if pd.notna(row[col_ocorrencia]) else ''
        justificativa = str(row[col_justificativa]) if pd.notna(row[col_justificativa]) else ''
        data_raw = row[col_data]
        
        # Converte data para string DD/MM e também obtém o dia da semana
        data_str = ''
        dia_semana = None  # 0=segunda, 6=domingo
        if pd.notna(data_raw):
            try:
                if isinstance(data_raw, (datetime, pd.Timestamp)):
                    data_dt = data_raw
                else:
                    data_dt = pd.to_datetime(str(data_raw), dayfirst=True, errors='coerce')
                
                if pd.notna(data_dt):
                    data_str = data_dt.strftime('%d/%m')
                    dia_semana = data_dt.weekday()  # 0=segunda, 6=domingo
            except:
                pass
        
        if not data_str:
            continue
        
        # ===== REGRA DE FIM DE SEMANA =====
        if dia_semana is not None:
            if dia_semana == 6:  # Domingo: sempre é descanso
                marcacoes[(nome, data_str)] = 'D'
                continue
            elif dia_semana == 5:  # Sábado: verifica se a jornada tem 6x1
                escala = str(row[col_escala]) if pd.notna(row[col_escala]) else ''
                if not jornada_trabalha_sabado(escala):
                    # Não trabalha sábado -> descanso
                    marcacoes[(nome, data_str)] = 'D'
                    continue
                # Se trabalha sábado (6x1), processa normal
        
        # ===== PROCESSAMENTO NORMAL (dias de semana ou sábado com 6x1) =====
        marcacao = determinar_marcacao_por_ponto(ocorrencia, justificativa)
        
        if marcacao is not None:
            chave = (nome, data_str)
            if chave in marcacoes:
                marcacao_existente = marcacoes[chave]
                prioridade = {'FI': 4, 'FA': 3, 'FERIAS-BH': 2, 'P': 1, 'D': 0}
                if prioridade.get(marcacao, 0) > prioridade.get(marcacao_existente, 0):
                    marcacoes[chave] = marcacao
            else:
                marcacoes[chave] = marcacao
    
    return marcacoes


def extrair_datas_da_mestra(df_mestra) -> List[str]:
    datas = []
    for col in df_mestra.columns:
        col_str = str(col).strip()
        try:
            if isinstance(col, (datetime, pd.Timestamp)):
                datas.append(col.strftime('%d/%m'))
                continue
            match = re.match(r'^(\d{1,2})/(\d{1,2})$', col_str)
            if match:
                dia, mes = int(match.group(1)), int(match.group(2))
                if 1 <= dia <= 31 and 1 <= mes <= 12:
                    datas.append(col_str)
                    continue
            match = re.match(r'^(\d{1,2})/([a-zA-Z]{3})$', col_str)
            if match:
                datas.append(col_str)
                continue
        except:
            pass
    return datas


def comparar_mestra_com_ponto(
    df_mestra: pd.DataFrame,
    marcacoes_ponto: Dict[Tuple[str, str], str],
    datas_mestra: List[str]
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df_mestra['NOME_LIMPO'] = df_mestra['NOME'].apply(limpar_nome)
    nomes_mestra = set(df_mestra['NOME_LIMPO'].dropna().unique())
    
    df_corrigida = df_mestra.copy()
    divergencias = []
    nao_encontrados_ponto = []
    
    for (nome_ponto, data_str), marcacao_correta in marcacoes_ponto.items():
        if nome_ponto in nomes_mestra:
            idx_mestra = df_mestra[df_mestra['NOME_LIMPO'] == nome_ponto].index[0]
            
            col_data_encontrada = None
            for col in datas_mestra:
                col_norm = normalizar(col)
                data_norm = normalizar(data_str)
                if col_norm == data_norm or col_norm.endswith(data_norm) or data_norm.endswith(col_norm):
                    col_data_encontrada = col
                    break
            
            if col_data_encontrada and col_data_encontrada in df_mestra.columns:
                valor_atual = str(df_mestra.at[idx_mestra, col_data_encontrada]) if pd.notna(df_mestra.at[idx_mestra, col_data_encontrada]) else ''
                valor_atual = valor_atual.strip().upper()
                
                # ===== NORMALIZAÇÃO: FÉRIAS-BH pode vir com ou sem acento =====
                if valor_atual == 'FÉRIAS-BH':
                    valor_atual = 'FERIAS-BH'
                comparavel_marcacao = marcacao_correta
                if comparavel_marcacao == 'FÉRIAS-BH':
                    comparavel_marcacao = 'FERIAS-BH'
                
                if valor_atual != comparavel_marcacao:
                    divergencias.append({
                        'NOME': df_mestra.at[idx_mestra, 'NOME'],
                        'DATA': data_str,
                        'VALOR_NA_MESTRA': valor_atual if valor_atual else '(vazio)',
                        'VALOR_CORRETO_PONTO': marcacao_correta,
                        'STATUS': 'DIVERGENTE'
                    })
                    df_corrigida.at[idx_mestra, col_data_encontrada] = marcacao_correta
        else:
            nao_encontrados_ponto.append({
                'NOME': nome_ponto,
                'DATA': data_str,
                'MARCACAO_CORRETA': marcacao_correta
            })
    
    df_divergencias = pd.DataFrame(divergencias)
    df_nao_encontrados = pd.DataFrame(nao_encontrados_ponto)
    
    if 'NOME_LIMPO' in df_corrigida.columns:
        df_corrigida = df_corrigida.drop(columns=['NOME_LIMPO'])
    
    return df_corrigida, df_divergencias, df_nao_encontrados


def gerar_mestra_corrigida_excel(df_corrigida: pd.DataFrame, datas_mestra: List[str]) -> bytes:
    """
    Gera a planilha mestra corrigida com a MESMA FORMATAÇÃO do Controle_de_Absenteismo.py
    """
    out = io.BytesIO()
    
    with pd.ExcelWriter(out, engine='openpyxl') as w:
        df_corrigida.to_excel(w, index=False, sheet_name='Dados')
        worksheet = w.sheets['Dados']
        
        # ===== FORMATAÇÃO DO HEADER (igual ao Controle_de_Absenteismo.py) =====
        header_fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFFFF', size=11)
        
        for col_idx in range(1, len(df_corrigida.columns) + 1):
            header_cell = worksheet.cell(row=1, column=col_idx)
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # ===== FORMATAÇÃO DAS COLUNAS ESPECÍFICAS =====
        col_names = df_corrigida.columns.tolist()
        
        def calc_width(df, col_name, min_width=10, max_width=50):
            if col_name not in df.columns:
                return min_width
            max_len = df[col_name].astype(str).str.len().max()
            header_len = len(str(col_name))
            largest = max(max_len, header_len)
            width = min(max(largest + 2, min_width), max_width)
            return width
        
        for col_idx, col_name in enumerate(col_names, 1):
            if col_name == 'NOME':
                col_fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                width = calc_width(df_corrigida, col_name, min_width=15, max_width=40)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = width
            elif col_name == 'AREA':
                col_fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 25
            elif col_name == 'GESTOR':
                col_fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                width = calc_width(df_corrigida, col_name, min_width=15, max_width=40)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = width
            else:
                col_fill = None
                try:
                    datetime.strptime(str(col_name), '%d/%m')
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = 7
                except:
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = 10
            
            if col_fill is not None:
                for row_idx in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = col_fill
        
        # ===== FORMATAÇÃO DAS CÉLULAS DE DATA (cores iguais ao Controle_de_Absenteismo.py) =====
        for col_data_nome in datas_mestra:
            if col_data_nome not in df_corrigida.columns:
                continue
            col_idx = list(df_corrigida.columns).index(col_data_nome) + 1
            
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                valor = str(cell.value).strip() if cell.value else ''
                
                if valor == 'P':
                    cell.fill = PatternFill(start_color=MAPA_CORES['P'], end_color=MAPA_CORES['P'], fill_type='solid')
                elif valor == 'FI':
                    cell.fill = PatternFill(start_color=MAPA_CORES['FI'], end_color=MAPA_CORES['FI'], fill_type='solid')
                    cell.font = Font(color='FFFFFFFF')
                elif valor == 'FA':
                    cell.fill = PatternFill(start_color=MAPA_CORES['FA'], end_color=MAPA_CORES['FA'], fill_type='solid')
                elif valor in ['FÉRIAS-BH', 'FERIAS-BH']:
                    cell.fill = PatternFill(start_color=MAPA_CORES['FÉRIAS-BH'], end_color=MAPA_CORES['FÉRIAS-BH'], fill_type='solid')
                    cell.font = Font(color='FFFFFFFF')
                elif valor == 'DESLIGADO':
                    cell.fill = PatternFill(start_color=MAPA_CORES['DESLIGADO'], end_color=MAPA_CORES['DESLIGADO'], fill_type='solid')
                    cell.font = Font(color='FFFFFFFF')
                elif valor == 'D':
                    cell.fill = PatternFill(start_color=MAPA_CORES['DESCANSO'], end_color=MAPA_CORES['DESCANSO'], fill_type='solid')
    
    out.seek(0)
    return out.getvalue()


def gerar_relatorio_divergencias_excel(
    df_divergencias: pd.DataFrame,
    df_nao_encontrados: pd.DataFrame,
    df_mestra_original: pd.DataFrame
) -> bytes:
    wb = Workbook()
    
    verde_escuro = 'FF0D4F45'
    branco = 'FFFFFFFF'
    
    header_fill = PatternFill(start_color=verde_escuro, end_color=verde_escuro, fill_type='solid')
    header_font = Font(bold=True, color=branco, size=11)
    divergente_fill = PatternFill(start_color='FFFFE0E0', end_color='FFFFE0E0', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== ABA 1: DIVERGÊNCIAS =====
    ws_div = wb.active
    ws_div.title = 'Divergencias'
    
    ws_div.merge_cells('A1:F1')
    titulo = ws_div['A1']
    titulo.value = '🚨 RELATÓRIO DE DIVERGÊNCIAS - ABS vs PONTO'
    titulo.font = Font(bold=True, size=14, color=branco)
    titulo.fill = header_fill
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    
    if not df_divergencias.empty:
        headers = ['NOME', 'DATA', 'VALOR NA MESTRA (Encarregado)', 'VALOR CORRETO (Ponto)', 'STATUS']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_div.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        for row_idx, (_, row) in enumerate(df_divergencias.iterrows(), 4):
            ws_div.cell(row=row_idx, column=1, value=row['NOME']).border = thin_border
            ws_div.cell(row=row_idx, column=2, value=row['DATA']).border = thin_border
            ws_div.cell(row=row_idx, column=3, value=row['VALOR_NA_MESTRA']).border = thin_border
            ws_div.cell(row=row_idx, column=4, value=row['VALOR_CORRETO_PONTO']).border = thin_border
            cell_status = ws_div.cell(row=row_idx, column=5, value='❌ DIVERGENTE')
            cell_status.border = thin_border
            cell_status.fill = divergente_fill
            cell_status.font = Font(bold=True, color='FFCC0000')
            
            for col_idx in range(1, 6):
                ws_div.cell(row=row_idx, column=col_idx).fill = divergente_fill
        
        row_resumo = len(df_divergencias) + 5
        ws_div.cell(row=row_resumo, column=1, value='TOTAL DE DIVERGÊNCIAS:').font = Font(bold=True, size=12)
        ws_div.cell(row=row_resumo, column=2, value=len(df_divergencias)).font = Font(bold=True, size=12, color='FFCC0000')
        
        ws_div.column_dimensions['A'].width = 45
        ws_div.column_dimensions['B'].width = 12
        ws_div.column_dimensions['C'].width = 25
        ws_div.column_dimensions['D'].width = 25
        ws_div.column_dimensions['E'].width = 20
    else:
        ws_div.cell(row=3, column=1, value='✅ NENHUMA DIVERGÊNCIA ENCONTRADA!').font = Font(bold=True, size=14, color='FF00B050')
    
    # ===== ABA 2: NÃO ENCONTRADOS NA MESTRA =====
    ws_ne = wb.create_sheet('Nao Encontrados na Mestra')
    
    ws_ne.merge_cells('A1:C1')
    titulo_ne = ws_ne['A1']
    titulo_ne.value = '👥 COLABORADORES NO PONTO MAS NÃO NA MESTRA'
    titulo_ne.font = Font(bold=True, size=14, color=branco)
    titulo_ne.fill = header_fill
    titulo_ne.alignment = Alignment(horizontal='center', vertical='center')
    
    if not df_nao_encontrados.empty:
        headers_ne = ['NOME', 'DATA', 'MARCAÇÃO CORRETA']
        for col_idx, header in enumerate(headers_ne, 1):
            cell = ws_ne.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        for row_idx, (_, row) in enumerate(df_nao_encontrados.iterrows(), 4):
            ws_ne.cell(row=row_idx, column=1, value=row['NOME']).border = thin_border
            ws_ne.cell(row=row_idx, column=2, value=row['DATA']).border = thin_border
            ws_ne.cell(row=row_idx, column=3, value=row['MARCACAO_CORRETA']).border = thin_border
        
        ws_ne.column_dimensions['A'].width = 45
        ws_ne.column_dimensions['B'].width = 12
        ws_ne.column_dimensions['C'].width = 20
    else:
        ws_ne.cell(row=3, column=1, value='✅ Todos os colaboradores do ponto estão na planilha mestra!').font = Font(bold=True, size=12, color='FF00B050')
    
    # ===== ABA 3: RESUMO POR GESTOR =====
    ws_resumo = wb.create_sheet('Resumo por Gestor')
    
    ws_resumo.merge_cells('A1:D1')
    titulo_res = ws_resumo['A1']
    titulo_res.value = '📊 RESUMO DE DIVERGÊNCIAS POR GESTOR'
    titulo_res.font = Font(bold=True, size=14, color=branco)
    titulo_res.fill = header_fill
    titulo_res.alignment = Alignment(horizontal='center', vertical='center')
    
    if not df_divergencias.empty and 'GESTOR' in df_mestra_original.columns:
        df_mestra_original['NOME_LIMPO'] = df_mestra_original['NOME'].apply(limpar_nome)
        div_com_gestor = df_divergencias.copy()
        div_com_gestor['GESTOR'] = div_com_gestor['NOME'].apply(
            lambda x: df_mestra_original[df_mestra_original['NOME_LIMPO'] == limpar_nome(x)]['GESTOR'].values[0]
            if len(df_mestra_original[df_mestra_original['NOME_LIMPO'] == limpar_nome(x)]['GESTOR'].values) > 0
            else 'NÃO IDENTIFICADO'
        )
        
        resumo_gestor = div_com_gestor.groupby('GESTOR').size().reset_index(name='DIVERGÊNCIAS')
        resumo_gestor = resumo_gestor.sort_values('DIVERGÊNCIAS', ascending=False)
        
        headers_res = ['GESTOR', 'TOTAL DE DIVERGÊNCIAS', '% DO TOTAL']
        for col_idx, header in enumerate(headers_res, 1):
            cell = ws_resumo.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        total_div = len(df_divergencias)
        for row_idx, (_, row) in enumerate(resumo_gestor.iterrows(), 4):
            ws_resumo.cell(row=row_idx, column=1, value=row['GESTOR']).border = thin_border
            ws_resumo.cell(row=row_idx, column=2, value=int(row['DIVERGÊNCIAS'])).border = thin_border
            pct = (row['DIVERGÊNCIAS'] / total_div * 100) if total_div > 0 else 0
            cell_pct = ws_resumo.cell(row=row_idx, column=3, value=f"{pct:.1f}%")
            cell_pct.border = thin_border
            cell_pct.alignment = Alignment(horizontal='center')
        
        ws_resumo.column_dimensions['A'].width = 45
        ws_resumo.column_dimensions['B'].width = 25
        ws_resumo.column_dimensions['C'].width = 15
    else:
        if df_divergencias.empty:
            ws_resumo.cell(row=3, column=1, value='✅ Nenhuma divergência encontrada!').font = Font(bold=True, size=12, color='FF00B050')
        else:
            ws_resumo.cell(row=3, column=1, value='Coluna GESTOR não encontrada na planilha mestra original.').font = Font(bold=True, size=12)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================
# INTERFACE
# ============================================================

st.divider()
st.subheader("📤 Upload dos Arquivos")

col1, col2 = st.columns(2)

with col1:
    file_mestra = st.file_uploader(
        "Planilha MESTRA de ABS (gerada pelos encarregados)",
        type=["xlsx", "xlsm"],
        key="mestra_abs"
    )

with col2:
    files_ponto = st.file_uploader(
        "Planilha(s) de PONTO (verdade absoluta) - múltiplas permitidas",
        type=["xlsx", "xlsm"],
        accept_multiple_files=True,
        key="ponto_abs"
    )

# ===== CONFIGURAÇÃO (ANO/MÊS) - DEVE VIR ANTES DO PROCESSAMENTO =====
st.divider()
st.subheader("📋 Configuração")

col_ano, col_mes = st.columns(2)
with col_ano:
    ano = st.number_input("Ano", 2020, 2050, datetime.now().year, key="ano_validacao")
with col_mes:
    mes = st.number_input("Mês", 1, 12, datetime.now().month, key="mes_validacao")

# ===== PROCESSAR ARQUIVOS DE PONTO (ACUMULA) =====
if files_ponto:
    total_atual = len(files_ponto)
    
    with st.expander(f"📋 {total_atual} arquivo(s) de PONTO carregado(s)", expanded=False):
        for i, f in enumerate(files_ponto):
            st.write(f"  **[{i+1}]** - `{f.name}`")
    
    # Botão para processar/REPROCESSAR os pontos
    if st.button("🔄 Processar Arquivos de PONTO", type="secondary", use_container_width=True):
        st.session_state.marcacoes_ponto_consolidadas = {}
        st.session_state.total_pontos_processados = 0
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, file_ponto in enumerate(files_ponto):
            status_text.text(f"📖 Processando [{i+1}/{total_atual}]: {file_ponto.name}...")
            try:
                df_ponto = pd.read_excel(file_ponto)
                
                if len(df_ponto.columns) < 39:
                    st.warning(f"⚠️ {file_ponto.name} tem apenas {len(df_ponto.columns)} colunas (mínimo 39). Pulando...")
                    continue
                
                # Processa este arquivo de ponto
                marcacoes_arquivo = processar_ponto_para_marcacoes(df_ponto, ano, mes)
                
                # Faz merge com as marcações já existentes (prioridade: maior prioridade vence)
                for chave, marcacao in marcacoes_arquivo.items():
                    if chave in st.session_state.marcacoes_ponto_consolidadas:
                        marcacao_existente = st.session_state.marcacoes_ponto_consolidadas[chave]
                        prioridade = {'FI': 4, 'FA': 3, 'FERIAS-BH': 2, 'P': 1, 'D': 0}
                        if prioridade.get(marcacao, 0) > prioridade.get(marcacao_existente, 0):
                            st.session_state.marcacoes_ponto_consolidadas[chave] = marcacao
                    else:
                        st.session_state.marcacoes_ponto_consolidadas[chave] = marcacao
                
                st.session_state.total_pontos_processados += len(marcacoes_arquivo)
                
            except Exception as e:
                st.warning(f"⚠️ Erro ao processar {file_ponto.name}: {str(e)}")
            
            progress_bar.progress((i + 1) / total_atual)
        
        status_text.success(f"✅ Processados {total_atual} arquivo(s)! Total de {st.session_state.total_pontos_processados} marcações consolidadas.")
        progress_bar.progress(1.0)

if file_mestra and st.session_state.total_pontos_processados > 0:
    if st.button("🔍 Validar e Comparar", type="primary", use_container_width=True):
        with st.spinner("Comparando MESTRA vs PONTO consolidado..."):
            try:
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                status_text.text("📖 Carregando planilha MESTRA...")
                df_mestra = pd.read_excel(file_mestra)
                progress_bar.progress(20)
                
                marcacoes_ponto = st.session_state.marcacoes_ponto_consolidadas
                
                st.info(f"📊 Usando {len(marcacoes_ponto)} marcações consolidadas de {len(files_ponto)} arquivo(s) de PONTO")
                
                status_text.text("📅 Identificando colunas de data na MESTRA...")
                datas_mestra = extrair_datas_da_mestra(df_mestra)
                progress_bar.progress(40)
                
                st.info(f"📅 Encontradas {len(datas_mestra)} colunas de data na MESTRA")
                
                status_text.text("🔍 Comparando MESTRA vs PONTO...")
                df_corrigida, df_divergencias, df_nao_encontrados = comparar_mestra_com_ponto(
                    df_mestra, marcacoes_ponto, datas_mestra
                )
                progress_bar.progress(70)
                
                # ===== EXIBE RESULTADOS =====
                st.divider()
                st.subheader("📊 Resultados da Validação")
                
                col_res1, col_res2, col_res3 = st.columns(3)
                with col_res1:
                    st.metric("Total de Marcações no PONTO", len(marcacoes_ponto))
                with col_res2:
                    st.metric("Divergências Encontradas", len(df_divergencias),
                              delta_color="inverse" if len(df_divergencias) > 0 else "normal")
                with col_res3:
                    st.metric("Colab. no Ponto sem Mestra", len(df_nao_encontrados),
                              delta_color="inverse" if len(df_nao_encontrados) > 0 else "normal")
                
                if not df_divergencias.empty:
                    with st.expander(f"❌ Ver {len(df_divergencias)} divergências encontradas", expanded=True):
                        st.dataframe(df_divergencias, use_container_width=True, hide_index=True)
                        
                        st.subheader("📈 Estatísticas")
                        tipos_erro = df_divergencias['VALOR_NA_MESTRA'].value_counts().reset_index()
                        tipos_erro.columns = ['Valor na MESTRA (errado)', 'Quantidade']
                        st.dataframe(tipos_erro, use_container_width=True, hide_index=True)
                else:
                    st.success("✅ Nenhuma divergência encontrada! A planilha MESTRA está correta!")
                
                if not df_nao_encontrados.empty:
                    with st.expander(f"👥 Ver {len(df_nao_encontrados)} colaboradores no PONTO mas não na MESTRA"):
                        st.dataframe(df_nao_encontrados, use_container_width=True, hide_index=True)
                
                # ===== GERA ARQUIVOS =====
                status_text.text("📦 Gerando arquivos de saída...")
                
                # Gera relatório de divergências
                excel_divergencias = gerar_relatorio_divergencias_excel(
                    df_divergencias, df_nao_encontrados, df_mestra
                )
                
                # Gera planilha mestra corrigida COM FORMATAÇÃO
                excel_mestra_corrigida = gerar_mestra_corrigida_excel(df_corrigida, datas_mestra)
                
                progress_bar.progress(100)
                status_text.text("✅ Processamento concluído!")
                
                st.divider()
                
                # ===== DOWNLOAD EM ZIP =====
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                    zf.writestr('Relatorio_Divergencias_ABS_vs_PONTO.xlsx', excel_divergencias)
                    zf.writestr('MESTRA_CORRIGIDA_PELO_PONTO.xlsx', excel_mestra_corrigida)
                
                st.download_button(
                    "📦 Baixar ZIP (Relatório + Mestra Corrigida)",
                    zip_buffer.getvalue(),
                    f"Validacao_ABS_vs_PONTO.zip",
                    "application/zip",
                    use_container_width=True
                )
                
                st.success("""
                ### 📌 Instruções
                
                1. **Relatório de Divergências**: Use para mostrar aos encarregados onde estão errando
                2. **MESTRA CORRIGIDA**: Planilha com os dados VERDADEIROS baseados no PONTO (com formatação idêntica à original)
                
                > ⚠️ **Filtro ativo:** Apenas colaboradores com cargo **AUXILIAR DEPOSITO I, II ou III**
                > **Fins de semana:** Domingo = sempre "D". Sábado = "D" a menos que a jornada tenha **6x1**
                > A MESTRA CORRIGIDA substitui APENAS as células onde havia divergência.
                > Afastamentos > 15 dias foram IGNORADOS (não entram na contagem de absenteísmo).
                > **Múltiplos pontos:** Foram consolidados {len(files_ponto)} arquivo(s) de PONTO.
                """)
                
            except Exception as e:
                st.error(f"❌ Erro durante o processamento: {str(e)}")
                import traceback
                st.error(traceback.format_exc())
else:
    st.info("🌐 Faça upload dos dois arquivos para começar a validação.")
    
    st.divider()
    st.subheader("ℹ️ Como funciona")
    st.markdown("""
    ### Fluxo de Validação
    
    1. **Upload da MESTRA de ABS**: Planilha gerada pelos encarregados (com P, FI, FA, etc.)
    2. **Upload do PONTO**: Planilha de ponto eletrônico (verdade absoluta)
    3. **Filtro**: Apenas colaboradores com cargo **AUXILIAR DEPOSITO I, II ou III**
    4. **Processamento**: O sistema compara célula por célula
    5. **Geração de ZIP** com:
       - **MESTRA CORRIGIDA**: Nova planilha com os dados corretos do ponto (mesma formatação da original)
       - **Relatório de Divergências**: Mostra onde os encarregados erraram
    
    ### Mapeamento de Ocorrências
    
    | Ocorrência no PONTO | Marcação na MESTRA |
    |---|---|
    | Falta (sem justificativa) | **FI** |
    | Falta abonada (Obito, Amamentação, etc.) | **FÉRIAS-BH** |
    | Afastamento Doença/Acidente ≤ 15 dias | **FA** |
    | Afastamento > 15 dias | Ignorado |
    | Férias / Licença Maternidade | **FÉRIAS-BH** |
    | Sem marcação de entrada | **FI** |
    | Sem marcação de saída | **P** |
    | Entrada em atraso | **P** |
    
    ### Tratamento de Fins de Semana
    
    | Dia | Regra |
    |---|---|
    | **Domingo** | Sempre **D** (descanso) |
    | **Sábado** | **D** se jornada NÃO tiver **6x1** |
    | **Sábado com 6x1** | Processa normal (trabalha) |
    | Segunda a Sexta | Processa normal |
    """)