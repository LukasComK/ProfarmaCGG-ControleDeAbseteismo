import streamlit as st
import pandas as pd
import io
import re
import csv
import datetime
import math
import urllib.request
import json
import os
import time
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

st.set_page_config(page_title="Relatório Integrado", page_icon="📄", layout="wide")

st.title("📄 Relatório Integrado de Absenteísmo")
st.markdown("""
Esta página cruza dados de 4 planilhas diferentes para gerar um relatório automático 
agrupando os colaboradores com **Faltas Injustificadas (FI)** e **Faltas por Atestado (FA)**.
""")

st.divider()

col1, col2, col3 = st.columns(3)
with col1:
    f_abs = st.file_uploader("1️⃣ Planilha de Absenteísmo (Aba 'Dados')", type=["xlsx", "xls"])
    f_med = st.file_uploader("2️⃣ Planilha de Medida Disciplinar", type=["xlsx", "xls"])
with col2:
    f_dem = st.file_uploader("3️⃣ Planilha de Demissões", type=["xlsx", "xls"])
    f_ent = st.file_uploader("4️⃣ Planilha de Entrevista", type=["xlsx", "xls"])
with col3:
    f_gest = st.file_uploader("5️⃣ Base OCI/Gestor (CSV de Colaboradores)", type=["csv", "xlsx", "xls"])
    st.markdown("### 📅 Filtrar Período")
    filtrar_periodo = st.checkbox("Habilitar Filtro por Datas")
    if filtrar_periodo:
        cd1, cd2 = st.columns(2)
        with cd1:
            data_inicio = st.date_input("Data Início", format="DD/MM/YYYY")
        with cd2:
            data_fim = st.date_input("Data Final", format="DD/MM/YYYY")
    else:
        data_inicio, data_fim = None, None

def limpar_nome(nome):
    if pd.isna(nome) or str(nome).strip().upper() == "NAN":
        return ""
    # Remover aspas que possam vir do CSV (quoting=3)
    nome_limpo = str(nome).replace('"', '').replace("'", "")
    # Evitar problemas com espaços múltiplos no meio do nome
    return " ".join(nome_limpo.strip().upper().split())

def buscar_info_aproximada(nome_alvo, base_dados):
    """
    Tenta achar a chave no dicionário ou set. Prioriza match exato.
    Se não achar, testa se a chave está contida no nome, ou o nome na chave ("LIKE").
    """
    if not nome_alvo:
        return None if isinstance(base_dados, dict) else False

    # 1. Match exato
    if nome_alvo in base_dados:
        val = base_dados[nome_alvo] if isinstance(base_dados, dict) else True
        if isinstance(val, bool) and isinstance(base_dados, dict):
            return None
        return val
        
    # 2. Match por contéudo ("LIKE") - Ex: nome_alvo="DANIEL NUNES DE ALMEIDA" e chave da planilha de demissoes="DANIEL NUNES DE ALMEIDA(PCD) chamado 2802079"
    # Fazemos isso garantindo um mínimo de 10 letras pra não confundir "ANA" com "JULIANA" (falso positivo)
    if len(nome_alvo) >= 8:
        for chave in base_dados:
            if isinstance(chave, str) and len(chave) >= 8:
                if nome_alvo in chave or chave in nome_alvo:
                            val = base_dados[chave]
                            if isinstance(val, bool) and isinstance(base_dados, dict):
                                return None # Avoid returning boolean if dict value is boolean and unexpected
                            return val if isinstance(base_dados, dict) else True
    return None if isinstance(base_dados, dict) else False

def carregar_arquivo(f, sheet=None, todas_abas=False):
    raw_bytes = f.getvalue()
    
    # 0. Checa explicitamente pelo final do arquivo se é CSV
    if f.name.lower().endswith('.csv'):
        try:
            # Planilhas CSV do sistema brasileiro comumente usam sep=';' ou ',' e charset latin1 ou utf-8
            csv_str = raw_bytes.decode('latin1', errors='ignore')
            # Busca separador dinâmico nas primeiras 5 linhas, pois a linha 1 pode ser apenas "Título" sem colunas
            linhas = csv_str.split('\n')[:5]
            sep = ';' if any(';' in l for l in linhas) else ','
            # names=range(250) força o pandas a aceitar linhas largas em arquivos que começam com uma linha curta
            # Ignora quebra de aspas dupla (csv.QUOTE_NONE é 3) para evitar "',' expected after '"'"
            df_csv = pd.read_csv(io.StringIO(csv_str), sep=sep, header=None, names=range(250), quoting=3, on_bad_lines='skip', engine='python')
            if todas_abas:
                return {"Aba_CSV": df_csv}
            return df_csv
        except Exception as e:
            raise ValueError(f"Erro ao processar o arquivo CSV '{f.name}': {str(e)}")

    # 1. Tentativa padrão (pandas decide o engine via extensão) para XLSX/XLS
    try:
        if todas_abas:
            return pd.read_excel(io.BytesIO(raw_bytes), sheet_name=None, header=None)
        
        if sheet:
            try:
                return pd.read_excel(io.BytesIO(raw_bytes), sheet_name=sheet, header=None)
            except:
                pass
        return pd.read_excel(io.BytesIO(raw_bytes), header=None)
    except Exception as e_excel:
        # 2. Tentativa forçando openpyxl (pode ser um .xls que na verdade é .xlsx)
        try:
            if todas_abas:
                return pd.read_excel(io.BytesIO(raw_bytes), sheet_name=None, header=None, engine='openpyxl')
                
            if sheet:
                try:
                    return pd.read_excel(io.BytesIO(raw_bytes), sheet_name=sheet, header=None, engine='openpyxl')
                except:
                    pass
            return pd.read_excel(io.BytesIO(raw_bytes), header=None, engine='openpyxl')
        except:
            # 3. Fallback para HTML masquerade (muito comum em sistemas de controle de ponto e catracas)
            try:
                html_str = raw_bytes.decode('latin1', errors='ignore')
                dfs = pd.read_html(io.StringIO(html_str), header=None)
                if dfs:
                    if todas_abas:
                        return {"Aba_HTML": dfs[0]}
                    return dfs[0]
            except:
                # 4. Fallback final para CSV
                try:
                    csv_str = raw_bytes.decode('latin1', errors='ignore')
                    df_csv = pd.read_csv(io.StringIO(csv_str), sep=';', header=None)
                    if todas_abas:
                        return {"Aba_CSV": df_csv}
                    return df_csv
                except:
                    raise ValueError(f"Não foi possível ler '{f.name}'. Ele pode estar protegido por senha, criptografado ou em um formato desconhecido. (Erro: {str(e_excel)})")

if f_abs and f_med and f_dem and f_ent and f_gest:
    if st.button("🚀 Gerar Relatório Integrado", type="primary", use_container_width=True):
        with st.spinner("Analisando e cruzando as planilhas. Isso pode levar alguns segundos..."):
            try:
                # ---------------------------------------------------------
                # 1. PROCESSAR PLANILHA DE ABSENTEÍSMO
                # ---------------------------------------------------------
                df_abs = carregar_arquivo(f_abs, sheet="Dados")
                
                absencias = {} # Formato: { "NOME": {"FI": [], "FA": []} }
                
                # Mapear os cabeçalhos de data (Colunas J até AN -> Índices 9 a 39)
                datas_colunas = {}
                col_fim = min(50, len(df_abs.columns)) # Aumentado para 50 para garantir que pega as datas após a M
                
                # Tenta achar a linha de cabeçalho (que tem as datas) olhando as primeiras linhas
                linha_cabecalho = 0
                for r in range(min(5, len(df_abs))):
                    # Procura na linha algo que pareça um cabeçalho (ex: Nome na pos 0) ou se a J/K tem algo
                    if pd.notna(df_abs.iloc[r, 0]) and pd.notna(df_abs.iloc[r, 9]):
                        linha_cabecalho = r
                        break
                        
                # Achar a verdadeira coluna de Data de Admissão (padrão é M=12)
                col_admissao = 12
                for c in range(len(df_abs.columns)):
                    val_str = str(df_abs.iloc[linha_cabecalho, c]).upper()
                    if "ADMISS" in val_str:
                        col_admissao = c
                        break
                        
                with st.expander("🛠️ Diagnóstico de Admissão (Clique para ver os dados puxados da planilha)"):
                    st.write(f"Índice da coluna identificada como Admissão: **{col_admissao}** (A=0, B=1... M=12)")
                    if col_admissao < len(df_abs.columns):
                        st.write(f"Título lido nesta coluna: **{df_abs.iloc[linha_cabecalho, col_admissao]}**")
                        
                        amostra_dados = []
                        validos = 0
                        # Vamos puxar os próximos 10 colaboradores para ver o que tá vindo
                        for test_r in range(linha_cabecalho + 1, min(linha_cabecalho + 15, len(df_abs))):
                            nome_teste = df_abs.iloc[test_r, 0]
                            val_admissao = df_abs.iloc[test_r, col_admissao]
                            if pd.notna(nome_teste):
                                amostra_dados.append(f"Nome: {nome_teste} -> Admissão lida: {val_admissao} (Tipo: {type(val_admissao).__name__})")
                                validos += 1
                        
                        st.write("Amostra das 10 primeiras linhas:")
                        for info in amostra_dados:
                            st.text(info)
                    else:
                        st.write("A coluna identificada extrapolou o número de colunas da planilha!")
                        
                for c in range(9, col_fim):
                    # Não tenta ler a coluna de Admissão como data de faltas
                    if c == col_admissao:
                        continue
                        
                    val = df_abs.iloc[linha_cabecalho, c]
                    if pd.isna(val): continue
                    
                    dt_col = None
                    try:
                        # Tenta extrair a data real do cabeçalho
                        if isinstance(val, pd.Timestamp):
                            dt_col = val.date()
                        elif isinstance(val, datetime.date) and not isinstance(val, pd.Timestamp):
                            dt_col = val
                        else:
                            val_str = str(val).strip()
                            # Caso seja apenas formato dia/mês estilo "01/04" ou "01-04"
                            if re.match(r'^\d{1,2}[/-]\d{1,2}$', val_str):
                                val_str += '/2026' # Adiciona ano fixo temporário para conseguir ordenar a semana
                            val_dt = pd.to_datetime(val_str, dayfirst=True, errors='coerce')
                            if pd.notna(val_dt):
                                dt_col = val_dt.date()
                    except:
                        pass
                        
                    # Filtra data se habilitado
                    if filtrar_periodo and dt_col and data_inicio and data_fim:
                        if not (data_inicio <= dt_col <= data_fim):
                            continue # Pula a coluna dessa data se estiver fora do período
                            
                    # Manter a data real na lista
                    if dt_col is not None:
                        datas_colunas[c] = dt_col # Salva como datetime.date real
                    else:
                        datas_colunas[c] = str(val)[:10]

                # Percorrer os colaboradores
                for r in range(linha_cabecalho + 1, len(df_abs)):
                    nome = limpar_nome(df_abs.iloc[r, 0]) # Coluna A (Índice 0)
                    if not nome: continue
                    
                    data_adm = df_abs.iloc[r, col_admissao] if len(df_abs.columns) > col_admissao else None
                    
                    # Itera apenas pelas colunas que foram mapeadas/aprovadas pelo Filtro de Datas
                    for c in datas_colunas.keys():
                        status = limpar_nome(df_abs.iloc[r, c])
                        if status in ['FI', 'FA', 'P']:
                            if nome not in absencias:
                                absencias[nome] = {'FI': [], 'FA': [], 'P': [], 'admissao': data_adm}
                            if 'P' not in absencias[nome]:
                                absencias[nome]['P'] = []
                                
                            # Se a primeira vez for lido e era vazio, tenta repopular
                            if absencias[nome].get('admissao') is None or str(absencias[nome].get('admissao')).strip() in ['', 'nan', 'NaT', 'NaN']:
                                absencias[nome]['admissao'] = data_adm
                                
                            absencias[nome][status].append(datas_colunas[c])

                # ---------------------------------------------------------
                # 2. PROCESSAR MEDIDAS DISCIPLINARES
                # ---------------------------------------------------------
                df_med = carregar_arquivo(f_med)
                medidas_dict = {}
                
                col_max_med = len(df_med.columns)
                for r in range(len(df_med)):
                    # Col B (1), Col D (3), Col E (4)
                    if col_max_med > 1:
                        nome = limpar_nome(df_med.iloc[r, 1])
                    else: continue
                    
                    tipo = limpar_nome(df_med.iloc[r, 3]) if col_max_med > 3 else ""
                    detalhe_data = str(df_med.iloc[r, 4]).strip() if col_max_med > 4 and pd.notna(df_med.iloc[r, 4]) else ""
                    
                    if nome and "FALTA INJUSTIFICADA" in tipo:
                        medidas_dict[nome] = detalhe_data

                # ---------------------------------------------------------
                # 3. PROCESSAR DEMISSÕES (Leitura Completa em Toda a Pasta de Trabalho)
                # ---------------------------------------------------------
                dfs_dem = carregar_arquivo(f_dem, todas_abas=True)
                demissoes_dict = {}
                
                # Percorre todas as abas lidas
                for nome_aba, df_dem_aba in dfs_dem.items():
                    col_max_dem = len(df_dem_aba.columns)
                    
                    for r in range(len(df_dem_aba)):
                        # Procura na Col B (1), D (3), F (5)
                        if col_max_dem > 1:
                            nome = limpar_nome(df_dem_aba.iloc[r, 1])
                        else: 
                            continue
                        
                        if nome:
                            val_data = df_dem_aba.iloc[r, 3] if col_max_dem > 3 else None
                            if pd.notna(val_data) and str(val_data).strip() != "NaT":
                                try:
                                    d_data = pd.to_datetime(val_data).strftime('%d/%m/%Y')
                                except:
                                    d_data = str(val_data)[:10]
                            else:
                                d_data = "Data Indisponível"
                                
                            val_tipo = df_dem_aba.iloc[r, 5] if col_max_dem > 5 else None
                            d_tipo = str(val_tipo).strip().capitalize() if pd.notna(val_tipo) else "Tipo Indisponível"
                            demissoes_dict[nome] = {'data': d_data, 'tipo': d_tipo}
                
                # ---------------------------------------------------------
                # 4. PROCESSAR ENTREVISTAS DE ABSENTEÍSMO
                # ---------------------------------------------------------
                df_ent = carregar_arquivo(f_ent, sheet="2026")
                    
                entrevistas_dict = {}
                col_max_ent = len(df_ent.columns)
                for r in range(len(df_ent)):
                    if col_max_ent > 1:
                        nome = limpar_nome(df_ent.iloc[r, 1]) # Col B (1)
                        if nome:
                            motivo = str(df_ent.iloc[r, 8]).strip() if col_max_ent > 8 and pd.notna(df_ent.iloc[r, 8]) else ""
                            entrevistas_dict[nome] = motivo

                # ---------------------------------------------------------
                # PROCESSAR GESTORES E SUPERVISORES E ADMISSÃO (BASE CSV)
                # ---------------------------------------------------------
                df_gest = carregar_arquivo(f_gest)
                gestores_dict = {}
                admissoes_dict = {}
                ceps_dict = {}
                
                # Busca automática pela linha e índices das colunas no CSV/Excel
                linha_cab_gest = 0
                col_colab = 3  # Padrão original no CSV era 3 se desse erro
                col_nome_gest = 25 # Padrão original era 25
                col_admissao = 12 # Padrão é coluna M (índice 12)
                col_cep = 63  # Padrão é coluna BL (índice 63)
                
                for r in range(min(15, len(df_gest))):
                    linha_txt = [str(v).upper() for v in df_gest.iloc[r, :]]
                    # Muitos CSVs vêm com aspas duplas, então limpamos para garantir o match
                    linha_txt_limpa = [v.replace('"', '').replace("'", "").strip() for v in linha_txt]
                    if any("COLABORADOR" in v for v in linha_txt_limpa) and any("NOME GESTOR" in v for v in linha_txt_limpa):
                        linha_cab_gest = r
                        for idx, v in enumerate(linha_txt_limpa):
                            if "COLABORADOR" in v: col_colab = idx
                            if "NOME GESTOR" in v: col_nome_gest = idx
                            if "ADMISS" in v: col_admissao = idx
                            # Procura exato a coluna CEP
                            if v == "CEP" or v == "C.E.P" or " CEP " in f" {v} ": col_cep = idx
                        break
                        
                for r in range(linha_cab_gest + 1, len(df_gest)):
                    if len(df_gest.columns) > max(col_colab, col_nome_gest):
                        colab_nome = limpar_nome(df_gest.iloc[r, col_colab])
                        
                        if colab_nome and str(colab_nome) != "NAN":
                            # Gestor
                            gest_nome = limpar_nome(df_gest.iloc[r, col_nome_gest])
                            if gest_nome:
                                gestores_dict[colab_nome] = gest_nome
                            
                            # Admissão
                            if len(df_gest.columns) > col_admissao:
                                val_adm = df_gest.iloc[r, col_admissao]
                                if pd.notna(val_adm) and str(val_adm).strip() != "":
                                    # Limpa possíveis aspas no texto da data (ex: CSV bugado)
                                    val_adm = str(val_adm).replace('"', '').replace("'", "").strip()
                                    admissoes_dict[colab_nome] = val_adm
                                    
                            # CEP
                            if len(df_gest.columns) > col_cep:
                                val_cep = df_gest.iloc[r, col_cep]
                                if pd.notna(val_cep) and str(val_cep).strip() != "":
                                    # Formata padronizando remover aspas
                                    val_cep = str(val_cep).replace('"', '').replace("'", "").strip()
                                    ceps_dict[colab_nome] = val_cep

                # =========================================================
                # FUNÇÕES AUXILIARES DE CÁLCULO
                # =========================================================
                def calcular_distancia_cep(cep_colab):
                    if not cep_colab or str(cep_colab).strip() in ["", "nan", "NaN", "NaT", "N/A", "None"]:
                        return "Não Informado"
                        
                    cep_limpo = str(cep_colab).replace("-", "").replace(".", "").strip()
                    if len(cep_limpo) != 8 or not cep_limpo.isdigit():
                        return f"CEP Inválido ({cep_limpo})"
                        
                    # Utilize st.session_state para cache de CEP dinâmico, evitando bloquear leitura em disco ou timeout repetitivo
                    if "dict_ceps_cache" not in st.session_state:
                        st.session_state["dict_ceps_cache"] = {}
                        
                    cache_local = st.session_state["dict_ceps_cache"]
                            
                    if cep_limpo in cache_local:
                        if cache_local[cep_limpo] == "Erro": return "CEP Não Encontrado na API"
                        lat1, lon1 = cache_local[cep_limpo]
                    else:
                        # Buscar na Awesome API (Especialista em CEPs Brasileiros com Lat/Lng)
                        try:
                            import urllib.request
                            import urllib.error
                            url = f'https://cep.awesomeapi.com.br/json/{cep_limpo}'
                            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'})
                            with urllib.request.urlopen(req, timeout=3) as res:
                                data = json.loads(res.read())
                                lat_str = data.get('lat')
                                lng_str = data.get('lng')
                                if lat_str and lng_str:
                                    lat1, lon1 = float(lat_str), float(lng_str)
                                    cache_local[cep_limpo] = (lat1, lon1)
                                else:
                                    cache_local[cep_limpo] = "Erro"
                                    return "CEP Sem Coordenadas"
                        except Exception as e:
                            # Se a API falhar (ex: bloqueio de rede ou timeout), salva como Erro temporário para não explodir
                            cache_local[cep_limpo] = "Erro"
                            return "API Offline/Limitada"
                            
                    if lat1 is None or lon1 is None:
                        return "Não Encontrado"
                        
                    # Coordenadas da Empresa (Profarma CGG / Modular - 23078-001)
                    lat2, lon2 = -22.866352, -43.5856617
                    
                    # Se for o exato local da empresa
                    if (lat1 == lat2 and lon1 == lon2) or cep_limpo == "23078001":
                        return "0.0 km"
                        
                    R = 6371.0 # Raio da Terra
                    lat1_r, lon1_r, lat2_r, lon2_r = map(math.radians, [lat1, lon1, lat2, lon2])
                    dlat = lat2_r - lat1_r
                    dlon = lon2_r - lon1_r
                    
                    a = math.sin(dlat/2)**2 + math.cos(lat1_r)*math.cos(lat2_r)*math.sin(dlon/2)**2
                    c = 2 * math.asin(math.sqrt(a))
                    dist_km = R * c
                    
                    # Evita exibir algo como "0.0 km" para CEPs distintos se a fórmula não for precisa
                    if dist_km < 0.1:
                        return "< 0.1 km"
                        
                    return f"{dist_km:.1f} km"

                def calcular_tempo_servico(data_adm, data_ref=None):
                    if data_adm is None or pd.isna(data_adm) or str(data_adm).strip() in ["", "NaT", "NaN", "nan"]:
                        return "N/A"
                    
                    if not isinstance(data_adm, datetime.date):
                        try:
                            # Se for Timestamp, extrai a data
                            if isinstance(data_adm, pd.Timestamp):
                                data_adm = data_adm.date()
                            else:
                                dt = pd.to_datetime(str(data_adm), dayfirst=True, errors='coerce')
                                if pd.isna(dt):
                                    return "N/A"
                                data_adm = dt.date()
                        except:
                            return "N/A"
                    
                    if data_adm is None or pd.isna(data_adm) or not isinstance(data_adm, datetime.date):
                        return "N/A"
                        
                    if data_ref is None:
                        data_ref = datetime.date.today()
                        
                    anos = data_ref.year - data_adm.year
                    meses = data_ref.month - data_adm.month
                    if data_ref.day < data_adm.day:
                        meses -= 1
                    if meses < 0:
                        anos -= 1
                        meses += 12
                    
                    if anos < 0:
                        return "0a 0m"
                    return f"{anos}a {meses}m"

                # =========================================================
                # FUNÇÃO PARA CONFRONTAR MESES DAS FALTAS E DA MEDIDA
                # =========================================================
                def tem_mes_correspondente(lista_datas_fi, texto_medida):
                    if not texto_medida:
                        return False
                    
                    texto_medida = str(texto_medida).lower()
                    
                    # 1. Pega os meses das faltas (Ex: "07/04", "15/04/2026")
                    meses_fi = set()
                    for d in lista_datas_fi:
                        match_d = re.search(r'\d{1,2}/(\d{1,2})', str(d))
                        if match_d:
                            meses_fi.add(match_d.group(1).zfill(2))
                        else:
                            # Tenta yyyy-mm-dd
                            match_t = re.search(r'\d{4}-(\d{1,2})-\d{1,2}', str(d))
                            if match_t:
                                meses_fi.add(match_t.group(1).zfill(2))
                    
                    # 2. Pega os meses descritos no texto da Medida
                    meses_med = set()
                    # Busca dd/mm ou dd/mm/yyyy
                    matches_b = re.findall(r'\b\d{1,2}/(\d{1,2})\b', texto_medida)
                    meses_med.update([m.zfill(2) for m in matches_b])
                    
                    # Busca meses por extenso, caso alguém escreva "abril"
                    mapa_meses = {
                        'jan': '01', 'fev': '02', 'mar': '03', 'abr': '04',
                        'mai': '05', 'jun': '06', 'jul': '07', 'ago': '08',
                        'set': '09', 'out': '10', 'nov': '11', 'dez': '12'
                    }
                    for nome_mes, num_mes in mapa_meses.items():
                        if nome_mes in texto_medida:
                            meses_med.add(num_mes)
                            
                    # Se não detectou data/mês nenhum na medida, assume True (pra não perder nada só por causa do texto genérico)
                    if not meses_med:
                        return True
                        
                    # 3. Retorna se há intercessão
                    return bool(meses_fi.intersection(meses_med))

                # =========================================================
                # CONSTRUÇÃO DO RELATÓRIO EM PLANILHA (GERAL + SEMANAL)
                # =========================================================
                st.success("✅ Cruzamento de dados concluído com sucesso!")
                
                # -------------------------------
                # RELATÓRIO GERAL (CONSOLIDADO)
                # -------------------------------
                lista_fi_geral = []
                lista_fa_geral = []
                lista_entrevistas_pendentes = []
                
                for nome, rec in sorted(absencias.items()):
                    # Pegar Gestor e Supervisor
                    gestor_nome = buscar_info_aproximada(nome, gestores_dict)
                    gestor_final = gestor_nome if gestor_nome else "Sem Gestor Mapeado"
                    
                    supervisor_final = "Sem Supervisor Mapeado"
                    if gestor_nome:
                        supervisor_nome = buscar_info_aproximada(gestor_nome, gestores_dict)
                        if supervisor_nome:
                            supervisor_final = supervisor_nome
                            
                    # Distância CEP
                    cep_val = buscar_info_aproximada(nome, ceps_dict)
                    distancia_residencia = calcular_distancia_cep(cep_val)

                    # VERIFICAÇÃO DE ENTREVISTAS PENDENTES (FA -> P)
                    entrevista_motivo = buscar_info_aproximada(nome, entrevistas_dict)
                    if not entrevista_motivo and rec.get('FA'):
                        timeline = sorted([(d, 'FA') for d in rec.get('FA', [])] + [(d, 'P') for d in rec.get('P', [])], key=lambda x: x[0] if isinstance(x[0], datetime.date) else datetime.date.min)
                        current_fa_block = []
                        for d, st_timeline in timeline:
                            if st_timeline == 'FA':
                                current_fa_block.append(d)
                            elif st_timeline == 'P':
                                if current_fa_block:
                                    dias_str = ", ".join([df.strftime('%d/%m') if isinstance(df, datetime.date) and pd.notna(df) else str(df) for df in current_fa_block])
                                    ret_str = d.strftime('%d/%m') if isinstance(d, datetime.date) and pd.notna(d) else str(d)
                                    lista_entrevistas_pendentes.append({
                                        "Colaborador": nome,
                                        "Gestor": gestor_final,
                                        "Supervisor": supervisor_final,
                                        "Data de Retorno": ret_str,
                                        "Datas de Ausência": dias_str,
                                        "Quantidade das faltas": len(current_fa_block),
                                        "Distância Residência x Trabalho": distancia_residencia
                                    })
                                    current_fa_block = []

                    if rec.get('FI'):
                        med = buscar_info_aproximada(nome, medidas_dict)
                        if med and tem_mes_correspondente(rec['FI'], med):
                            texto_medida = f"Sim ({med})"
                        else:
                            texto_medida = "Não solicitada"
                        
                        dem = buscar_info_aproximada(nome, demissoes_dict)
                        texto_dem = f"{dem['data']} - {dem['tipo']}" if dem else "Sem projeção"
                        
                        data_admissao_csv = buscar_info_aproximada(nome, admissoes_dict)
                        ts_geral = calcular_tempo_servico(data_admissao_csv)
                        
                        lista_fi_geral.append({
                            "Colaborador": nome, "Gestor": gestor_final, "Supervisor": supervisor_final,
                            "Datas das Faltas (FI)": ", ".join([d.strftime('%d/%m') if isinstance(d, datetime.date) and pd.notna(d) else str(d) for d in rec['FI']]),
                            "Quantidade de Faltas": len(rec['FI']), "Medida Disciplinar": texto_medida, "Desligamento": texto_dem,
                            "Tempo de Serviço": ts_geral, "Distância Residência x Trabalho": distancia_residencia
                        })
                        
                    if rec['FA']:
                        entrevista_motivo = buscar_info_aproximada(nome, entrevistas_dict)
                        if entrevista_motivo is not None:
                            texto_entrevista = "Sim"
                            texto_motivo = entrevista_motivo if entrevista_motivo else "Sem motivo detalhado"
                        else:
                            texto_entrevista = "Não possui"
                            texto_motivo = "N/A"
                            
                        dem = buscar_info_aproximada(nome, demissoes_dict)
                        texto_dem = f"{dem['data']} - {dem['tipo']}" if dem else "Sem projeção"
                        
                        data_admissao_csv = buscar_info_aproximada(nome, admissoes_dict)
                        ts_geral = calcular_tempo_servico(data_admissao_csv)
                        
                        lista_fa_geral.append({
                            "Colaborador": nome, "Gestor": gestor_final, "Supervisor": supervisor_final,
                            "Datas dos Atestados (FA)": ", ".join([d.strftime('%d/%m') if isinstance(d, datetime.date) and pd.notna(d) else str(d) for d in rec['FA']]),
                            "Quantidade de Faltas": len(rec['FA']), "Entrevista de Absenteísmo": texto_entrevista,
                            "Motivo": texto_motivo,
                            "Desligamento": texto_dem,
                            "Tempo de Serviço": ts_geral,
                            "Distância Residência x Trabalho": distancia_residencia
                        })

                df_fi_geral = pd.DataFrame(lista_fi_geral)
                if not df_fi_geral.empty:
                    df_fi_geral = df_fi_geral.sort_values(by="Quantidade de Faltas", ascending=False)
                    
                df_fa_geral = pd.DataFrame(lista_fa_geral)
                if not df_fa_geral.empty:
                    df_fa_geral = df_fa_geral.sort_values(by="Quantidade de Faltas", ascending=False)
                    
                df_entrevistas_pendentes = pd.DataFrame(lista_entrevistas_pendentes)
                if not df_entrevistas_pendentes.empty:
                    df_entrevistas_pendentes = df_entrevistas_pendentes.sort_values(by="Data de Retorno", ascending=True)

                # -------------------------------
                # RELATÓRIO NOVO (OFENSORES SEMANAIS)
                # -------------------------------
                datas_validas = sorted([d for d in datas_colunas.values() if isinstance(d, datetime.date) and pd.notna(d)])
                semanas = {}
                for d in datas_validas:
                    ano, sem_num, _ = d.isocalendar()
                    k = (ano, sem_num)
                    if k not in semanas:
                        semanas[k] = []
                    semanas[k].append(d)
                
                semanas_lista = []
                for idx, (k, dias) in enumerate(sorted(semanas.items())):
                    inicio = min(dias)
                    fim = max(dias)
                    semanas_lista.append({
                        "nome": f"Semana {idx+1} - {inicio.strftime('%d/%m')} a {fim.strftime('%d/%m')}",
                        "dias": dias
                    })
                
                if not semanas_lista:
                    semanas_lista.append({"nome": "Período Único", "dias": []})
                
                lista_ofensores_fi = []
                lista_ofensores_fa = []
                
                for sem in semanas_lista:
                    # Linha Separadora de Semana
                    lista_ofensores_fi.append({
                        "Nome": sem["nome"], "Gestor": "", "Dias das Faltas": "", "Faltas Injustificadas": None, "Desligamento": "", "Tempo de Serviço": "", "Distância Residência x Trabalho": ""
                    })
                    lista_ofensores_fa.append({
                        "Nome": sem["nome"], "Gestor": "", "Dias das Faltas": "", "Faltas por Atestado": None, "Entrevista de Absenteísmo": "", "Motivo": "", "Desligamento": "", "Tempo de Serviço": "", "Distância Residência x Trabalho": ""
                    })
                    
                    pessoas_fi = []
                    pessoas_fa = []
                    
                    for nome, rec in sorted(absencias.items()):
                        gestor_nome = buscar_info_aproximada(nome, gestores_dict)
                        gestor_final = gestor_nome if gestor_nome else "Sem Gestor Mapeado"
                        
                        fi_na_semana = [d for d in rec['FI'] if (isinstance(d, datetime.date) and pd.notna(d) and d in sem["dias"]) or (not sem["dias"])]
                        fa_na_semana = [d for d in rec['FA'] if (isinstance(d, datetime.date) and pd.notna(d) and d in sem["dias"]) or (not sem["dias"])]
                        qtd_fi = len(fi_na_semana)
                        qtd_fa = len(fa_na_semana)
                        
                        data_admissao_csv = buscar_info_aproximada(nome, admissoes_dict)
                        ts_semanal = calcular_tempo_servico(data_admissao_csv)
                        
                        # Distância CEP
                        cep_val = buscar_info_aproximada(nome, ceps_dict)
                        distancia_residencia = calcular_distancia_cep(cep_val)
                        
                        dem = buscar_info_aproximada(nome, demissoes_dict)
                        texto_dem = f"{dem['data']} - {dem['tipo']}" if dem else "Sem projeção"
                        
                        # Se faltou INJUSTIFICADO na semana, coloca na aba de Semanais FI
                        if qtd_fi > 0:
                            dias_fi_str = ", ".join([d.strftime('%d/%m') if isinstance(d, datetime.date) and pd.notna(d) else str(d) for d in fi_na_semana])
                            pessoas_fi.append({
                                "Nome": nome, "Gestor": gestor_final, "Dias das Faltas": dias_fi_str, "Faltas Injustificadas": qtd_fi, "Desligamento": texto_dem, "Tempo de Serviço": ts_semanal, "Distância Residência x Trabalho": distancia_residencia
                            })
                            
                        # Se teve ATESTADO na semana, coloca na aba de Semanais FA
                        if qtd_fa > 0:
                            dias_fa_str = ", ".join([d.strftime('%d/%m') if isinstance(d, datetime.date) and pd.notna(d) else str(d) for d in fa_na_semana])
                            entrevista_motivo = buscar_info_aproximada(nome, entrevistas_dict)
                            if entrevista_motivo is not None:
                                texto_entrevista = "Sim"
                                texto_motivo = entrevista_motivo if entrevista_motivo else "Sem motivo detalhado"
                            else:
                                texto_entrevista = "Não possui"
                                texto_motivo = "N/A"
                                
                            pessoas_fa.append({
                                "Nome": nome, "Gestor": gestor_final, "Dias das Faltas": dias_fa_str, "Faltas por Atestado": qtd_fa, "Entrevista de Absenteísmo": texto_entrevista, "Motivo": texto_motivo, "Desligamento": texto_dem, "Tempo de Serviço": ts_semanal, "Distância Residência x Trabalho": distancia_residencia
                            })

                    # Ordenar ofensores
                    if pessoas_fi:
                        pessoas_fi.sort(key=lambda x: x["Faltas Injustificadas"], reverse=True)
                        lista_ofensores_fi.extend(pessoas_fi)
                    if pessoas_fa:
                        pessoas_fa.sort(key=lambda x: x["Faltas por Atestado"], reverse=True)
                        lista_ofensores_fa.extend(pessoas_fa)

                df_of_fi = pd.DataFrame(lista_ofensores_fi)
                df_of_fa = pd.DataFrame(lista_ofensores_fa)

                # =========================================================
                # EXIBIÇÃO E EXPORTAÇÃO PARA EXCEL
                # =========================================================
                tab1, tab2, tab3, tab4 = st.tabs(["🔴 Faltas Injustificadas (Geral)", "🟡 Faltas por Atestado (Geral)", "📅 Ofensores Semanais (Prévia)", "⚠️ Entrevistas Pendentes"])
                
                with tab1:
                    if not df_fi_geral.empty:
                        st.dataframe(df_fi_geral, use_container_width=True, hide_index=True)
                    else:
                        st.info("💡 Nenhum colaborador com Falta Injustificada (FI) encontrado no período!")
                        
                with tab2:
                    if not df_fa_geral.empty:
                        st.dataframe(df_fa_geral, use_container_width=True, hide_index=True)
                    else:
                        st.info("💡 Nenhum colaborador com Falta por Atestado (FA) encontrado no período!")
                        
                with tab3:
                    st.write("##### Ofensores Injustificados (FI)")
                    if not df_of_fi.empty:
                        st.dataframe(df_of_fi, use_container_width=True, hide_index=True)
                        
                    st.write("##### Ofensores de Atestado (FA)")
                    if not df_of_fa.empty:
                        st.dataframe(df_of_fa, use_container_width=True, hide_index=True)
                        
                with tab4:
                    st.write("##### Colaboradores com Faltas por Atestado (FA) pendentes de Entrevista de Absenteísmo após Retorno (P)")
                    if not df_entrevistas_pendentes.empty:
                        st.dataframe(df_entrevistas_pendentes, use_container_width=True, hide_index=True)
                    else:
                        st.success("✅ Nenhuma entrevista pendente! Todos os colaboradores com FA que retornaram já possuem entrevista.")
                
                # Monta arquivo XLSX em memória
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Planilhas originais consolidadas
                    if not df_fi_geral.empty:
                        df_fi_geral.to_excel(writer, sheet_name='Falta Injustificada', index=False)
                    else:
                        pd.DataFrame({'Aviso': ['Sem faltas']}).to_excel(writer, sheet_name='Falta Injustificada', index=False)
                        
                    if not df_fa_geral.empty:
                        df_fa_geral.to_excel(writer, sheet_name='Faltas por Atestado', index=False)
                    else:
                        pd.DataFrame({'Aviso': ['Sem faltas']}).to_excel(writer, sheet_name='Faltas por Atestado', index=False)
                    
                    # Novas Planilhas Semanais
                    if not df_of_fi.empty:
                        df_of_fi.to_excel(writer, sheet_name='Semanal Injustificadas', index=False)
                    else:
                        pd.DataFrame({'Aviso': ['Sem faltas semanais']}).to_excel(writer, sheet_name='Semanal Injustificadas', index=False)
                        
                    if not df_of_fa.empty:
                        df_of_fa.to_excel(writer, sheet_name='Semanal Atestados', index=False)
                    else:
                        pd.DataFrame({'Aviso': ['Sem faltas semanais']}).to_excel(writer, sheet_name='Semanal Atestados', index=False)

                    # Entrevistas Pendentes
                    if not df_entrevistas_pendentes.empty:
                        df_entrevistas_pendentes.to_excel(writer, sheet_name='Entrevistas Pendentes', index=False)
                    else:
                        pd.DataFrame({'Aviso': ['Nenhuma entrevista pendente']}).to_excel(writer, sheet_name='Entrevistas Pendentes', index=False)
                    
                    # Estilizando as planilhas
                    workbook = writer.book
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    cor_profarma = PatternFill(start_color="0D4F45", end_color="0D4F45", fill_type="solid") # Verde Profarma
                    
                    for nome_sheet in workbook.sheetnames:
                        ws = workbook[nome_sheet]
                        ws.row_dimensions[1].height = 30
                        
                        is_semanal = ('Semanal' in nome_sheet)
                        
                        for r_idx, row in enumerate(ws.iter_rows()):
                            # Se for as novas abas Semanais, queremos destacar a linha da Semana se o Nome começar com "Semana"
                            is_linha_semana_sep = False
                            if is_semanal and r_idx > 0:
                                val_cell = row[0].value
                                if val_cell and str(val_cell).startswith("Semana "):
                                    is_linha_semana_sep = True
                                    
                            for c_idx, cell in enumerate(row):
                                if r_idx == 0 or is_linha_semana_sep:
                                    cell.fill = cor_profarma
                                    cell.font = Font(color="FFFFFF", bold=True)
                                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                                else:
                                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                                
                                cell.border = thin_border
                                
                                # Lógica para fazer a linha mesclada caso seja o separador de semanas ("Semana 1...")
                                # OpenPyxl precisa mesclar depois, então tratamos o estilo da célula primeiro.
                                
                        # Largura das Colunas
                        for col in ws.columns:
                            max_length = 0
                            col_letter = col[0].column_letter
                            
                            for cell in col:
                                val_str = str(cell.value) if cell.value else ""
                                if not val_str.startswith("Semana "):
                                    max_length = max(max_length, len(val_str))
                            
                            adjusted_width = min(max_length + 4, 40)
                            ws.column_dimensions[col_letter].width = adjusted_width
                            
                        # Mescla a linha de separação da semana em toda a tabela nas abas Semanais
                        if is_semanal:
                            max_col = ws.max_column
                            # iter_rows(min_row=2)
                            for r_idx in range(2, ws.max_row + 1):
                                val_cell = str(ws.cell(row=r_idx, column=1).value)
                                if val_cell and val_cell.startswith("Semana "):
                                    ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=max_col)
                
                excel_data = output.getvalue()
                
                st.divider()
                # Botão verde destacado da planilha
                st.download_button(
                    label="📥 Exportar Relatório em Planilha (XLSX)",
                    data=excel_data,
                    file_name="Relatorio_Cruzado_Absenteismo.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
                        
            except Exception as e:
                st.error(f"❌ Ocorreu um erro durante o processamento das planilhas: {e}")
                st.exception(e)
else:
    st.info("⚠️ Aguardando o envio das 4 planilhas para habilitar o relatório.")
