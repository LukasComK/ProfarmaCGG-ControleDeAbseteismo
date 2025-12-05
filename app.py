# Ferramenta de Lançamento de Absenteísmo com Busca LIKE
import streamlit as st
import pandas as pd
from unidecode import unidecode
import io
import datetime
from dateutil.relativedelta import relativedelta
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher
import plotly.express as px
from copy import copy
from openpyxl import load_workbook, Workbook

# Nota: A página "👥 Colaboradores" foi criada em pages/1_👥_Colaboradores.py
# Ela será exibida automaticamente pelo Streamlit como uma página multipage

def obter_feriados_brasil(ano):
    """
    Busca feriados nacionais do Brasil para um ano específico via API Brasil API.
    Retorna um dicionário {data: nome_feriado}
    """
    import requests
    feriados = {}
    try:
        url = f"https://brasilapi.com.br/api/feriados/v1/{ano}"
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            dados = response.json()
            for feriado in dados:
                try:
                    data = datetime.datetime.strptime(feriado['date'], '%Y-%m-%d').date()
                    nome = feriado.get('name', 'Feriado')
                    feriados[data] = nome
                except:
                    pass
    except Exception as e:
        print(f"Erro ao buscar feriados: {e}")
    
    return feriados

def marcar_feriados_na_workbook(workbook, feriados, mapa_datas, mapa_cores):
    """
    Marca colunas inteiras de feriados na workbook como "FERIADO" (sobrescreve tudo)
    """
    if not feriados:
        return
    
    for sheet_name in workbook.sheetnames:
        if sheet_name == 'Dados':
            ws = workbook[sheet_name]
            
            # Para cada feriado, marca a coluna inteira
            for data_feriado, nome_feriado in feriados.items():
                if data_feriado in mapa_datas:
                    col_data = mapa_datas[data_feriado]
                    col_idx = list(ws[1]) 
                    
                    # Procura a coluna pela data
                    for col_letter_idx, cell in enumerate(ws[1], 1):
                        if cell.value == col_data or str(cell.value) == str(col_data):
                            # Marca toda a coluna (exceto header) como FERIADO
                            for row_idx in range(2, ws.max_row + 1):
                                cell_data = ws.cell(row=row_idx, column=col_letter_idx)
                                cell_data.value = "FERIADO"
                                # Aplica cor preta com texto branco
                                if 'FERIADO' in mapa_cores:
                                    cell_data.fill = PatternFill(
                                        start_color=mapa_cores['FERIADO'],
                                        end_color=mapa_cores['FERIADO'],
                                        fill_type='solid'
                                    )
                                    cell_data.font = Font(color='FFFFFFFF')  # Texto branco
                            break

def eh_fim_de_semana(data):

    """Retorna True se é sábado (5) ou domingo (6)"""
    return data.weekday() in [5, 6]

def ler_dataframe_do_workbook(workbook):
    """
    Lê o dataframe da sheet 'Dados' do workbook (após marcações de FERIADO e AFASTAMENTO)
    Retorna um pandas DataFrame com os dados atualizados
    """
    ws = workbook['Dados']
    dados = []
    
    # Lê header
    header = []
    for cell in ws[1]:
        header.append(cell.value)
    
    # Lê dados
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, col_name in enumerate(header, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data[col_name] = cell.value
        dados.append(row_data)
    
    return pd.DataFrame(dados)

def calcular_similaridade(s1, s2):
    """Calcula similaridade entre duas strings (0 a 1)"""
    return SequenceMatcher(None, s1, s2).ratio()

def limpar_nome(nome):
    if isinstance(nome, str):
        return unidecode(nome).upper().strip()
    return ""

def extrair_dia_do_cabecalho(label_dia, mes, ano):
    """
    Extrai a data do cabeçalho da coluna, detectando automaticamente o formato.
    Aceita: "01/nov", "01/11", "01", "1/nov", "1/11", etc.
    """
    if pd.isna(label_dia):
        return None
    
    label_str = str(label_dia).strip().lower()
    
    # Mapa de meses em português (tanto nomes curtos quanto abreviações)
    mapa_mes_curto = {'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6, 
                      'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12}
    
    dia_num = None
    mes_encontrado = None
    
    # Formato 1: "DD/mmm" ou "D/mmm" (ex: "01/nov", "1/nov")
    for nome_mes, num_mes in mapa_mes_curto.items():
        if nome_mes in label_str:
            if num_mes == mes:
                mes_encontrado = num_mes
                # Extrair número antes do mês
                parts = label_str.split(nome_mes)
                if parts[0]:
                    try:
                        # Remove tudo que não é número
                        dia_num = int("".join(filter(str.isdigit, parts[0])))
                    except:
                        pass
            if mes_encontrado is not None:
                break
    
    # Formato 2: "DD/MM" ou "D/M" (ex: "01/11", "1/11")
    if mes_encontrado is None:
        # Tenta com separadores comuns: /, -, .
        partes = re.split(r'[/.\-]', label_str.strip())
        if len(partes) >= 2:
            try:
                dia_candidato = int(partes[0].strip())
                mes_candidato = int(partes[1].strip())
                # Valida se é o mês certo e dia válido
                if mes_candidato == mes and 1 <= dia_candidato <= 31:
                    dia_num = dia_candidato
                    mes_encontrado = mes_candidato
            except:
                pass
    
    # Formato 3: "DD" (só o dia, sem separador)
    if mes_encontrado is None:
        try:
            # Se for só número, assume que é o dia
            dia_num = int(label_str.strip())
            if 1 <= dia_num <= 31:
                mes_encontrado = mes
        except:
            pass
    
    # Se encontrou dia e mês válidos, retornar data
    if dia_num is not None and mes_encontrado == mes and 1 <= dia_num <= 31:
        try:
            return datetime.date(ano, mes, dia_num)
        except:
            pass
    
    return None
    
    return None

    return None

def marcar_afastamentos_na_workbook(workbook, mapa_cores, afastamentos=None, df_mest=None, mapa_datas=None):
    """
    Marca células como "Afastamento" onde foi detectado afastamento (>15 FA em sequência).
    Usa os dados de afastamentos detectados pela função detectar_afastamentos_no_dataframe().
    
    Parâmetros:
    - workbook: openpyxl Workbook
    - mapa_cores: dicionário de cores
    - afastamentos: dicionário {index_row: [(col_inicio, col_fim), ...]}
    - df_mest: dataframe original (para mapear índices para linhas)
    - mapa_datas: dicionário de mapeamento de datas
    """
    if not afastamentos or df_mest is None or mapa_datas is None:
        return
    
    ws = workbook['Dados']
    
    # Pega todas as colunas de data em ordem
    colunas_datas = sorted([col for col in df_mest.columns if col in mapa_datas.values()])
    
    # Para cada colaborador com afastamento
    for row_idx_df, sequencias in afastamentos.items():
        # row_idx_df é o índice no dataframe, precisa converter para linha do Excel (row_idx_excel = row_idx_df + 2)
        row_idx_excel = row_idx_df + 2
        
        # Para cada sequência de afastamento detectada
        for col_inicio_idx, col_fim_idx in sequencias:
            # col_inicio_idx e col_fim_idx são índices em colunas_datas
            col_inicio_nome = colunas_datas[col_inicio_idx]
            col_fim_nome = colunas_datas[col_fim_idx]
            
            # Encontra posição dessas colunas no worksheet
            col_inicio_excel = None
            col_fim_excel = None
            
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value == col_inicio_nome:
                    col_inicio_excel = col_idx
                if cell.value == col_fim_nome:
                    col_fim_excel = col_idx
            
            # Marca todas as células nesse intervalo como "Afastamento"
            if col_inicio_excel and col_fim_excel:
                for col_idx in range(col_inicio_excel, col_fim_excel + 1):
                    cell = ws.cell(row=row_idx_excel, column=col_idx)
                    valor_original = str(cell.value).strip().upper() if cell.value else ''
                    
                    # Substitui FA/FI/etc por Afastamento
                    if valor_original not in ['FERIADO', '']:
                        cell.value = 'Afastamento'
                    
                    # Aplica cor
                    if 'Afastamento' in mapa_cores:
                        cell.fill = PatternFill(
                            start_color=mapa_cores['Afastamento'],
                            end_color=mapa_cores['Afastamento'],
                            fill_type='solid'
                        )

def detectar_afastamentos_no_dataframe(df, mapa_datas):
    """
    Detecta colaboradores com sequências contendo > 15 FA (ignorando D, FERIADO e AFASTAMENTO).
    Retorna um dicionário {index_row: [(col_inicio, col_fim), ...]}
    
    Lógica: 
    - Procura por sequências que começam com FA
    - Continua enquanto houver FA, D, FERIADO ou AFASTAMENTO
    - Conta apenas FA (ignora D, FERIADO, AFASTAMENTO)
    - Se total de FA > 15, marca toda a sequência como afastamento
    """
    afastamentos = {}  # {index_row: [(col_inicio, col_fim), ...]}
    
    # Pega todas as colunas de data em ordem
    colunas_datas = sorted([col for col in df.columns if col in mapa_datas.values()])
    
    for idx, row in df.iterrows():
        afastamentos_row = []
        
        i = 0
        while i < len(colunas_datas):
            col = colunas_datas[i]
            valor = str(row[col]).strip().upper() if pd.notna(row[col]) else ''
            
            # Procura por sequências que começam com FA, D, FERIADO ou AFASTAMENTO
            if valor in ['FA', 'D', 'FERIADO', 'AFASTAMENTO']:
                fa_total = 0
                col_inicio = i
                
                # Percorre enquanto tiver FA, D, FERIADO ou AFASTAMENTO (ignora outros valores)
                j = i
                while j < len(colunas_datas):
                    col_j = colunas_datas[j]
                    valor_j = str(row[col_j]).strip().upper() if pd.notna(row[col_j]) else ''
                    
                    if valor_j == 'FA':
                        fa_total += 1
                        j += 1
                    elif valor_j in ['D', 'FERIADO', 'AFASTAMENTO']:
                        # Ignora (pula) mas continua a sequência
                        j += 1
                    else:
                        # Quebra a sequência
                        break
                
                # Se encontrou > 15 FA na sequência, registra
                if fa_total > 15:
                    afastamentos_row.append((col_inicio, j - 1))
                
                i = j if j > i else i + 1
            else:
                i += 1
        
        if afastamentos_row:
            afastamentos[idx] = afastamentos_row
    
    return afastamentos

def criar_sheet_ofensores_abs(df_mest, w, mapa_datas, mapa_cores, afastamentos=None):
    """
    Cria sheet 'Ofensores de ABS' mostrando por GESTOR e TURNO:
    - PERÍODO INTEIRO
    - Semana 1, 2, 3, 4 (dados na mesma sheet)
    
    afastamentos: dicionário com índices de linhas que têm afastamento
    """
    if afastamentos is None:
        afastamentos = {}
    try:
        from openpyxl.styles import Border, Side
        
        # Extrai lista única de gestores
        gestores = df_mest['GESTOR'].dropna().unique()
        gestores = sorted([g for g in gestores if str(g).strip()])
        
        # Colunas de datas no dataframe
        colunas_datas = [col for col in df_mest.columns if col not in ['NOME', 'FUNÇÃO', 'SITUAÇÃO', 'AREA', 'GESTOR', 'SUPERVISOR', 'NOME_LIMPO']]
        
        # Define bordas para as células (MEDIUM para melhor visibilidade)
        border_style = Side(style='medium', color='000000')
        thin_border = Border(
            left=border_style,
            right=border_style,
            top=border_style,
            bottom=border_style
        )
        
        # Cria o sheet
        ws = w.book.create_sheet('Ofensores de ABS', 1)
        
        # Header principal
        titulo_cell = ws['A1']
        titulo_cell.value = '🚨 OFENSORES DE ABSENTEÍSMO POR GESTOR'
        titulo_cell.font = Font(bold=True, size=14, color='FFFFFF')
        titulo_cell.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
        ws.merge_cells('A1:I1')
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        titulo_cell.border = thin_border
        
        # Agrupa datas por período com labels de data (segunda a domingo, ou dias restantes)
        datas_obj = sorted([d for d in mapa_datas.keys() if isinstance(d, datetime.date)])
        periodos_dict = {}  # {label: [colunas_datas], ...}
        
        if datas_obj:
            import calendar
            
            ano_dados = datas_obj[0].year
            mes_dados = datas_obj[0].month
            
            # monthcalendar retorna semanas (segunda a domingo)
            cal = calendar.monthcalendar(ano_dados, mes_dados)
            
            # Processa cada semana do calendário
            periodo_num = 1
            for semana_dias in cal:
                # Filtra apenas dias que existem em nosso dataset
                dias_na_semana = [dia for dia in semana_dias if dia != 0]
                
                # Encontra quais datas do nosso dataset estão nesta semana
                datas_nesta_semana = [d for d in datas_obj if d.day in dias_na_semana]
                
                if datas_nesta_semana:
                    # Cria label com as datas (exemplo: "3/11 a 8/11")
                    data_inicio = min(datas_nesta_semana)
                    data_fim = max(datas_nesta_semana)
                    
                    label = f"{data_inicio.day}/{data_inicio.month:02d} a {data_fim.day}/{data_fim.month:02d}"
                    
                    # Adiciona colunas de data neste período
                    periodos_dict[label] = [mapa_datas[d] for d in sorted(datas_nesta_semana)]
                    periodo_num += 1
        
        # Função para processar análise
        def processar_analise(colunas_processar):
            dados_gestores = []
            
            for gestor in gestores:
                colaboradores_gestor = df_mest[df_mest['GESTOR'] == gestor]
                total_colab = len(colaboradores_gestor)
                
                # Encontra o turno do GESTOR: pega o TURNO mais frequente dos colaboradores deste gestor
                gestor_turno = 'N/A'
                if 'TURNO' in df_mest.columns and len(colaboradores_gestor) > 0:
                    # Extrai todos os turnos dos colaboradores deste gestor
                    turnos = colaboradores_gestor['TURNO'].dropna()
                    turnos = [str(t).strip() for t in turnos if str(t).strip()]
                    
                    if turnos:
                        # Pega o turno mais frequente (moda)
                        from collections import Counter
                        turno_counts = Counter(turnos)
                        gestor_turno = turno_counts.most_common(1)[0][0]
                    else:
                        gestor_turno = 'N/A'
                
                total_fi = 0
                total_fa = 0
                
                for idx, row in colaboradores_gestor.iterrows():
                    # Verifica se o colaborador tem afastamento (já detectado antes)
                    tem_afastamento = idx in afastamentos
                    
                    for col_data in colunas_processar:
                        if col_data not in df_mest.columns:
                            continue
                        
                        valor = str(row[col_data]).strip().upper() if pd.notna(row[col_data]) else ''
                        
                        # Ignora FERIADO nas contagens
                        if valor == 'FERIADO':
                            continue
                        
                        if valor == 'FI':
                            total_fi += 1
                        elif valor == 'FA':
                            # Só conta FA se NÃO for afastamento
                            if not tem_afastamento:
                                total_fa += 1
                
                total_faltas = total_fi + total_fa
                dias_uteis = len(colunas_processar)
                percentual = (total_faltas / dias_uteis / total_colab * 100) if total_colab > 0 and dias_uteis > 0 else 0
                
                # OPÇÃO 1: % Colaboradores com faltas (Número de colaboradores que tiveram pelo menos 1 falta)
                colab_com_faltas = 0
                for idx, row in colaboradores_gestor.iterrows():
                    tem_afastamento = idx in afastamentos
                    tem_falta = False
                    
                    for col_data in colunas_processar:
                        if col_data not in df_mest.columns:
                            continue
                        valor = str(row[col_data]).strip().upper() if pd.notna(row[col_data]) else ''
                        
                        if valor == 'FERIADO':
                            continue
                        if valor == 'FI' or (valor == 'FA' and not tem_afastamento):
                            tem_falta = True
                            break
                    
                    if tem_falta:
                        colab_com_faltas += 1
                
                # Porcentagem de colaboradores com faltas
                pct_colab_com_faltas = (colab_com_faltas / total_colab * 100) if total_colab > 0 else 0
                
                # OPÇÃO 2: Média de faltas por colaborador (Total Faltas / Total Colaboradores)
                media_faltas_por_colab = (total_faltas / total_colab) if total_colab > 0 else 0
                
                # ÍNDICE DE CONCENTRAÇÃO: Mede o quanto as faltas estão concentradas em poucas pessoas
                # Usa o Índice de Gini (0-100): 0 = distribuído entre todos, 100 = concentrado em 1 pessoa
                indice_concentracao = 0
                if total_faltas > 0:
                    # Conta faltas de cada colaborador
                    faltas_por_colab = {}
                    for idx, row in colaboradores_gestor.iterrows():
                        tem_afastamento = idx in afastamentos
                        faltas_pessoa = 0
                        
                        for col_data in colunas_processar:
                            if col_data not in df_mest.columns:
                                continue
                            valor = str(row[col_data]).strip().upper() if pd.notna(row[col_data]) else ''
                            
                            if valor == 'FERIADO':
                                continue
                            if valor == 'FI':
                                faltas_pessoa += 1
                            elif valor == 'FA' and not tem_afastamento:
                                faltas_pessoa += 1
                        
                        if faltas_pessoa > 0:
                            faltas_por_colab[idx] = faltas_pessoa
                    
                    # Calcula Índice de Gini simplificado (0-100)
                    if len(faltas_por_colab) > 0:
                        faltas_sorted = sorted(faltas_por_colab.values(), reverse=True)
                        cumsum = 0
                        sum_desvios = 0
                        
                        for i, faltas in enumerate(faltas_sorted):
                            cumsum += faltas
                            # Quanto mais distante da média, maior o desvio
                            sum_desvios += abs(cumsum - (i + 1) * (total_faltas / len(faltas_sorted)))
                        
                        # Normaliza para 0-100
                        indice_concentracao = min(100, (sum_desvios / total_faltas * 10) if total_faltas > 0 else 0)
                
                if percentual > 20:
                    status = '🔴 CRÍTICO'
                    status_color = 'FFFF0000'
                elif percentual > 10:
                    status = '🟡 ATENÇÃO'
                    status_color = 'FFFFFF00'
                else:
                    status = '🟢 OK'
                    status_color = 'FF00B050'
                
                dados_gestores.append({
                    'gestor': gestor,
                    'turno': gestor_turno,
                    'total_colab': total_colab,
                    'total_fi': total_fi,
                    'total_fa': total_fa,
                    'total_faltas': total_faltas,
                    'percentual': percentual,
                    'pct_colab_com_faltas': pct_colab_com_faltas,
                    'colab_com_faltas': colab_com_faltas,
                    'media_faltas_por_colab': media_faltas_por_colab,
                    'indice_concentracao': indice_concentracao,
                    'status': status,
                    'status_color': status_color
                })
            
            # Ordena por % de colaboradores com faltas (descendente) - maiores porcentagens primeiro
            dados_gestores.sort(key=lambda x: x['pct_colab_com_faltas'], reverse=True)
            return dados_gestores
        
        # PERÍODO INTEIRO
        dados_periodo = processar_analise(colunas_datas)
        
        # PERÍODOS (com labels de datas)
        dados_periodos = {}
        for label, colunas_periodo in periodos_dict.items():
            dados_periodos[label] = processar_analise(colunas_periodo)
        
        # Preenche o sheet com PERÍODO + PERÍODOS
        row_idx = 3
        
        # Título PERÍODO INTEIRO
        ws.cell(row=row_idx, column=1, value='PERÍODO INTEIRO DE (MÊS)')
        ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
        ws.merge_cells(f'A{row_idx}:I{row_idx}')
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left')
        ws.cell(row=row_idx, column=1).border = thin_border
        row_idx += 1
        
        # Headers
        headers = ['GESTOR', 'TURNO', 'Total de Colaboradores', 'Com Faltas (FI)', 'Com Faltas (FA)', 'Total de Faltas', '% Colab. com Faltas', 'Com Faltas (X/Y)', 'Índice Concentração']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')  # Verde escuro corporativo
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        row_idx += 1
        
        # Dados do período
        for dado in dados_periodo:
            values = [dado['gestor'], dado['turno'], dado['total_colab'], dado['total_fi'], dado['total_fa'], dado['total_faltas']]
            
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = thin_border
                
                # Cores - Paleta Corporativa Profarma
                if col_idx == 1:  # GESTOR - Verde escuro
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                elif col_idx == 2:  # TURNO - Cinza claro
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                elif col_idx == 3:  # Total de Colaboradores - Cinza claro
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                elif col_idx == 4:  # FI - Verde médio
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFFFF')
                elif col_idx == 5:  # FA - Verde claro
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFFFF')
                elif col_idx == 6:  # TOTAL - Verde escuro com texto branco
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFFFF')
            
            # Coluna 7: % Colaboradores com Faltas (fórmula)
            cell_pct_colab = ws.cell(row=row_idx, column=7)
            cell_pct_colab.value = dado['pct_colab_com_faltas']
            cell_pct_colab.number_format = '0.00"%"'
            cell_pct_colab.alignment = Alignment(horizontal='center', vertical='center')
            cell_pct_colab.fill = PatternFill(start_color='FF8CC850', end_color='FF8CC850', fill_type='solid')  # Verde light
            cell_pct_colab.font = Font(bold=True, color='FF000000')
            cell_pct_colab.border = thin_border
            
            # Coluna 8: X/Y Colaboradores com Faltas
            cell_colab_ratio = ws.cell(row=row_idx, column=8)
            cell_colab_ratio.value = f"{dado['colab_com_faltas']}/{dado['total_colab']}"
            cell_colab_ratio.alignment = Alignment(horizontal='center', vertical='center')
            cell_colab_ratio.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            cell_colab_ratio.font = Font(bold=True)
            cell_colab_ratio.border = thin_border
            
            # Coluna 9: Índice de Concentração (0-100)
            cell_indice = ws.cell(row=row_idx, column=9)
            cell_indice.value = round(dado['indice_concentracao'], 1)
            cell_indice.number_format = '0.0'
            cell_indice.alignment = Alignment(horizontal='center', vertical='center')
            cell_indice.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            cell_indice.font = Font(bold=True)
            cell_indice.border = thin_border
            
            row_idx += 1
        
        # PERÍODOS (com labels de data)
        for label, dados_periodo_especifico in dados_periodos.items():
            row_idx += 1
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
            ws.merge_cells(f'A{row_idx}:I{row_idx}')
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left')
            ws.cell(row=row_idx, column=1).border = thin_border
            row_idx += 1
            
            # Headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color='FFFFFF', size=11)
                cell.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')  # Verde escuro corporativo
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            row_idx += 1
            
            # Dados do período
            for dado in dados_periodo_especifico:
                values = [dado['gestor'], dado['turno'], dado['total_colab'], dado['total_fi'], dado['total_fa'], dado['total_faltas']]
                
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = thin_border
                    
                    # Cores corporativas Profarma
                    if col_idx == 1:  # GESTOR - cinza claro
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    elif col_idx == 2:  # TURNO - cinza claro
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    elif col_idx == 3:  # Total de Colaboradores - cinza claro
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    elif col_idx == 4:  # FI - verde médio
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFFFF')
                    elif col_idx == 5:  # FA - verde claro
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFFFF')
                    elif col_idx == 6:  # TOTAL - verde escuro
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFFFF')
                
                # Coluna 7: % Colaboradores com Faltas - verde light
                cell_pct_colab = ws.cell(row=row_idx, column=7)
                cell_pct_colab.value = dado['pct_colab_com_faltas']
                cell_pct_colab.number_format = '0.00"%"'
                cell_pct_colab.alignment = Alignment(horizontal='center', vertical='center')
                cell_pct_colab.fill = PatternFill(start_color='FF8CC850', end_color='FF8CC850', fill_type='solid')
                cell_pct_colab.font = Font(bold=True, color='FF000000')
                cell_pct_colab.border = thin_border
                
                # Coluna 8: X/Y Colaboradores com Faltas - cinza claro
                cell_colab_ratio = ws.cell(row=row_idx, column=8)
                cell_colab_ratio.value = f"{dado['colab_com_faltas']}/{dado['total_colab']}"
                cell_colab_ratio.alignment = Alignment(horizontal='center', vertical='center')
                cell_colab_ratio.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                cell_colab_ratio.font = Font(bold=True)
                cell_colab_ratio.border = thin_border
                
                # Coluna 9: Índice de Concentração (0-100) - cinza claro
                cell_indice = ws.cell(row=row_idx, column=9)
                cell_indice.value = round(dado['indice_concentracao'], 1)
                cell_indice.number_format = '0.0'
                cell_indice.alignment = Alignment(horizontal='center', vertical='center')
                cell_indice.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                cell_indice.font = Font(bold=True)
                cell_indice.border = thin_border
                
                row_idx += 1
        
        # Ajusta largura das colunas (A 30% maior)
        ws.column_dimensions['A'].width = 25 * 1.3  # 30% maior
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 18  # % Colab. com Faltas
        ws.column_dimensions['H'].width = 14  # Com Faltas (X/Y)
        ws.column_dimensions['I'].width = 18  # Índice Concentração
        
        return True
    except Exception as e:
        st.error(f"Erro ao criar sheet de ofensores: {str(e)}")
        import traceback
        st.write(traceback.format_exc())
        return False

def criar_sheet_ranking_abs(df_mest, w, mapa_colors, top10_fa_enriquecido=None, top10_fi_enriquecido=None):
    """
    Cria sheet 'Ranking ABS' com TOP 10 colaboradores com mais FA e TOP 10 com mais FI
    
    Args:
        df_mest: DataFrame da planilha mestra
        w: Workbook
        mapa_colors: Dicionário de cores
        top10_fa_enriquecido: TOP 10 FA com dados enriquecidos (opcional)
        top10_fi_enriquecido: TOP 10 FI com dados enriquecidos (opcional)
    """
    try:
        from openpyxl.styles import Border, Side
        
        # Colunas de datas no dataframe (contêm dados de FA/FI)
        colunas_datas = [col for col in df_mest.columns if col not in ['NOME', 'FUNÇÃO', 'SITUAÇÃO', 'AREA', 'GESTOR', 'SUPERVISOR', 'NOME_LIMPO']]
        
        # Conta FA e FI para cada colaborador
        df_ranking = pd.DataFrame({
            'NOME': df_mest['NOME'],
            'GESTOR': df_mest['GESTOR'],
            'FUNÇÃO': df_mest['FUNÇÃO'],
            'AREA': df_mest['AREA'],
            'FI': df_mest[colunas_datas].apply(lambda row: (row == 'FI').sum(), axis=1),
            'FA': df_mest[colunas_datas].apply(lambda row: (row == 'FA').sum(), axis=1),
        }).copy()
        
        # Remove registros vazios
        df_ranking = df_ranking[df_ranking['NOME'].notna() & (df_ranking['NOME'] != '')]
        
        # TOP 10 FA e FI
        top10_fa = df_ranking.nlargest(10, 'FA')
        top10_fi = df_ranking.nlargest(10, 'FI')
        
        # Se foram passados dados enriquecidos, use-os
        if top10_fa_enriquecido is not None:
            top10_fa = top10_fa_enriquecido
        if top10_fi_enriquecido is not None:
            top10_fi = top10_fi_enriquecido
        
        # Cria o sheet
        ws = w.book.create_sheet('Ranking ABS')
        
        # Define bordas
        border_style = Side(style='medium', color='000000')
        thin_border = Border(
            left=border_style,
            right=border_style,
            top=border_style,
            bottom=border_style
        )
        
        row_idx = 1
        
        # Título geral
        ws.merge_cells('A1:I1')
        title_cell = ws.cell(row=row_idx, column=1, value='🏆 RANKING DE ABSENTEÍSMO')
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row_idx].height = 25
        row_idx += 2
        
        # ===== TOP 10 FA =====
        ws.merge_cells(f'A{row_idx}:I{row_idx}')
        fa_header = ws.cell(row=row_idx, column=1, value='TOP 10 - FALTAS POR ATESTADO (FA)')
        fa_header.font = Font(bold=True, size=12, color='FFFFFFFF')
        fa_header.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
        fa_header.alignment = Alignment(horizontal='center', vertical='center')
        row_idx += 1
        
        # Headers FA
        headers_fa = ['Posição', 'Nome', 'Gestor', 'Função', 'Área', 'FA', 'Data Admissão', 'Tempo de Serviço', 'Gênero']
        for col_idx, header in enumerate(headers_fa, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        row_idx += 1
        
        # Dados TOP 10 FA
        for idx, (_, row) in enumerate(top10_fa.iterrows(), 1):
            # Posição
            cell_pos = ws.cell(row=row_idx, column=1, value=idx)
            cell_pos.border = thin_border
            cell_pos.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
            cell_pos.font = Font(bold=True, color='FFFFFFFF')
            cell_pos.alignment = Alignment(horizontal='center', vertical='center')
            
            # Nome
            cell_nome = ws.cell(row=row_idx, column=2, value=row['NOME'])
            cell_nome.border = thin_border
            cell_nome.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            
            # Gestor
            cell_gestor = ws.cell(row=row_idx, column=3, value=row['GESTOR'])
            cell_gestor.border = thin_border
            cell_gestor.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            
            # Função
            cell_func = ws.cell(row=row_idx, column=4, value=row['FUNÇÃO'])
            cell_func.border = thin_border
            cell_func.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            
            # Área
            cell_area = ws.cell(row=row_idx, column=5, value=row['AREA'])
            cell_area.border = thin_border
            cell_area.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            
            # FA
            cell_fa = ws.cell(row=row_idx, column=6, value=row['FA'])
            cell_fa.border = thin_border
            cell_fa.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
            cell_fa.font = Font(bold=True, color='FFFFFFFF')
            cell_fa.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data Admissão
            data_adm = row.get('Data Admissão', 'N/A') if 'Data Admissão' in row.index else 'N/A'
            cell_data = ws.cell(row=row_idx, column=7, value=data_adm)
            cell_data.border = thin_border
            cell_data.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            cell_data.alignment = Alignment(horizontal='center', vertical='center')
            
            # Tempo de Serviço
            tempo_srv = row.get('Tempo de Serviço', 'N/A') if 'Tempo de Serviço' in row.index else 'N/A'
            cell_tempo = ws.cell(row=row_idx, column=8, value=tempo_srv)
            cell_tempo.border = thin_border
            cell_tempo.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            cell_tempo.alignment = Alignment(horizontal='center', vertical='center')
            
            # Gênero
            genero = row.get('Gênero', 'N/A') if 'Gênero' in row.index else 'N/A'
            cell_genero = ws.cell(row=row_idx, column=9, value=genero)
            cell_genero.border = thin_border
            cell_genero.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            cell_genero.alignment = Alignment(horizontal='center', vertical='center')
            
            row_idx += 1
        
        row_idx += 2
        
        # ===== TOP 10 FI =====
        ws.merge_cells(f'A{row_idx}:I{row_idx}')
        fi_header = ws.cell(row=row_idx, column=1, value='TOP 10 - FALTAS INJUSTIFICADAS (FI)')
        fi_header.font = Font(bold=True, size=12, color='FFFFFFFF')
        fi_header.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
        fi_header.alignment = Alignment(horizontal='center', vertical='center')
        row_idx += 1
        
        # Headers FI
        headers_fi = ['Posição', 'Nome', 'Gestor', 'Função', 'Área', 'FI', 'Data Admissão', 'Tempo de Serviço', 'Gênero']
        for col_idx, header in enumerate(headers_fi, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        row_idx += 1
        
        # Dados TOP 10 FI
        for idx, (_, row) in enumerate(top10_fi.iterrows(), 1):
            # Posição
            cell_pos = ws.cell(row=row_idx, column=1, value=idx)
            cell_pos.border = thin_border
            cell_pos.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
            cell_pos.font = Font(bold=True, color='FFFFFFFF')
            cell_pos.alignment = Alignment(horizontal='center', vertical='center')
            
            # Nome
            cell_nome = ws.cell(row=row_idx, column=2, value=row['NOME'])
            cell_nome.border = thin_border
            cell_nome.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            
            # Gestor
            cell_gestor = ws.cell(row=row_idx, column=3, value=row['GESTOR'])
            cell_gestor.border = thin_border
            cell_gestor.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            
            # Função
            cell_func = ws.cell(row=row_idx, column=4, value=row['FUNÇÃO'])
            cell_func.border = thin_border
            cell_func.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            
            # Área
            cell_area = ws.cell(row=row_idx, column=5, value=row['AREA'])
            cell_area.border = thin_border
            cell_area.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            
            # FI
            cell_fi = ws.cell(row=row_idx, column=6, value=row['FI'])
            cell_fi.border = thin_border
            cell_fi.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
            cell_fi.font = Font(bold=True, color='FFFFFFFF')
            cell_fi.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data Admissão
            data_adm = row.get('Data Admissão', 'N/A') if 'Data Admissão' in row.index else 'N/A'
            cell_data = ws.cell(row=row_idx, column=7, value=data_adm)
            cell_data.border = thin_border
            cell_data.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            cell_data.alignment = Alignment(horizontal='center', vertical='center')
            
            # Tempo de Serviço
            tempo_srv = row.get('Tempo de Serviço', 'N/A') if 'Tempo de Serviço' in row.index else 'N/A'
            cell_tempo = ws.cell(row=row_idx, column=8, value=tempo_srv)
            cell_tempo.border = thin_border
            cell_tempo.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            cell_tempo.alignment = Alignment(horizontal='center', vertical='center')
            
            # Gênero
            genero = row.get('Gênero', 'N/A') if 'Gênero' in row.index else 'N/A'
            cell_genero = ws.cell(row=row_idx, column=9, value=genero)
            cell_genero.border = thin_border
            cell_genero.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
            cell_genero.alignment = Alignment(horizontal='center', vertical='center')
            
            row_idx += 1
        
        # Ajusta largura das colunas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 42
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 38
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15  # Data Admissão
        ws.column_dimensions['H'].width = 18  # Tempo de Serviço
        ws.column_dimensions['I'].width = 14  # Gênero
        
        return (top10_fa, top10_fi)
    except Exception as e:
        st.error(f"Erro ao criar sheet de ranking: {str(e)}")
        import traceback
        st.write(traceback.format_exc())
        return False


def enriquecer_ranking_com_dados_csv(top_10_fa, top_10_fi, df_colaboradores):
    """
    Enriquece os TOP 10 FA e FI com dados do CSV de colaboradores.
    Usa fuzzy matching (LIKE) para encontrar nomes mesmo com pequenas diferenças.
    
    Extrai:
    - Data de Admissão (coluna "Data Admissão")
    - Gênero (coluna "Sexo")
    - Calcula Tempo de Serviço em Anos e Meses
    
    Args:
        top_10_fa: DataFrame com TOP 10 FA
        top_10_fi: DataFrame com TOP 10 FI
        df_colaboradores: DataFrame com dados dos colaboradores (CSV)
    
    Returns:
        tuple: (df_fa_enriquecido, df_fi_enriquecido)
    """
    from difflib import SequenceMatcher
    
    def calcular_tempo_admissao(data_admissao):
        """Calcula Anos e Meses desde a data de admissão (formato: XaYm)"""
        try:
            if pd.isna(data_admissao):
                return "N/A"
            
            # Converte para datetime se necessário
            if isinstance(data_admissao, str):
                # Tenta diferentes formatos
                try:
                    data = pd.to_datetime(data_admissao, format='%d/%m/%Y')
                except:
                    try:
                        data = pd.to_datetime(data_admissao, format='%Y-%m-%d')
                    except:
                        return "N/A"
            else:
                data = pd.to_datetime(data_admissao)
            
            hoje = datetime.datetime.now()
            
            # Calcula diferença
            diff = relativedelta(hoje, data)
            anos = diff.years
            meses = diff.months
            
            return f"{anos}a {meses}m"
        except:
            return "N/A"
    
    def similarity_ratio(a, b):
        """Calcula o índice de similaridade entre duas strings (0 a 1)"""
        return SequenceMatcher(None, a.upper().strip(), b.upper().strip()).ratio()
    
    # Fazer merge dos DataFrames
    df_fa_enriquecido = top_10_fa.copy()
    df_fi_enriquecido = top_10_fi.copy()
    
    # Procurar colunas pelo padrão correto
    # No seu CSV: coluna 4 (índice 3) é "Colaborador", coluna 13 (índice 12) é "Data Admissão", coluna 51 (índice 50) é "Sexo"
    
    col_nome_csv = None
    col_data_adm = None
    col_sexo = None
    
    # Tenta encontrar by exact match na lista de colunas
    for i, col in enumerate(df_colaboradores.columns):
        col_upper = col.upper().strip()
        
        # Coluna de Colaborador
        if col_upper == 'COLABORADOR':
            col_nome_csv = col
        
        # Coluna de Data Admissão
        if 'DATA ADMISS' in col_upper:
            col_data_adm = col
        
        # Coluna de Sexo
        if col_upper == 'SEXO':
            col_sexo = col
    
    # Fallback: se não encontrou, usa índices conhecidos
    if col_nome_csv is None and len(df_colaboradores.columns) > 3:
        col_nome_csv = df_colaboradores.columns[3]
    
    if col_data_adm is None and len(df_colaboradores.columns) > 12:
        col_data_adm = df_colaboradores.columns[12]
    
    if col_sexo is None and len(df_colaboradores.columns) > 50:
        col_sexo = df_colaboradores.columns[50]
    
    # Função para buscar dados do colaborador com fuzzy matching
    def buscar_dados_colaborador(nome_ranking):
        if not col_nome_csv:
            return {
                'Data Admissão': 'Não consta',
                'Tempo de Serviço': 'Não consta',
                'Gênero': 'Não consta'
            }
        
        nome_ranking_upper = str(nome_ranking).strip().upper()
        
        # Busca por correspondência exata primeiro
        matches_exatos = [
            idx for idx, row in df_colaboradores.iterrows()
            if str(row[col_nome_csv]).strip().upper() == nome_ranking_upper
        ]
        
        if matches_exatos:
            idx_match = matches_exatos[0]
        else:
            # Busca por fuzzy matching (similaridade >= 0.75)
            melhor_idx = None
            melhor_score = 0
            
            for idx, row in df_colaboradores.iterrows():
                nome_csv = str(row[col_nome_csv]).strip()
                score = similarity_ratio(nome_ranking_upper, nome_csv)
                
                if score > melhor_score:
                    melhor_score = score
                    melhor_idx = idx
            
            # Só aceita se a similaridade for >= 75%
            if melhor_score >= 0.75:
                idx_match = melhor_idx
            else:
                idx_match = None
        
        if idx_match is not None:
            row_match = df_colaboradores.iloc[idx_match]
            
            # Extrai dados com segurança
            data_adm = 'Não consta'
            sexo = 'Não consta'
            tempo_servico = 'Não consta'
            
            try:
                if col_data_adm and col_data_adm in df_colaboradores.columns:
                    data_adm_val = row_match[col_data_adm]
                    if pd.notna(data_adm_val):
                        data_adm = str(data_adm_val).strip()
                        tempo_servico = calcular_tempo_admissao(data_adm_val)
            except:
                pass
            
            try:
                if col_sexo and col_sexo in df_colaboradores.columns:
                    sexo_val = row_match[col_sexo]
                    if pd.notna(sexo_val):
                        sexo = str(sexo_val).strip()
            except:
                pass
            
            return {
                'Data Admissão': data_adm,
                'Tempo de Serviço': tempo_servico,
                'Gênero': sexo
            }
        else:
            return {
                'Data Admissão': 'Não consta',
                'Tempo de Serviço': 'Não consta',
                'Gênero': 'Não consta'
            }
    
    # Aplicar busca para FA
    for idx, row in df_fa_enriquecido.iterrows():
        dados = buscar_dados_colaborador(row['NOME'])
        for col, val in dados.items():
            df_fa_enriquecido.at[idx, col] = val
    
    # Aplicar busca para FI
    for idx, row in df_fi_enriquecido.iterrows():
        dados = buscar_dados_colaborador(row['NOME'])
        for col, val in dados.items():
            df_fi_enriquecido.at[idx, col] = val
    
    return df_fa_enriquecido, df_fi_enriquecido


def colorir_celulas_incomuns_dados(w, MAPA_CORES, mapa_datas):
    """
    Pinta com cinza apenas as CÉLULAS DE DATAS na planilha Dados que contêm marcadores incomuns.
    Marcadores "comuns" são: P, FI, FA, FÉRIAS-BH, DESLIGADO, FERIADO, DOMINGO
    Qualquer outro valor (exceto vazio) será marcado com cinza fundo + texto preto.
    
    Args:
        w: Workbook wrapper object
        MAPA_CORES: Dicionário de cores
        mapa_datas: Dicionário de mapeamento de datas para colunas
    """
    try:
        # Lista de códigos "comuns" que NÃO devem ser marcados
        codigos_comuns = {'P', 'FI', 'FA', 'FÉRIAS-BH', 'DESLIGADO', 'FERIADO', 'DOMINGO', ''}
        
        # Cor cinza para células incomuns
        gray_fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
        black_font = Font(color='FF000000', bold=False)
        
        # Tenta encontrar a sheet com nome "Dados"
        ws_dados = None
        for sheet_name in w.book.sheetnames:
            if sheet_name.upper() == 'DADOS':
                ws_dados = w.book[sheet_name]
                break
        
        if not ws_dados:
            print("Sheet 'Dados' não encontrada")
            return False
        
        # Obtém lista de nomes de colunas de datas a partir de mapa_datas
        # mapa_datas.values() contém os nomes das colunas de datas
        colunas_data_nomes = set(mapa_datas.values())
        
        # Lê o header para identificar qual coluna é qual
        header = []
        for cell in ws_dados[1]:
            header.append(cell.value)
        
        # Identifica os índices das colunas de datas (usando mapa_datas como fonte de verdade)
        colunas_data_indices = []
        for col_idx, col_name in enumerate(header, 1):
            if col_name in colunas_data_nomes:
                colunas_data_indices.append(col_idx)
        
        # Percorre apenas as LINHAS DE DADOS nas COLUNAS DE DATAS
        for row_idx in range(2, ws_dados.max_row + 1):
            for col_idx in colunas_data_indices:
                cell = ws_dados.cell(row=row_idx, column=col_idx)
                cell_value = str(cell.value).strip() if cell.value is not None else ''
                
                # Se o valor não está vazio E não está na lista de códigos comuns
                if cell_value and cell_value not in codigos_comuns:
                    # Aplica cor cinza E texto preto
                    cell.fill = gray_fill
                    cell.font = black_font
        
        return True
        
    except Exception as e:
        print(f"Erro ao colorir células incomuns: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return False

st.set_page_config(layout="wide", initial_sidebar_state="collapsed")

# CSS para expandir containers em full width
st.markdown("""
<style>
    .main > .block-container {
        padding: 1rem;
        max-width: 100%;
    }
    .stContainer {
        max-width: 100%;
    }
    /* Fix para columns ocuparem full width */
    [data-testid="column"] {
        flex: 1 1 calc(100% - 1rem) !important;
    }
    .st-emotion-cache-9edo8l.e1wguzas2 {
        flex: 1 1 calc(100% - 1rem) !important;
    }
</style>
""", unsafe_allow_html=True)

# Inicializa session_state globalmente
if 'idx_arquivo_nav' not in st.session_state:
    st.session_state.idx_arquivo_nav = 0
if 'config_arquivos' not in st.session_state:
    st.session_state.config_arquivos = {}

st.title("🤖 Lançamento de Absenteísmo")
st.write("Com busca LIKE (aproximada) para nomes")

MAPA_CODIGOS = {1: 'P', 2: 'FI', 4: 'FA', 3: 'FÉRIAS-BH', 5: 'DESLIGADO'}

MAPA_CORES = {
    'P': 'FF90EE90',      # Verde claro
    'FI': 'FFFF0000',     # Vermelho puro (mais nítido)
    'FA': 'FFFFFF00',     # Amarelo puro (mais nítido)
    'Afastamento': 'FFC0C0C0',  # Cinza (mesma cor de D)
    'FERIADO': 'FF000000',      # Preto (com texto branco)
    'FÉRIAS-BH': 'FF000000',    # Preto (com texto branco)
    'DESLIGADO': 'FF800080',   # Roxo
    'DESCANSO': 'FFC0C0C0'  # Cinza
}

col1, col2 = st.columns(2)

with col1:
    st.header("Upload")
    file_mestra = st.file_uploader("Planilha MESTRA", type=["xlsx"])
    file_colaboradores = st.file_uploader("CSV de Colaboradores (para enriquecer Ranking)", type=["csv", "xlsx"])
    files_encarregado = st.file_uploader("Planilhas ENCARREGADO (múltiplas permitidas)", type=["xlsx"], accept_multiple_files=True)

with col2:
    st.header("Config")
    ano = st.number_input("Ano", 2020, 2050, datetime.date.today().year)
    mes = st.number_input("Mês", 1, 12, datetime.date.today().month)

# Valida arquivos de encarregado
arquivos_invalidos = []
arquivos_validos = []

if files_encarregado:
    st.divider()
    
    # Valida cada arquivo
    for file_enc in files_encarregado:
        try:
            # Tenta detectar sheets com múltiplos engines
            guias = None
            try:
                guias = pd.ExcelFile(io.BytesIO(file_enc.getvalue()), engine='openpyxl').sheet_names
            except:
                try:
                    file_enc.seek(0)
                    guias = pd.ExcelFile(io.BytesIO(file_enc.getvalue())).sheet_names
                except:
                    guias = None
            
            if guias:
                arquivos_validos.append(file_enc)
            else:
                arquivos_invalidos.append(file_enc.name)
        except Exception as e:
            arquivos_invalidos.append(f"{file_enc.name} (Erro: {str(e)[:50]}...)")
    
    # Mostra avisos dos arquivos inválidos
    if arquivos_invalidos:
        st.warning(f"⚠️ **{len(arquivos_invalidos)} arquivo(s) inválido(s) ou confidencial(is)**:")
        for arquivo_invalido in arquivos_invalidos:
            st.error(f"❌ {arquivo_invalido}")
        st.info("💡 **Dica:** Remova esses arquivos ou salve-os como novos arquivos sem proteção/confidencialidade")
    
    # Continua com arquivos válidos apenas
    files_encarregado = arquivos_validos
    
    if not files_encarregado:
        st.error("❌ Nenhum arquivo válido encontrado! Por favor, envie arquivos Excel válidos.")
        st.stop()
    
    st.header("Pré-Visualização")
    
    # Se há apenas 1 arquivo, processa normalmente
    # Se há múltiplos, mostra navegação
    if len(files_encarregado) == 1:
        file_encarregado = files_encarregado[0]
        idx_arquivo_atual = 0
    else:
        col_prev, col_info, col_next = st.columns([1, 3, 1])
        
        with col_prev:
            if st.button("⬅️ Anterior", key="btn_prev_arquivo"):
                st.session_state.idx_arquivo_nav = max(0, st.session_state.idx_arquivo_nav - 1)
                st.rerun()
        
        with col_info:
            nomes_arquivos = [f.name for f in files_encarregado]
            idx_arq = st.session_state.idx_arquivo_nav
            # Mostra se está configurado
            status = "✅" if nomes_arquivos[idx_arq] in st.session_state.config_arquivos else "⚠️"
            st.info(f"{status} {nomes_arquivos[idx_arq]} ({idx_arq + 1}/{len(files_encarregado)})")
        
        with col_next:
            if st.button("Próximo ➡️", key="btn_next_arquivo"):
                st.session_state.idx_arquivo_nav = min(len(files_encarregado) - 1, st.session_state.idx_arquivo_nav + 1)
                st.rerun()
        
        idx_arquivo_atual = st.session_state.idx_arquivo_nav
        file_encarregado = files_encarregado[idx_arquivo_atual]
    
    # Detecta as guias (sheets) disponíveis no arquivo
    guias_disponiveis = pd.ExcelFile(io.BytesIO(file_encarregado.getvalue())).sheet_names
    
    # Detecta qual é a guia ATIVA no arquivo Excel
    wb_temp = load_workbook(io.BytesIO(file_encarregado.getvalue()), data_only=True)
    guia_ativa_arquivo = wb_temp.active.title  # Pega o título da guia ativa
    wb_temp.close()
    
    # Define a guia ativa do arquivo como padrão
    if guia_ativa_arquivo in guias_disponiveis:
        default_guia = guia_ativa_arquivo
    else:
        default_guia = guias_disponiveis[0]
    
    # Se há múltiplas guias, deixa o usuário escolher
    if len(guias_disponiveis) > 1:
        guia_selecionada = st.selectbox("Selecione a guia:", guias_disponiveis, index=guias_disponiveis.index(default_guia), key=f"guia_{idx_arquivo_atual}")
    else:
        guia_selecionada = guias_disponiveis[0]
    
    # Guarda a guia selecionada no session_state
    st.session_state.guia_selecionada = guia_selecionada
    buf = io.BytesIO(file_encarregado.getvalue())
    df_raw = pd.read_excel(buf, sheet_name=guia_selecionada, header=None, dtype=str)  # Especifica dtype direto
    
    st.write(f"**Linhas detectadas:** {len(df_raw)} | **Colunas:** {len(df_raw.columns)}")
    
    # Cria nomes em formato Excel (A, B, C...)
    letras_disponíveis = []
    for i in range(len(df_raw.columns)):
        if i < 26:
            letras_disponíveis.append(chr(65 + i))  # A-Z
        else:
            letras_disponíveis.append(f"{chr(65 + i//26 - 1)}{chr(65 + i%26)}")  # AA, AB, etc
    
    # DETECTA AUTOMATICAMENTE qual é a coluna com nomes (testando múltiplas linhas)
    keywords_nomes = ['NOME', 'NOMES', 'COLABORADOR', 'COLABORADORES', 'FUNCIONARIO', 'FUNCIONARIOS', 'EMPLOYEE', 'EMPLOYEES', 'PESSOAL', 'PERSON', 'STAFF']
    col_detectada_auto = None
    idx_col_detectada_auto = None
    
    # Testa as primeiras 10 linhas procurando por keywords
    for linha_teste in range(min(10, len(df_raw))):
        for i in range(len(df_raw.columns) - 1, -1, -1):  # De trás para frente
            header = str(df_raw.iloc[linha_teste, i]).upper().strip()
            for keyword in keywords_nomes:
                if keyword in header:
                    col_detectada_auto = letras_disponíveis[i]
                    idx_col_detectada_auto = i
                    break
            if col_detectada_auto:
                break
        if col_detectada_auto:
            break
    
    # Se não encontrou pela keyword, detecta por conteúdo (muitas letras)
    if col_detectada_auto is None:
        for i in range(len(df_raw.columns) - 1, -1, -1):
            valores = df_raw.iloc[:, i].astype(str).str.strip()
            tem_letras = valores.apply(lambda x: any(c.isalpha() for c in x)).sum() > len(valores) * 0.7
            if tem_letras:
                col_detectada_auto = letras_disponíveis[i]
                idx_col_detectada_auto = i
                break
    
    # Detecta automaticamente qual é a linha com os dias
    # Procura pela primeira linha que tem números em sequência (1, 2, 3, 4, 5...)
    linha_detectada = None
    for tentativa_linha in range(min(20, len(df_raw))):
        valores_linha = [str(df_raw.iloc[tentativa_linha, i]).strip() for i in range(len(df_raw.columns))]
        numeros_encontrados = [v for v in valores_linha if v.isdigit() or v in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31']]
        if len(numeros_encontrados) >= 15:  # Se tem pelo menos 15 números (dias do mês)
            linha_detectada = tentativa_linha
            break
    
    linhas = [f"Linha {i+1}" for i in range(min(20, len(df_raw)))]
    
    # Carrega configuração salva do arquivo se existir, senão usa default
    nome_arquivo = file_encarregado.name
    if nome_arquivo in st.session_state.config_arquivos:
        config = st.session_state.config_arquivos[nome_arquivo]
        default_linha = f"Linha {config['linha_idx'] + 1}"
        default_col = config['col_idx']
        default_nome_encarregado = config.get('nome_encarregado', '')
    else:
        default_linha = linhas[0]
        default_col = 0
        default_nome_encarregado = ''
    
    # Inicializa selectbox state se não existir (usa configuração salva)
    if f'l_{idx_arquivo_atual}' not in st.session_state:
        st.session_state[f'l_{idx_arquivo_atual}'] = default_linha
    if f'c_{idx_arquivo_atual}' not in st.session_state:
        st.session_state[f'c_{idx_arquivo_atual}'] = letras_disponíveis[default_col]
    if f'encarregado_{idx_arquivo_atual}' not in st.session_state:
        st.session_state[f'encarregado_{idx_arquivo_atual}'] = default_nome_encarregado
    
    c1, c2 = st.columns(2)
    with c1:
        linha_sel = st.selectbox("Linha com DATAS:", linhas, key=f"l_{idx_arquivo_atual}")
        idx_linha = int(linha_sel.split()[1]) - 1  # -1 para voltar ao índice 0 do pandas
    
    with c2:
        col_sel = st.selectbox("Coluna NOMES:", letras_disponíveis, key=f"c_{idx_arquivo_atual}")
        idx_col = letras_disponíveis.index(col_sel)  # Pega o índice baseado na letra

    # Mostra dicas se há diferenças entre detecção e seleção
    tem_dica_linha = linha_detectada is not None and idx_linha != linha_detectada
    tem_dica_coluna = col_detectada_auto is not None and idx_col_detectada_auto != idx_col
    
    if tem_dica_linha:
        st.info(f"💡 **Dica:** Detectei que a linha {linha_detectada + 1} tem os DIAS em sequência. Você selecionou a linha {idx_linha + 1}.")  # +1 para mostrar como Excel
    
    # Mostra dica da coluna logo após a dica da linha
    if tem_dica_coluna:
        st.info(f"💡 **Dica:** Detectei que a coluna **{col_detectada_auto}** tem nomes. Você selecionou a coluna **{col_sel}**.")
    
    # Botão "Aderir Dica" logo após as dicas - só mostra se há dicas
    if tem_dica_linha or tem_dica_coluna:
        col_dica_btn, col_dica_space = st.columns([1, 4])
        with col_dica_btn:
            def aderir_dica():
                if tem_dica_linha:
                    st.session_state[f'l_{idx_arquivo_atual}'] = f"Linha {linha_detectada + 1}"
                if tem_dica_coluna:
                    st.session_state[f'c_{idx_arquivo_atual}'] = col_detectada_auto
            
            st.button("✅ Aderir Dica", key=f"btn_aderir_{idx_arquivo_atual}", on_click=aderir_dica)
    
    # Caixa de texto para o nome do encarregado
    st.write("**👤 Informações do Encarregado:**")
    nome_encarregado = st.text_input("Nome do Encarregado:", placeholder="Digite o nome do encarregado", key=f"encarregado_{idx_arquivo_atual}")
    st.session_state.nome_encarregado = nome_encarregado

    # Salva configuração deste arquivo
    nome_arquivo = file_encarregado.name
    st.session_state.config_arquivos[nome_arquivo] = {
        'linha_idx': idx_linha,
        'col_idx': idx_col,
        'guia': guia_selecionada,
        'nome_encarregado': nome_encarregado
    }
    
    st.session_state.linha_idx = idx_linha
    st.session_state.col_idx = idx_col
    st.session_state.df_raw = df_raw
    
    st.success(f"✅ Linha {idx_linha + 1} + Coluna {col_sel} - Configurado!")  # +1 para mostrar como Excel
    
    # Se a coluna mudou, recarrega o preview com auto-fit
    if 'col_idx_anterior' not in st.session_state or st.session_state.col_idx_anterior != idx_col:
        st.session_state.col_idx_anterior = idx_col
        st.rerun()
    
    # Mostra TODAS as linhas de dados (não só 10)
    # PULA a primeira linha (idx_linha) porque é a linha de cabeçalho
    # INCLUI: a coluna de nomes E TODAS as colunas DEPOIS dela
    colunas_para_manter = [i for i in range(idx_col, len(df_raw.columns))]  # Inclui idx_col também!
    df_prev = df_raw.iloc[idx_linha+1:, colunas_para_manter].copy()
    
    # Cria índice começando em idx_linha+2 (próxima linha após o cabeçalho, em formato Excel)
    df_prev.index = range(idx_linha + 2, idx_linha + 2 + len(df_prev))
    
    # Renomeia colunas para letras (A, B, C, D...) como no Excel APÓS remover
    letras = []
    for i in range(len(df_prev.columns)):
        if i < 26:
            letras.append(chr(65 + i))  # A-Z
        else:
            letras.append(f"{chr(65 + i//26 - 1)}{chr(65 + i%26)}")  # AA, AB, etc
    
    df_prev.columns = letras
    
    # Remove "nan", "None" e NaN do preview - substitui por vazio
    df_prev = df_prev.replace(['nan', 'None', 'NaN', '<NA>'], '')
    df_prev = df_prev.fillna('')
    
    # Remove decimais desnecessários (1.0 -> 1, 4.0 -> 4)
    def remove_decimais(x):
        try:
            if isinstance(x, str) and '.' in x and x.replace('.', '').replace('-', '').isdigit():
                return str(int(float(x)))
        except:
            pass
        return x
    
    for col in df_prev.columns:
        df_prev[col] = df_prev[col].apply(remove_decimais)
    
    # Exibe com st.dataframe normal - key dinâmica força rerender
    st.dataframe(df_prev, width='stretch', height=600, key=f"preview_{idx_col}")

# Botão de processamento com validação
col_btn_processar, col_status = st.columns([1, 3])

with col_btn_processar:
    # Verifica se todos os arquivos foram configurados
    nomes_arquivos_upload = [f.name for f in files_encarregado] if files_encarregado else []
    configs_salvas = list(st.session_state.get('config_arquivos', {}).keys())
    todos_configurados = len(nomes_arquivos_upload) > 0 and all(nome in configs_salvas for nome in nomes_arquivos_upload)
    
    if st.button("🚀 Processar TODOS os Arquivos", disabled=not todos_configurados):
        if file_mestra and files_encarregado and todos_configurados:
            try:
                # Carrega a planilha mestra UMA VEZ
                df_mest = None
                
                try:
                    # Primeiro tenta ler direto com pandas (automático)
                    file_mestra.seek(0)
                    df_mest = pd.read_excel(file_mestra, header=0)
                except Exception as e1:
                    try:
                        # Segunda tentativa: especifica openpyxl
                        file_mestra.seek(0)
                        df_mest = pd.read_excel(file_mestra, header=0, engine='openpyxl')
                    except Exception as e2:
                        try:
                            # Terceira tentativa: lê como bytes
                            file_mestra.seek(0)
                            file_bytes = file_mestra.read()
                            import io as io_module
                            df_mest = pd.read_excel(io_module.BytesIO(file_bytes), header=0, engine='openpyxl')
                        except Exception as e3:
                            try:
                                # Quarta tentativa: tenta reparar ZIP interno
                                import zipfile
                                file_mestra.seek(0)
                                file_bytes = file_mestra.read()
                                import io as io_module
                                
                                try:
                                    # Tenta verificar como ZIP
                                    zip_file = zipfile.ZipFile(io_module.BytesIO(file_bytes), 'r')
                                    zip_file.testzip()  # Testa integridade
                                    df_mest = pd.read_excel(io_module.BytesIO(file_bytes), header=0, engine='openpyxl')
                                except zipfile.BadZipFile:
                                    st.warning("⚠️ Arquivo ZIP corrompido, tentando recuperar...")
                                    # Tenta reparar procurando por PK (assinatura ZIP)
                                    pk_index = file_bytes.find(b'PK\x03\x04')
                                    if pk_index > 0:
                                        file_bytes_repaired = file_bytes[pk_index:]
                                        df_mest = pd.read_excel(io_module.BytesIO(file_bytes_repaired), header=0, engine='openpyxl')
                                    else:
                                        raise Exception("Não foi possível reparar o arquivo")
                            except Exception as e4:
                                st.error(f"❌ Erro ao ler planilha mestra:\n\n**Tentativa 1 (automática):** {str(e1)}\n**Tentativa 2 (openpyxl):** {str(e2)}\n**Tentativa 3 (bytes):** {str(e3)}\n**Tentativa 4 (reparar ZIP):** {str(e4)}\n\n**Solução:** O arquivo está severamente corrompido e não pode ser recuperado. Tente:\n1. Abrir o arquivo no LibreOffice/Excel\n2. Salvar como novo arquivo (.xlsx)\n3. Fazer upload novamente")
                                st.stop()
                
                if df_mest is None:
                    st.error("❌ Não foi possível carregar a planilha mestra (DataFrame vazio)")
                    st.stop()
                
                if 'NOME' not in df_mest.columns:
                    st.error("Coluna NOME não encontrada!")
                    st.stop()
                
                df_mest['NOME_LIMPO'] = df_mest['NOME'].apply(limpar_nome)
                
                mapa_datas = {}
                for col in df_mest.columns:
                    # Detecta colunas de data: datetime, date, ou string no formato DD/MM ou DD/mmm
                    if isinstance(col, (datetime.datetime, datetime.date)):
                        mapa_datas[col.date()] = col
                    elif isinstance(col, str):
                        # Tenta extrair data de strings no formato "DD/MM" ou "DD/mmm"
                        try:
                            # Tenta parse como data
                            data_obj = extrair_dia_do_cabecalho(col, mes, ano)
                            if data_obj:
                                mapa_datas[data_obj] = col
                        except:
                            pass
                
                # Debug: mostra quantas datas foram encontradas
                st.write(f"📅 Encontradas {len(mapa_datas)} colunas de data")
                if len(mapa_datas) == 0:
                    st.warning("⚠️ Nenhuma coluna de data encontrada! Colunas disponíveis: " + str(list(df_mest.columns)))
                
                # Pré-preenche APENAS sábados e domingos VAZIOS com "D" (Descanso)
                st.info("🗓️ Pré-preenchendo fins de semana vazios com 'D'...")
                for data_obj, col_data_obj in mapa_datas.items():
                    # data_obj já é uma datetime.date, col_data_obj é o nome da coluna
                    if eh_fim_de_semana(data_obj):
                        for idx in df_mest.index:
                            # Verifica se a célula está vazia antes de preencher
                            valor_atual = df_mest.at[idx, col_data_obj]
                            
                            # Converte para string e limpa espaços
                            valor_str = str(valor_atual).strip() if valor_atual is not None else ''
                            
                            # Considera vazio se for: '', 'nan', 'none', '<na>', ou se for NaN
                            eh_vazio = (
                                valor_str == '' or 
                                valor_str.lower() in ['nan', 'none', '<na>', 'nat'] or 
                                pd.isna(valor_atual)
                            )
                            
                            # Só preenche se estiver realmente vazio
                            if eh_vazio:
                                df_mest.at[idx, col_data_obj] = 'D'
                
                # Processa CADA arquivo de encarregado
                total_sucesso = 0
                total_erros = []  # Agora será uma lista de tuplas: (nome_colaborador, nome_arquivo)
                total_nomes_unicos = set()
                total_linhas_processadas = set()
                
                with st.spinner('Processando todos os arquivos...'):
                    for idx_arquivo, file_enc in enumerate(files_encarregado):
                        # Recupera a configuração salva deste arquivo
                        config = st.session_state.config_arquivos.get(file_enc.name)
                        if not config:
                            st.warning(f"⚠️ Arquivo {file_enc.name} não foi configurado, pulando...")
                            continue
                        
                        idx_linha = config['linha_idx']
                        idx_col = config['col_idx']
                        guia_usar = config['guia']
                        nome_encarregado = config['nome_encarregado']
                        
                        st.write(f"📄 Processando: **{file_enc.name}**")
                        
                        df_enc = None
                        try:
                            # Primeira tentativa: leitura direta
                            file_enc.seek(0)
                            df_enc = pd.read_excel(file_enc, sheet_name=guia_usar, header=None, dtype=str)
                        except Exception as e1:
                            try:
                                # Segunda tentativa: com openpyxl
                                file_enc.seek(0)
                                df_enc = pd.read_excel(file_enc, sheet_name=guia_usar, header=None, dtype=str, engine='openpyxl')
                            except Exception as e2:
                                try:
                                    # Terceira tentativa: bytes
                                    file_enc.seek(0)
                                    file_bytes = file_enc.read()
                                    import io as io_module
                                    df_enc = pd.read_excel(io_module.BytesIO(file_bytes), sheet_name=guia_usar, header=None, dtype=str, engine='openpyxl')
                                except Exception as e3:
                                    try:
                                        # Quarta tentativa: tenta reparar ZIP
                                        import zipfile
                                        file_enc.seek(0)
                                        file_bytes = file_enc.read()
                                        import io as io_module
                                        
                                        try:
                                            zip_file = zipfile.ZipFile(io_module.BytesIO(file_bytes), 'r')
                                            zip_file.testzip()
                                            df_enc = pd.read_excel(io_module.BytesIO(file_bytes), sheet_name=guia_usar, header=None, dtype=str, engine='openpyxl')
                                        except zipfile.BadZipFile:
                                            pk_index = file_bytes.find(b'PK\x03\x04')
                                            if pk_index > 0:
                                                file_bytes_repaired = file_bytes[pk_index:]
                                                df_enc = pd.read_excel(io_module.BytesIO(file_bytes_repaired), sheet_name=guia_usar, header=None, dtype=str, engine='openpyxl')
                                            else:
                                                raise Exception("Não foi possível reparar")
                                    except Exception as e4:
                                        st.error(f"❌ Erro ao ler arquivo {file_enc.name}: {str(e1)}")
                                        continue
                        
                        if df_enc is None:
                            continue
                        
                        cols_nomes = [str(df_enc.iloc[idx_linha, i]) for i in range(len(df_enc.columns))]
                        df_enc = df_enc.iloc[idx_linha+1:].copy()
                        df_enc.columns = cols_nomes
                        df_enc.reset_index(drop=True, inplace=True)
                        
                        col_nome = cols_nomes[idx_col]
                        cols_datas = cols_nomes[idx_col + 1:]
                        
                        df_enc = df_enc.dropna(how='all')
                        # Usa iloc para pegar a coluna por índice para evitar problema com nomes duplicados
                        df_enc = df_enc[df_enc.iloc[:, idx_col].astype(str).str.strip() != '']
                        df_enc.reset_index(drop=True, inplace=True)
                        
                        # Renomeia a coluna de nomes para algo único para evitar problemas com colunas duplicadas
                        df_enc_temp = df_enc.iloc[:, idx_col:].copy()
                        df_enc_temp.columns = ['___NOME___'] + list(df_enc_temp.columns[1:])
                        
                        df_long = df_enc_temp.melt(
                            id_vars=['___NOME___'],
                            value_vars=cols_datas,
                            var_name='DIA', value_name='COD'
                        )
                        
                        df_long.rename(columns={'___NOME___': 'NOME'}, inplace=True)
                        df_long['NOME_LIMPO'] = df_long['NOME'].apply(limpar_nome)
                        df_long['CODIGO'] = pd.to_numeric(df_long['COD'], errors='coerce')
                        df_long['DATA'] = df_long['DIA'].apply(lambda x: extrair_dia_do_cabecalho(x, mes, ano))
                        df_long = df_long[df_long['NOME_LIMPO'].astype(str).str.strip() != '']
                        df_long = df_long.dropna(subset=['DATA', 'NOME_LIMPO'])
                        
                        sucesso = 0
                        erros = []
                        nomes_com_erro = set()  # Rastreia nomes únicos que não foram encontrados
                        linhas_processadas = set()
                        nomes_unicos = set()
                        
                        for _, row in df_long.iterrows():
                            nome = row['NOME_LIMPO']
                            cod = row['CODIGO']
                            data = row['DATA']
                            
                            nomes_unicos.add(nome)
                            total_nomes_unicos.add(nome)
                            
                            if pd.isna(cod) or cod not in MAPA_CODIGOS or data not in mapa_datas:
                                continue
                            
                            col_data = mapa_datas[data]
                            
                            # BUSCA EXATA
                            match = df_mest['NOME_LIMPO'] == nome
                            
                            # BUSCA FUZZY (por similaridade)
                            if match.sum() == 0:
                                similaridades = df_mest['NOME_LIMPO'].apply(lambda x: calcular_similaridade(nome, x))
                                match = similaridades >= 0.85
                            
                            # BUSCA POR PALAVRAS-CHAVE
                            if match.sum() == 0:
                                palavras_nome = nome.split()[:3]
                                
                                def contem_palavras_iniciais(nome_mestra):
                                    palavras_mestra = nome_mestra.split()
                                    return all(p in palavras_mestra for p in palavras_nome[:2])
                                
                                match = df_mest['NOME_LIMPO'].apply(contem_palavras_iniciais)
                            
                            if match.sum() > 0:
                                indices_match = df_mest[match].index.tolist()
                                
                                for idx in indices_match:
                                    if df_mest[col_data].dtype != 'object':
                                        df_mest[col_data] = df_mest[col_data].astype('object')
                                    
                                    df_mest.at[idx, col_data] = MAPA_CODIGOS[cod]
                                    linhas_processadas.add(idx)
                                    total_linhas_processadas.add(idx)
                                
                                sucesso += 1
                            else:
                                # Só adiciona ao erro se não foi adicionado antes
                                if nome not in nomes_com_erro:
                                    erros.append(nome)
                                    nomes_com_erro.add(nome)
                        
                        # Atualiza GESTOR para este arquivo (usa o nome_encarregado da configuração)
                        if nome_encarregado and nome_encarregado.strip() != '':
                            if 'GESTOR' in df_mest.columns:
                                for idx in linhas_processadas:
                                    df_mest.at[idx, 'GESTOR'] = nome_encarregado
                        
                        # Agrega erros locais para o total com nome do arquivo
                        for erro_nome in erros:
                            total_erros.append((erro_nome, file_enc.name))
                        
                        st.success(f"  ✅ {sucesso} lançamentos | 👥 {len(nomes_unicos)} colaboradores únicos")
                        total_sucesso += sucesso
                
                st.divider()
                st.success(f"🎉 Total: ✅ {total_sucesso} lançamentos | 👥 {len(total_nomes_unicos)} colaboradores processados")
                
                if total_erros:
                    with st.expander(f"⚠️ {len(set(total_erros))} não encontrados (de todos os arquivos)"):
                        for e in list(set(total_erros))[:15]:
                            st.write(f"- {e}")
                
                # ===== GERADOR DE RELATÓRIO =====
                st.divider()
                st.header("📊 Relatório Detalhado")
                
                # Seção 1: Colaboradores não processados
                st.subheader("❌ Colaboradores não encontrados")
                
                if total_erros:
                    col1, col2 = st.columns([2, 1])
                    with col1:
                        st.write(f"**Total:** {len(total_erros)} colaboradores")
                    with col2:
                        st.write(f"**Motivo:** Não encontrados na Planilha Mestra")
                    
                    with st.expander(f"📋 Ver lista completa ({len(total_erros)} nomes)"):
                        # Cria uma tabela com nome e arquivo
                        for nome_colaborador, nome_arquivo in sorted(total_erros):
                            st.write(f"• **{nome_colaborador}** - Arquivo: `{nome_arquivo}`")
                else:
                    st.success("✅ Todos os colaboradores foram encontrados e processados!")
                
                st.divider()
                out = io.BytesIO()
                df_mest_final = df_mest.drop(columns=['NOME_LIMPO'])
                
                with pd.ExcelWriter(out, engine='openpyxl') as w:
                    df_mest_final.to_excel(w, index=False, sheet_name='Dados')
                    
                    worksheet = w.sheets['Dados']
                    
                    # ===== FORMATAÇÃO DO HEADER =====
                    header_fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')  # Azul escuro
                    header_font = Font(bold=True, color='FFFFFFFF', size=11)  # Texto branco
                    
                    # Formata todas as colunas do header
                    for col_idx in range(1, len(df_mest_final.columns) + 1):
                        header_cell = worksheet.cell(row=1, column=col_idx)
                        header_cell.fill = header_fill
                        header_cell.font = header_font
                        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # ===== FORMATAÇÃO DAS COLUNAS ESPECÍFICAS =====
                    # Mapeamento de colunas com cores
                    col_names = df_mest_final.columns.tolist()
                    
                    # Função para calcular largura baseada no maior valor da coluna
                    def calc_width(df, col_name, min_width=10, max_width=50):
                        if col_name not in df.columns:
                            return min_width
                        max_len = df[col_name].astype(str).str.len().max()
                        header_len = len(str(col_name))
                        largest = max(max_len, header_len)
                        width = min(max(largest + 2, min_width), max_width)
                        return width
                    
                    for col_idx, col_name in enumerate(col_names, 1):
                        # Define preenchimento e largura para cada tipo de coluna
                        if col_name == 'NOME':
                            col_fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')  # Azul claro suave
                            width = calc_width(df_mest_final, col_name, min_width=15, max_width=40)
                            worksheet.column_dimensions[get_column_letter(col_idx)].width = width
                        elif col_name == 'AREA':
                            col_fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')  # Verde claro suave
                            worksheet.column_dimensions[get_column_letter(col_idx)].width = 25
                        elif col_name == 'GESTOR':
                            col_fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')  # Laranja #ffbf5e
                            width = calc_width(df_mest_final, col_name, min_width=15, max_width=40)
                            worksheet.column_dimensions[get_column_letter(col_idx)].width = width
                        else:
                            col_fill = None
                            # Largura fixa para outras colunas
                            try:
                                datetime.datetime.strptime(str(col_name), '%d/%m')
                                worksheet.column_dimensions[get_column_letter(col_idx)].width = 7
                            except:
                                worksheet.column_dimensions[get_column_letter(col_idx)].width = 10
                        
                        # Aplica a cor a todas as linhas de dados desta coluna
                        if col_fill is not None:
                            for row_idx in range(2, worksheet.max_row + 1):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.fill = col_fill
                    
                    for col_data_obj in mapa_datas.values():
                        col_idx = list(df_mest_final.columns).index(col_data_obj) + 1
                        
                        header_cell = worksheet.cell(row=1, column=col_idx)
                        if isinstance(col_data_obj, (datetime.datetime, datetime.date)):
                            data_formatada = col_data_obj.strftime('%d/%m') if isinstance(col_data_obj, datetime.date) else col_data_obj.date().strftime('%d/%m')
                            header_cell.value = data_formatada
                        
                        for row_idx, row in enumerate(worksheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2), start=2):
                            for cell in row:
                                cell.number_format = 'DD/MM'
                                
                                valor = str(cell.value).strip() if cell.value else ''
                                
                                if valor == 'P':
                                    cell.fill = PatternFill(start_color=MAPA_CORES['P'], end_color=MAPA_CORES['P'], fill_type='solid')
                                elif valor == 'FI':
                                    cell.fill = PatternFill(start_color=MAPA_CORES['FI'], end_color=MAPA_CORES['FI'], fill_type='solid')
                                    cell.font = Font(color='FFFFFFFF')
                                elif valor == 'FA':
                                    cell.fill = PatternFill(start_color=MAPA_CORES['FA'], end_color=MAPA_CORES['FA'], fill_type='solid')
                                elif valor == 'FÉRIAS-BH':
                                    cell.fill = PatternFill(start_color=MAPA_CORES['FÉRIAS-BH'], end_color=MAPA_CORES['FÉRIAS-BH'], fill_type='solid')
                                    cell.font = Font(color='FFFFFFFF')
                                elif valor == 'DESLIGADO':
                                    cell.fill = PatternFill(start_color=MAPA_CORES['DESLIGADO'], end_color=MAPA_CORES['DESLIGADO'], fill_type='solid')
                                    cell.font = Font(color='FFFFFFFF')
                                elif valor == 'D':
                                    cell.fill = PatternFill(start_color=MAPA_CORES['DESCANSO'], end_color=MAPA_CORES['DESCANSO'], fill_type='solid')
                    
                    # ===== OBTER FERIADOS PARA USO NO SHEET RELATÓRIO E PORCENTAGENS =====
                    if mapa_datas:
                        ano_feriados_temp = min(mapa_datas.keys()).year
                        feriados_temp = obter_feriados_brasil(ano_feriados_temp)
                    else:
                        feriados_temp = {}
                    
                    # ===== CRIAR GUIA DE RELATÓRIO =====
                    ws_relatorio = w.book.create_sheet('Relatório')
                    
                    # Linha 1: Título
                    titulo_cell = ws_relatorio.cell(row=1, column=1, value='📊 RELATÓRIO DE PROCESSAMENTO')
                    titulo_cell.font = Font(bold=True, size=14, color='FFFFFF')
                    titulo_cell.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                    ws_relatorio.merge_cells('A1:D1')
                    
                    # Linha 2: Data/Hora
                    ws_relatorio.cell(row=2, column=1, value='Data do Processamento:')
                    ws_relatorio.cell(row=2, column=2, value=datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
                    
                    ws_relatorio.cell(row=2, column=3, value='Mês/Ano:')
                    ws_relatorio.cell(row=2, column=4, value=f"{mes:02d}/{ano}")
                    
                    # Linha 4: Resumo por Dia
                    ws_relatorio.merge_cells('A4:F4')
                    cell_resumo = ws_relatorio.cell(row=4, column=1, value='RESUMO POR DIA')
                    cell_resumo.font = Font(bold=True, size=12, color='FFFFFFFF')
                    cell_resumo.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                    
                    # Headers da tabela de resumo
                    headers_resumo = ['Data', 'Dia', 'FI', 'FA', 'FÉRIAS-BH', 'Total']
                    for col_idx, header in enumerate(headers_resumo, 1):
                        cell = ws_relatorio.cell(row=5, column=col_idx, value=header)
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                    
                    # Preenche tabela de resumo
                    dias_semana_pt = {
                        'MON': 'SEG', 'TUE': 'TER', 'WED': 'QUA', 'THU': 'QUI',
                        'FRI': 'SEX', 'SAT': 'SÁB', 'SUN': 'DOM'
                    }
                    
                    row_idx = 6
                    for data_obj in sorted(mapa_datas.keys()):
                        col_data = mapa_datas[data_obj]
                        if col_data in df_mest.columns:
                            # Encontra o índice da coluna de dados
                            col_letter = get_column_letter(list(df_mest_final.columns).index(col_data) + 1)
                            
                            if row_idx == 11:  # Apenas calcula uma vez para referenciar depois
                                data_start_row = row_idx
                            
                            data_formatada = data_obj.strftime('%d/%m/%Y') if isinstance(data_obj, datetime.date) else str(data_obj)
                            dia_en = data_obj.strftime('%a').upper() if isinstance(data_obj, datetime.date) else '???'
                            dia_semana = dias_semana_pt.get(dia_en, dia_en)
                            
                            # Verifica se é feriado ou domingo
                            eh_feriado = data_obj in feriados_temp if 'feriados_temp' in locals() else False
                            eh_domingo = data_obj.weekday() == 6 if isinstance(data_obj, datetime.date) else False
                            
                            # Coluna Data
                            cell_data = ws_relatorio.cell(row=row_idx, column=1, value=data_formatada)
                            # Coluna Dia
                            cell_dia = ws_relatorio.cell(row=row_idx, column=2, value=dia_semana)
                            # Coluna FI
                            cell_fi = ws_relatorio.cell(row=row_idx, column=3)
                            cell_fi.value = f'=COUNTIF(Dados!{col_letter}:${col_letter},"FI")'
                            # Coluna FA
                            cell_fa = ws_relatorio.cell(row=row_idx, column=4)
                            cell_fa.value = f'=COUNTIF(Dados!{col_letter}:${col_letter},"FA")'
                            # Coluna FÉRIAS-BH
                            cell_ferias = ws_relatorio.cell(row=row_idx, column=5)
                            cell_ferias.value = f'=COUNTIF(Dados!{col_letter}:${col_letter},"FÉRIAS-BH")'
                            # Coluna Total
                            cell_total = ws_relatorio.cell(row=row_idx, column=6)
                            
                            # Aplica formatação condicional para FERIADO/DOMINGO
                            if eh_feriado:
                                for col_idx in range(1, 7):
                                    cell = ws_relatorio.cell(row=row_idx, column=col_idx)
                                    cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell.font = Font(bold=True, color='FFFFFFFF')
                                # Substitui valores por FERIADO
                                cell_fi.value = 'FERIADO'
                                cell_fa.value = 'FERIADO'
                                cell_ferias.value = 'FERIADO'
                                cell_total.value = 'FERIADO'
                            elif eh_domingo:
                                for col_idx in range(1, 7):
                                    cell = ws_relatorio.cell(row=row_idx, column=col_idx)
                                    cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell.font = Font(bold=True, color='FFFFFFFF')
                                # Substitui valores por DOMINGO
                                cell_fi.value = 'DOMINGO'
                                cell_fa.value = 'DOMINGO'
                                cell_ferias.value = 'DOMINGO'
                                cell_total.value = 'DOMINGO'
                            else:
                                cell_data.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                cell_dia.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                cell_fi.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                                cell_fi.font = Font(bold=True, color='FFFFFFFF')
                                cell_fa.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                                cell_fa.font = Font(bold=True, color='FFFFFFFF')
                                cell_ferias.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                                cell_ferias.font = Font(color='FFFFFFFF')
                                cell_total.value = f'=C{row_idx}+D{row_idx}+E{row_idx}'
                                cell_total.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                            
                            row_idx += 1
                    
                    # Linha de Resumo por Departamento (DIÁRIO)
                    row_departamento = row_idx + 2
                    ws_relatorio.merge_cells(f'A{row_departamento}:H{row_departamento}')
                    cell_depto = ws_relatorio.cell(row=row_departamento, column=1, value='RESUMO POR DEPARTAMENTO (DIÁRIO)')
                    cell_depto.font = Font(bold=True, size=12, color='FFFFFFFF')
                    cell_depto.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                    
                    # Mapeia setores para departamentos
                    setores_ma_bloq = ['MOVIMENTACAO E ARMAZENAGEM', 'PROJETO INTERPRISE - MOVIMENTACAO E ARMAZENAGEM', 'BLOQ', 'CD-RJ | FOB']
                    setores_crdk_de = ['CRDK D&E|CD-RJ HB', 'CROSSDOCK DISTRIBUICAO E EXPEDICAO', 'DISTRIBUICAO E EXPEDICAO']
                    
                    # Headers do resumo por departamento (com datas)
                    row_departamento += 1
                    headers_depto = ['Data', 'Dia', 'Depto', 'FI', 'FA', 'Total']
                    for col_idx, header in enumerate(headers_depto, 1):
                        cell = ws_relatorio.cell(row=row_departamento, column=col_idx, value=header)
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                    
                    # Função para contar FI e FA por departamento e data
                    def contar_fi_fa_por_depto_data(df, setores_lista, col_area, data_col):
                        total_fi = 0
                        total_fa = 0
                        
                        for setor in setores_lista:
                            # Filtra colaboradores deste setor
                            mask_setor = df[col_area].astype(str).str.contains(setor, case=False, na=False)
                            df_setor = df[mask_setor]
                            
                            if not df_setor.empty and data_col in df.columns:
                                # Conta FI e FA para esta data (exclui FERIADO)
                                total_fi += ((df_setor[data_col] == 'FI')).sum()
                                total_fa += ((df_setor[data_col] == 'FA')).sum()
                        
                        return total_fi, total_fa
                    
                    # Preenche tabela com dados por dia
                    if 'AREA' in df_mest.columns:
                        area_col_idx = list(df_mest_final.columns).index('AREA') + 1
                        area_col_letter = get_column_letter(area_col_idx)
                        row_departamento += 1
                        
                        for data_obj in sorted(mapa_datas.keys()):
                            col_data = mapa_datas[data_obj]
                            if col_data in df_mest.columns:
                                data_col_idx = list(df_mest_final.columns).index(col_data) + 1
                                data_col_letter = get_column_letter(data_col_idx)
                                
                                data_formatada = data_obj.strftime('%d/%m/%Y') if isinstance(data_obj, datetime.date) else str(data_obj)
                                dia_en = data_obj.strftime('%a').upper() if isinstance(data_obj, datetime.date) else '???'
                                dia_semana = dias_semana_pt.get(dia_en, dia_en)
                                
                                # Verifica se é feriado ou domingo
                                eh_feriado = data_obj in feriados_temp if 'feriados_temp' in locals() else False
                                eh_domingo = data_obj.weekday() == 6 if isinstance(data_obj, datetime.date) else False
                                
                                # M&A / BLOQ - Usar ordem específica para evitar duplicação
                                # 1. PROJETO INTERPRISE (mais específico)
                                # 2. MOVIMENTACAO (mas EXCLUI PROJETO INTERPRISE)
                                # 3. BLOQ
                                # 4. CD-RJ | FOB
                                fi_ma_bloq_formula = (
                                    f'=SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("PROJETO INTERPRISE - MOVIMENTACAO E ARMAZENAGEM",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FI"))'
                                    f'+SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("MOVIMENTACAO E ARMAZENAGEM",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'NOT(ISNUMBER(SEARCH("PROJETO INTERPRISE",Dados!{area_col_letter}:${area_col_letter})))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FI"))'
                                    f'+SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("BLOQ",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FI"))'
                                    f'+SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("CD-RJ | FOB",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FI"))'
                                )
                                fa_ma_bloq_formula = (
                                    f'=SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("PROJETO INTERPRISE - MOVIMENTACAO E ARMAZENAGEM",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FA"))'
                                    f'+SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("MOVIMENTACAO E ARMAZENAGEM",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'NOT(ISNUMBER(SEARCH("PROJETO INTERPRISE",Dados!{area_col_letter}:${area_col_letter})))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FA"))'
                                    f'+SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("BLOQ",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FA"))'
                                    f'+SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("CD-RJ | FOB",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FA"))'
                                )
                                
                                # Coluna Data (cinza)
                                cell_data = ws_relatorio.cell(row=row_departamento, column=1, value=data_formatada)
                                # Coluna Dia (cinza)
                                cell_dia = ws_relatorio.cell(row=row_departamento, column=2, value=dia_semana)
                                # Coluna Depto (verde suave)
                                cell_depto = ws_relatorio.cell(row=row_departamento, column=3, value='M&A / BLOQ')
                                # Coluna FI (vermelho suave) - Fórmula
                                cell_fi = ws_relatorio.cell(row=row_departamento, column=4)
                                cell_fi.value = fi_ma_bloq_formula
                                # Coluna FA (amarelo suave) - Fórmula
                                cell_fa = ws_relatorio.cell(row=row_departamento, column=5)
                                cell_fa.value = fa_ma_bloq_formula
                                # Coluna Total
                                cell_total = ws_relatorio.cell(row=row_departamento, column=6)
                                cell_total.value = f'=D{row_departamento}+E{row_departamento}'
                                
                                # Aplica formatação condicional para FERIADO/DOMINGO
                                if eh_feriado:
                                    for col_idx in range(1, 7):
                                        cell = ws_relatorio.cell(row=row_departamento, column=col_idx)
                                        cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                        cell.font = Font(bold=True, color='FFFFFFFF')
                                    # Substitui conteúdo por FERIADO
                                    cell_fi.value = 'FERIADO'
                                    cell_fa.value = 'FERIADO'
                                    cell_total.value = 'FERIADO'
                                elif eh_domingo:
                                    for col_idx in range(1, 7):
                                        cell = ws_relatorio.cell(row=row_departamento, column=col_idx)
                                        cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                        cell.font = Font(bold=True, color='FFFFFFFF')
                                    # Substitui conteúdo por DOMINGO
                                    cell_fi.value = 'DOMINGO'
                                    cell_fa.value = 'DOMINGO'
                                    cell_total.value = 'DOMINGO'
                                else:
                                    cell_data.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                    cell_dia.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                    cell_depto.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                    cell_fi.font = Font(bold=True, color='FFFFFFFF')
                                    cell_fi.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                                    cell_fa.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                                    cell_fa.font = Font(bold=True, color='FFFFFFFF')
                                    cell_total.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                row_departamento += 1
                                
                                # CRDK / D&E - Usar ordem específica para evitar duplicação
                                # 1. CROSSDOCK (mais específico)
                                # 2. CRDK D&E|CD-RJ HB
                                # 3. DISTRIBUICAO E EXPEDICAO (mas EXCLUI o que já foi contado em CROSSDOCK)
                                fi_crdk_de_formula = (
                                    f'=SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("CROSSDOCK DISTRIBUICAO E EXPEDICAO",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FI"))'
                                    f'+'
                                    f'SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("CRDK D&E|CD-RJ HB",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FI"))'
                                    f'+'
                                    f'SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("DISTRIBUICAO E EXPEDICAO",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'NOT(ISNUMBER(SEARCH("CROSSDOCK",Dados!{area_col_letter}:${area_col_letter})))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FI"))'
                                )
                                fa_crdk_de_formula = (
                                    f'=SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("CROSSDOCK DISTRIBUICAO E EXPEDICAO",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FA"))'
                                    f'+'
                                    f'SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("CRDK D&E|CD-RJ HB",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FA"))'
                                    f'+'
                                    f'SUMPRODUCT('
                                    f'ISNUMBER(SEARCH("DISTRIBUICAO E EXPEDICAO",Dados!{area_col_letter}:${area_col_letter}))*'
                                    f'NOT(ISNUMBER(SEARCH("CROSSDOCK",Dados!{area_col_letter}:${area_col_letter})))*'
                                    f'(Dados!{data_col_letter}:${data_col_letter}="FA"))'
                                )
                                
                                # Coluna Data (cinza)
                                cell_data = ws_relatorio.cell(row=row_departamento, column=1, value=data_formatada)
                                # Coluna Dia (cinza)
                                cell_dia = ws_relatorio.cell(row=row_departamento, column=2, value=dia_semana)
                                # Coluna Depto (verde suave)
                                cell_depto = ws_relatorio.cell(row=row_departamento, column=3, value='CRDK / D&E')
                                # Coluna FI (vermelho suave) - Fórmula
                                cell_fi = ws_relatorio.cell(row=row_departamento, column=4)
                                cell_fi.value = fi_crdk_de_formula
                                # Coluna FA (amarelo suave) - Fórmula
                                cell_fa = ws_relatorio.cell(row=row_departamento, column=5)
                                cell_fa.value = fa_crdk_de_formula
                                # Coluna Total
                                cell_total = ws_relatorio.cell(row=row_departamento, column=6)
                                cell_total.value = f'=D{row_departamento}+E{row_departamento}'
                                
                                # Aplica formatação condicional para FERIADO/DOMINGO
                                if eh_feriado:
                                    for col_idx in range(1, 7):
                                        cell = ws_relatorio.cell(row=row_departamento, column=col_idx)
                                        cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                        cell.font = Font(bold=True, color='FFFFFFFF')
                                    # Substitui conteúdo por FERIADO
                                    cell_fi.value = 'FERIADO'
                                    cell_fa.value = 'FERIADO'
                                    cell_total.value = 'FERIADO'
                                elif eh_domingo:
                                    for col_idx in range(1, 7):
                                        cell = ws_relatorio.cell(row=row_departamento, column=col_idx)
                                        cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                        cell.font = Font(bold=True, color='FFFFFFFF')
                                    # Substitui conteúdo por DOMINGO
                                    cell_fi.value = 'DOMINGO'
                                    cell_fa.value = 'DOMINGO'
                                    cell_total.value = 'DOMINGO'
                                else:
                                    cell_data.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                    cell_dia.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                    cell_depto.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                    cell_fi.font = Font(bold=True, color='FFFFFFFF')
                                    cell_fi.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                                    cell_fa.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                                    cell_fa.font = Font(bold=True, color='FFFFFFFF')
                                    cell_total.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                row_departamento += 1
                    
                    # ===== RESUMO POR TURNO, DIA E SETOR =====
                    # Verifica se coluna TURNO existe
                    if 'TURNO' in df_mest.columns:
                        row_turno_section = row_departamento + 2
                        ws_relatorio.merge_cells(f'A{row_turno_section}:H{row_turno_section}')
                        cell_turno_title = ws_relatorio.cell(row=row_turno_section, column=1, value='RESUMO POR TURNO (DIÁRIO)')
                        cell_turno_title.font = Font(bold=True, size=12, color='FFFFFFFF')
                        cell_turno_title.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                        
                        # Headers do resumo por turno
                        row_turno_section += 1
                        headers_turno = ['Turno', 'Data', 'Setor', 'FI', 'FA', 'Total']
                        for col_idx, header in enumerate(headers_turno, 1):
                            cell = ws_relatorio.cell(row=row_turno_section, column=col_idx, value=header)
                            cell.font = Font(bold=True, color='FFFFFF')
                            cell.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                        
                        row_turno_section += 1
                        turno_col_idx = list(df_mest_final.columns).index('TURNO') + 1
                        turno_col_letter = get_column_letter(turno_col_idx)
                        
                        # Para cada turno
                        for turno_num in [1, 2, 3]:
                            turno_label = f'TURNO {turno_num}'
                            
                            # Para cada data
                            for data_obj in sorted(mapa_datas.keys()):
                                col_data = mapa_datas[data_obj]
                                if col_data in df_mest.columns:
                                    data_col_idx = list(df_mest_final.columns).index(col_data) + 1
                                    data_col_letter = get_column_letter(data_col_idx)
                                    
                                    data_formatada = data_obj.strftime('%d/%m/%Y') if isinstance(data_obj, datetime.date) else str(data_obj)
                                    
                                    # Verifica se é feriado ou domingo
                                    eh_feriado = data_obj in feriados_temp if 'feriados_temp' in locals() else False
                                    eh_domingo = data_obj.weekday() == 6 if isinstance(data_obj, datetime.date) else False
                                    
                                    # Para cada setor
                                    setores_turno = [
                                        ('M&A / BLOQ', ['MOVIMENTACAO E ARMAZENAGEM', 'PROJETO INTERPRISE - MOVIMENTACAO E ARMAZENAGEM', 'BLOQ', 'CD-RJ | FOB']),
                                        ('CRDK / D&E', ['CROSSDOCK DISTRIBUICAO E EXPEDICAO', 'CRDK D&E|CD-RJ HB', 'DISTRIBUICAO E EXPEDICAO'])
                                    ]
                                    
                                    for setor_nome, keywords_setor in setores_turno:
                                        # Turno (azul claro)
                                        cell_turno = ws_relatorio.cell(row=row_turno_section, column=1, value=turno_label)
                                        # Data
                                        cell_data = ws_relatorio.cell(row=row_turno_section, column=2, value=data_formatada)
                                        # Setor (verde suave)
                                        cell_setor = ws_relatorio.cell(row=row_turno_section, column=3, value=setor_nome)
                                        # FI
                                        cell_fi = ws_relatorio.cell(row=row_turno_section, column=4)
                                        turno_text = f"TURNO {turno_num}"
                                        if setor_nome == 'M&A / BLOQ':
                                            cell_fi.value = '=SUMPRODUCT((ISNUMBER(SEARCH("' + turno_text + '";Dados!$' + turno_col_letter + ':$' + turno_col_letter + ')))*(ISNUMBER(SEARCH("PROJETO INTERPRISE - MOVIMENTACAO E ARMAZENAGEM";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("MOVIMENTACAO E ARMAZENAGEM";Dados!$' + area_col_letter + ':$' + area_col_letter + '))*NOT(ISNUMBER(SEARCH("PROJETO INTERPRISE";Dados!$' + area_col_letter + ':$' + area_col_letter + ')))+ISNUMBER(SEARCH("BLOQ";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("CD-RJ | FOB";Dados!$' + area_col_letter + ':$' + area_col_letter + ')))*(Dados!$' + data_col_letter + ':$' + data_col_letter + '="FI"))'
                                        else:  # CRDK / D&E
                                            cell_fi.value = '=SUMPRODUCT((ISNUMBER(SEARCH("' + turno_text + '";Dados!$' + turno_col_letter + ':$' + turno_col_letter + ')))*(ISNUMBER(SEARCH("CROSSDOCK DISTRIBUICAO E EXPEDICAO";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("CRDK D&E|CD-RJ HB";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("DISTRIBUICAO E EXPEDICAO";Dados!$' + area_col_letter + ':$' + area_col_letter + '))*NOT(ISNUMBER(SEARCH("CROSSDOCK";Dados!$' + area_col_letter + ':$' + area_col_letter + '))))*(Dados!$' + data_col_letter + ':$' + data_col_letter + '="FI"))'
                                        # FA
                                        cell_fa = ws_relatorio.cell(row=row_turno_section, column=5)
                                        if setor_nome == 'M&A / BLOQ':
                                            cell_fa.value = '=SUMPRODUCT((ISNUMBER(SEARCH("' + turno_text + '";Dados!$' + turno_col_letter + ':$' + turno_col_letter + ')))*(ISNUMBER(SEARCH("PROJETO INTERPRISE - MOVIMENTACAO E ARMAZENAGEM";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("MOVIMENTACAO E ARMAZENAGEM";Dados!$' + area_col_letter + ':$' + area_col_letter + '))*NOT(ISNUMBER(SEARCH("PROJETO INTERPRISE";Dados!$' + area_col_letter + ':$' + area_col_letter + ')))+ISNUMBER(SEARCH("BLOQ";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("CD-RJ | FOB";Dados!$' + area_col_letter + ':$' + area_col_letter + ')))*(Dados!$' + data_col_letter + ':$' + data_col_letter + '="FA"))'
                                        else:  # CRDK / D&E
                                            cell_fa.value = '=SUMPRODUCT((ISNUMBER(SEARCH("' + turno_text + '";Dados!$' + turno_col_letter + ':$' + turno_col_letter + ')))*(ISNUMBER(SEARCH("CROSSDOCK DISTRIBUICAO E EXPEDICAO";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("CRDK D&E|CD-RJ HB";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("DISTRIBUICAO E EXPEDICAO";Dados!$' + area_col_letter + ':$' + area_col_letter + '))*NOT(ISNUMBER(SEARCH("CROSSDOCK";Dados!$' + area_col_letter + ':$' + area_col_letter + '))))*(Dados!$' + data_col_letter + ':$' + data_col_letter + '="FA"))'
                                        # Total
                                        cell_total_turno = ws_relatorio.cell(row=row_turno_section, column=6)
                                        cell_total_turno.value = f'=D{row_turno_section}+E{row_turno_section}'
                                        
                                        # Aplica formatação condicional para FERIADO/DOMINGO
                                        if eh_feriado:
                                            for col_idx in range(1, 7):
                                                cell = ws_relatorio.cell(row=row_turno_section, column=col_idx)
                                                cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                                cell.font = Font(bold=True, color='FFFFFFFF')
                                            cell_fi.value = 'FERIADO'
                                            cell_fa.value = 'FERIADO'
                                            cell_total_turno.value = 'FERIADO'
                                        elif eh_domingo:
                                            for col_idx in range(1, 7):
                                                cell = ws_relatorio.cell(row=row_turno_section, column=col_idx)
                                                cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                                cell.font = Font(bold=True, color='FFFFFFFF')
                                            cell_fi.value = 'DOMINGO'
                                            cell_fa.value = 'DOMINGO'
                                            cell_total_turno.value = 'DOMINGO'
                                        else:
                                            cell_turno.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                            cell_turno.font = Font(bold=True)
                                            cell_data.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                            cell_setor.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                            cell_fi.font = Font(bold=True, color='FFFFFFFF')
                                            cell_fi.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                                            cell_fa.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                                            cell_fa.font = Font(bold=True, color='FFFFFFFF')
                                            cell_total_turno.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                                        
                                        row_turno_section += 1
                    
                    # Linha de Não Encontrados
                    row_nao_encontrados = row_turno_section + 2 if 'TURNO' in df_mest.columns else row_departamento + 2
                    ws_relatorio.merge_cells(f'A{row_nao_encontrados}:D{row_nao_encontrados}')
                    cell_nao_encontrados = ws_relatorio.cell(row=row_nao_encontrados, column=1, value='COLABORADORES NÃO ENCONTRADOS')
                    cell_nao_encontrados.font = Font(bold=True, size=12, color='FFFFFFFF')
                    cell_nao_encontrados.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                    
                    row_nao_encontrados += 1
                    # Headers para a tabela
                    cell_header_nome = ws_relatorio.cell(row=row_nao_encontrados, column=1, value='Colaborador')
                    cell_header_arquivo = ws_relatorio.cell(row=row_nao_encontrados, column=2, value='Arquivo')
                    cell_header_nome.font = Font(bold=True, color='FFFFFF')
                    cell_header_arquivo.font = Font(bold=True, color='FFFFFF')
                    cell_header_nome.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                    cell_header_arquivo.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                    
                    row_nao_encontrados += 1
                    if total_erros:
                        # Ordena por nome do colaborador
                        for nome_colaborador, nome_arquivo in sorted(total_erros, key=lambda x: x[0]):
                            ws_relatorio.cell(row=row_nao_encontrados, column=1, value=nome_colaborador)
                            ws_relatorio.cell(row=row_nao_encontrados, column=2, value=nome_arquivo)
                            row_nao_encontrados += 1
                    else:
                        ws_relatorio.cell(row=row_nao_encontrados, column=1, value='✅ Todos encontrados!')
                    
                    # Ajusta largura das colunas
                    ws_relatorio.column_dimensions['A'].width = 20
                    ws_relatorio.column_dimensions['B'].width = 15
                    ws_relatorio.column_dimensions['C'].width = 10
                    ws_relatorio.column_dimensions['D'].width = 10
                    ws_relatorio.column_dimensions['E'].width = 15
                    ws_relatorio.column_dimensions['F'].width = 10
                    
                    # ===== CRIAR GUIA PORCENTAGENS ABS =====
                    ws_porcentagens = w.book.create_sheet('Porcentagens ABS')
                    
                    # Linha 1: Título
                    ws_porcentagens.merge_cells('A1:Z1')
                    titulo_cell = ws_porcentagens.cell(row=1, column=1, value='📊 PORCENTAGENS DE ABSENTEÍSMO')
                    titulo_cell.font = Font(bold=True, size=14, color='FFFFFF')
                    titulo_cell.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                    
                    # Linha 3: Headers - Área, HC (agregado)
                    ws_porcentagens.cell(row=3, column=1, value='Área')
                    ws_porcentagens.cell(row=3, column=2, value='HC')
                    
                    # Formata header
                    for col_num in [1, 2]:
                        cell_header = ws_porcentagens.cell(row=3, column=col_num)
                        cell_header.font = Font(bold=True, color='FFFFFF', size=10)
                        cell_header.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                        cell_header.alignment = Alignment(horizontal='center', vertical='center')
                    
                    area_col_letter = get_column_letter(list(df_mest_final.columns).index('AREA') + 1)
                    
                    # Linha 4: M&A / BLOQ com HC
                    cell_ma = ws_porcentagens.cell(row=4, column=1, value='M&A / BLOQ')
                    cell_ma.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    cell_ma.font = Font(bold=True)
                    
                    cell_hc_ma = ws_porcentagens.cell(row=4, column=2)
                    hc_ma_formula = (
                        f'=SUMPRODUCT(ISNUMBER(SEARCH("PROJETO INTERPRISE - MOVIMENTACAO E ARMAZENAGEM",Dados!{area_col_letter}:${area_col_letter}))*1)'
                        f'+SUMPRODUCT(ISNUMBER(SEARCH("MOVIMENTACAO E ARMAZENAGEM",Dados!{area_col_letter}:${area_col_letter}))*NOT(ISNUMBER(SEARCH("PROJETO INTERPRISE",Dados!{area_col_letter}:${area_col_letter})))*1)'
                        f'+SUMPRODUCT(ISNUMBER(SEARCH("BLOQ",Dados!{area_col_letter}:${area_col_letter}))*1)'
                        f'+SUMPRODUCT(ISNUMBER(SEARCH("CD-RJ | FOB",Dados!{area_col_letter}:${area_col_letter}))*1)'
                    )
                    cell_hc_ma.value = hc_ma_formula
                    cell_hc_ma.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    cell_hc_ma.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Linha 5: CRDK / D&E com HC
                    cell_crdk = ws_porcentagens.cell(row=5, column=1, value='CRDK / D&E')
                    cell_crdk.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    cell_crdk.font = Font(bold=True)
                    
                    cell_hc_crdk = ws_porcentagens.cell(row=5, column=2)
                    hc_crdk_formula = (
                        f'=SUMPRODUCT(ISNUMBER(SEARCH("CROSSDOCK DISTRIBUICAO E EXPEDICAO",Dados!{area_col_letter}:${area_col_letter}))*1)'
                        f'+SUMPRODUCT(ISNUMBER(SEARCH("CRDK D&E|CD-RJ HB",Dados!{area_col_letter}:${area_col_letter}))*1)'
                        f'+SUMPRODUCT(ISNUMBER(SEARCH("DISTRIBUICAO E EXPEDICAO",Dados!{area_col_letter}:${area_col_letter}))*NOT(ISNUMBER(SEARCH("CROSSDOCK",Dados!{area_col_letter}:${area_col_letter})))*1)'
                    )
                    cell_hc_crdk.value = hc_crdk_formula
                    cell_hc_crdk.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    cell_hc_crdk.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Linha 6: TOTAL HC
                    cell_total_hc_label = ws_porcentagens.cell(row=6, column=1, value='TOTAL HC')
                    cell_total_hc_label.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    cell_total_hc_label.font = Font(bold=True)
                    
                    cell_total_hc_value = ws_porcentagens.cell(row=6, column=2)
                    cell_total_hc_value.value = '=B4+B5'
                    cell_total_hc_value.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    cell_total_hc_value.font = Font(bold=True)
                    cell_total_hc_value.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Linha 8: Headers com datas para porcentagens - TODOS os dias do mês
                    ws_porcentagens.cell(row=8, column=1, value='Área')
                    
                    # Gera todos os dias do mês
                    if mapa_datas:
                        mes_dados = min(mapa_datas.keys()).month
                        ano_dados = min(mapa_datas.keys()).year
                    else:
                        mes_dados = mes
                        ano_dados = ano
                    
                    import calendar
                    dias_no_mes = calendar.monthrange(ano_dados, mes_dados)[1]
                    
                    # Preenche header com todos os dias (mesmo sem dados)
                    for dia in range(1, dias_no_mes + 1):
                        data_obj = datetime.date(ano_dados, mes_dados, dia)
                        data_formatada = f"{dia:02d}/{mes_dados:02d}"
                        col_idx = dia + 1  # Coluna começa em 2 (coluna 1 é "Área")
                        cell_header = ws_porcentagens.cell(row=8, column=col_idx, value=data_formatada)
                        cell_header.font = Font(bold=True, color='FFFFFF', size=10)
                        cell_header.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                        cell_header.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Formata header coluna Área
                    cell_area_header = ws_porcentagens.cell(row=8, column=1)
                    cell_area_header.font = Font(bold=True, color='FFFFFF', size=10)
                    cell_area_header.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                    cell_area_header.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Setores para porcentagens
                    setores_info_pct = [
                        ('M&A / BLOQ', ['MOVIMENTACAO E ARMAZENAGEM', 'PROJETO INTERPRISE - MOVIMENTACAO E ARMAZENAGEM', 'BLOQ', 'CD-RJ | FOB']),
                        ('M&A / BLOQ - Porcentagem', ['MOVIMENTACAO E ARMAZENAGEM', 'PROJETO INTERPRISE - MOVIMENTACAO E ARMAZENAGEM', 'BLOQ', 'CD-RJ | FOB']),
                        ('CRDK / D&E', ['CROSSDOCK DISTRIBUICAO E EXPEDICAO', 'CRDK D&E|CD-RJ HB', 'DISTRIBUICAO E EXPEDICAO', '']),
                        ('CRDK / D&E - Porcentagem', ['CROSSDOCK DISTRIBUICAO E EXPEDICAO', 'CRDK D&E|CD-RJ HB', 'DISTRIBUICAO E EXPEDICAO', ''])
                    ]
                    
                    row_pct = 9
                    
                    for setor_idx, (setor_nome, keywords_setor) in enumerate(setores_info_pct):
                        # Nome do setor
                        cell_setor = ws_porcentagens.cell(row=row_pct, column=1, value=setor_nome)
                        # Títulos em VERDE PASTEL
                        cell_setor.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                        cell_setor.font = Font(bold=True)
                        
                        # Preenche cada data - TODOS os dias do mês
                        for dia in range(1, dias_no_mes + 1):
                            col_idx = dia + 1  # Coluna começa em 2
                            cell = ws_porcentagens.cell(row=row_pct, column=col_idx)
                            
                            # Verifica se existe data para este dia
                            data_obj = datetime.date(ano_dados, mes_dados, dia)
                            
                            # Verifica se é domingo ou feriado
                            eh_domingo = data_obj.weekday() == 6
                            eh_feriado = data_obj in feriados_temp
                            
                            if 'Porcentagem' not in setor_nome:
                                # Linhas de contagem FI+FA
                                if eh_feriado:
                                    cell.value = "FERIADO"
                                elif eh_domingo:
                                    cell.value = "DOMINGO"
                                elif data_obj in mapa_datas:
                                    col_data = mapa_datas[data_obj]
                                    data_col_idx = list(df_mest_final.columns).index(col_data) + 1
                                    data_col_letter = get_column_letter(data_col_idx)
                                    
                                    if setor_nome == 'M&A / BLOQ':
                                        formula = (
                                            f'=SUMPRODUCT('
                                            f'(ISNUMBER(SEARCH("PROJETO INTERPRISE - MOVIMENTACAO E ARMAZENAGEM",Dados!{area_col_letter}:${area_col_letter}))'
                                            f'+ISNUMBER(SEARCH("MOVIMENTACAO E ARMAZENAGEM",Dados!{area_col_letter}:${area_col_letter}))*NOT(ISNUMBER(SEARCH("PROJETO INTERPRISE",Dados!{area_col_letter}:${area_col_letter})))'
                                            f'+ISNUMBER(SEARCH("BLOQ",Dados!{area_col_letter}:${area_col_letter}))'
                                            f'+ISNUMBER(SEARCH("CD-RJ | FOB",Dados!{area_col_letter}:${area_col_letter})))*'
                                            f'((Dados!{data_col_letter}:${data_col_letter}="FI")+(Dados!{data_col_letter}:${data_col_letter}="FA")))'
                                        )
                                    else:  # CRDK / D&E
                                        formula = (
                                            f'=SUMPRODUCT('
                                            f'(ISNUMBER(SEARCH("CROSSDOCK DISTRIBUICAO E EXPEDICAO",Dados!{area_col_letter}:${area_col_letter}))'
                                            f'+ISNUMBER(SEARCH("CRDK D&E|CD-RJ HB",Dados!{area_col_letter}:${area_col_letter}))'
                                            f'+ISNUMBER(SEARCH("DISTRIBUICAO E EXPEDICAO",Dados!{area_col_letter}:${area_col_letter}))*NOT(ISNUMBER(SEARCH("CROSSDOCK",Dados!{area_col_letter}:${area_col_letter}))))*'
                                            f'((Dados!{data_col_letter}:${data_col_letter}="FI")+(Dados!{data_col_letter}:${data_col_letter}="FA")))'
                                        )
                                    
                                    cell.value = formula
                                else:
                                    # Se não tem dados para este dia, deixa vazio ou 0
                                    cell.value = 0
                                
                                if eh_feriado or eh_domingo:
                                    cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell.font = Font(bold=True, color='FFFFFFFF')
                                else:
                                    cell.fill = PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid')
                            else:
                                # Linhas de porcentagem: (contagem / HC) * 100
                                if eh_feriado:
                                    cell.value = "FERIADO"
                                    cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell.font = Font(bold=True, color='FFFFFFFF')
                                elif eh_domingo:
                                    cell.value = "DOMINGO"
                                    cell.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell.font = Font(bold=True, color='FFFFFFFF')
                                else:
                                    if 'M&A / BLOQ - Porcentagem' in setor_nome:
                                        contagem_row = row_pct - 1  # Linha anterior (M&A / BLOQ)
                                        hc_cell = 'B4'  # HC está em B4
                                    else:  # CRDK / D&E - Porcentagem
                                        contagem_row = row_pct - 1  # Linha anterior (CRDK / D&E)
                                        hc_cell = 'B5'  # HC está em B5
                                    
                                    col_letter = get_column_letter(col_idx)
                                    formula_pct = f'=IFERROR(({col_letter}{contagem_row}/{hc_cell})*100,0)'
                                    cell.value = formula_pct
                                    cell.number_format = '0.00"%"'
                                    cell.fill = PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid')
                                    cell.font = Font(bold=True)
                            
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        row_pct += 1
                    
                    # Linha de TOTAL HC - mostrar HC total em todas as colunas
                    cell_total_hc_label = ws_porcentagens.cell(row=row_pct, column=1, value='TOTAL HC')
                    cell_total_hc_label.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    cell_total_hc_label.font = Font(bold=True)
                    
                    # HC Total (soma de B4 e B5) - mostra em todas as datas também
                    cell_hc_total_label = ws_porcentagens.cell(row=row_pct, column=2)
                    cell_hc_total_label.value = '=B4+B5'
                    cell_hc_total_label.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    cell_hc_total_label.font = Font(bold=True)
                    cell_hc_total_label.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Replica o HC Total em todas as colunas de data (subtraindo DESLIGADOS)
                    for dia in range(1, dias_no_mes + 1):
                        col_idx = dia + 1
                        data_obj = datetime.date(ano_dados, mes_dados, dia)
                        
                        # Verifica se é domingo ou feriado
                        eh_domingo = data_obj.weekday() == 6
                        eh_feriado = data_obj in feriados_temp
                        
                        cell_hc_data = ws_porcentagens.cell(row=row_pct, column=col_idx)
                        
                        # Se é domingo ou feriado, escreve o texto com background preto
                        if eh_feriado:
                            cell_hc_data.value = "FERIADO"
                            cell_hc_data.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                            cell_hc_data.font = Font(bold=True, color='FFFFFFFF')
                        elif eh_domingo:
                            cell_hc_data.value = "DOMINGO"
                            cell_hc_data.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                            cell_hc_data.font = Font(bold=True, color='FFFFFFFF')
                        elif data_obj in mapa_datas:
                            col_data = mapa_datas[data_obj]
                            data_col_idx = list(df_mest_final.columns).index(col_data) + 1
                            data_col_letter = get_column_letter(data_col_idx)
                            
                            # Fórmula: HC Total (B4+B5) menos a contagem de DESLIGADO nesta data
                            # COUNTIF insensível a maiúsculas/minúsculas
                            cell_hc_data.value = f'=(B4+B5)-COUNTIF(Dados!{data_col_letter}:${data_col_letter},"DESLIGADO")'
                            cell_hc_data.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                            cell_hc_data.font = Font(bold=True)
                        else:
                            # Se não tem dados, coloca 0
                            cell_hc_data.value = 0
                            cell_hc_data.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                            cell_hc_data.font = Font(bold=True)
                        
                        cell_hc_data.alignment = Alignment(horizontal='center', vertical='center')
                    
                    row_total_hc = row_pct
                    row_pct += 1
                    
                    # Linha de FI - Faltas Injustificadas
                    cell_fi_label = ws_porcentagens.cell(row=row_pct, column=1, value='FI - Faltas Injustificadas')
                    cell_fi_label.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    cell_fi_label.font = Font(bold=True)
                    
                    # HC vazio para FI
                    cell_fi_hc = ws_porcentagens.cell(row=row_pct, column=2)
                    cell_fi_hc.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    
                    # Soma de FI por data (soma das linhas 9 e 11 de FI apenas)
                    for dia in range(1, dias_no_mes + 1):
                        col_idx = dia + 1
                        data_obj = datetime.date(ano_dados, mes_dados, dia)
                        
                        # Verifica se é domingo ou feriado
                        eh_domingo = data_obj.weekday() == 6
                        eh_feriado = data_obj in feriados_temp
                        
                        if data_obj in mapa_datas:
                            col_data = mapa_datas[data_obj]
                            data_col_idx = list(df_mest_final.columns).index(col_data) + 1
                            data_col_letter = get_column_letter(data_col_idx)
                            
                            cell_fi_data = ws_porcentagens.cell(row=row_pct, column=col_idx)
                            
                            # Se é domingo ou feriado, escreve o texto com background preto
                            if eh_feriado:
                                cell_fi_data.value = "FERIADO"
                                cell_fi_data.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                cell_fi_data.font = Font(bold=True, color='FFFFFFFF')
                            elif eh_domingo:
                                cell_fi_data.value = "DOMINGO"
                                cell_fi_data.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                cell_fi_data.font = Font(bold=True, color='FFFFFFFF')
                            else:
                                # Usa as linhas 9 (M&A FI) e 11 (CRDK FI), pegando apenas a parte de FI
                                cell_fi_data.value = f'=COUNTIF(Dados!{data_col_letter}:${data_col_letter},"FI")'
                                cell_fi_data.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                                cell_fi_data.font = Font(bold=True, color='FFFFFFFF')
                            
                            cell_fi_data.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            # Se não tem dados, coloca 0
                            cell_fi_data = ws_porcentagens.cell(row=row_pct, column=col_idx)
                            cell_fi_data.value = 0
                            cell_fi_data.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                            cell_fi_data.font = Font(bold=True, color='FFFFFFFF')
                            cell_fi_data.alignment = Alignment(horizontal='center', vertical='center')
                    
                    row_fi = row_pct
                    row_pct += 1
                    
                    # Linha de FA - Faltas por Atestado
                    cell_fa_label = ws_porcentagens.cell(row=row_pct, column=1, value='FA - Faltas por Atestado')
                    cell_fa_label.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    cell_fa_label.font = Font(bold=True)
                    
                    # HC vazio para FA
                    cell_fa_hc = ws_porcentagens.cell(row=row_pct, column=2)
                    cell_fa_hc.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    
                    # Soma de FA por data
                    for dia in range(1, dias_no_mes + 1):
                        col_idx = dia + 1
                        data_obj = datetime.date(ano_dados, mes_dados, dia)
                        
                        # Verifica se é domingo ou feriado
                        eh_domingo = data_obj.weekday() == 6
                        eh_feriado = data_obj in feriados_temp
                        
                        if data_obj in mapa_datas:
                            col_data = mapa_datas[data_obj]
                            data_col_idx = list(df_mest_final.columns).index(col_data) + 1
                            data_col_letter = get_column_letter(data_col_idx)
                            
                            cell_fa_data = ws_porcentagens.cell(row=row_pct, column=col_idx)
                            
                            # Se é domingo ou feriado, escreve o texto com background preto
                            if eh_feriado:
                                cell_fa_data.value = "FERIADO"
                                cell_fa_data.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                cell_fa_data.font = Font(bold=True, color='FFFFFFFF')
                            elif eh_domingo:
                                cell_fa_data.value = "DOMINGO"
                                cell_fa_data.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                cell_fa_data.font = Font(bold=True, color='FFFFFFFF')
                            else:
                                cell_fa_data.value = f'=COUNTIF(Dados!{data_col_letter}:${data_col_letter},"FA")'
                                cell_fa_data.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                                cell_fa_data.font = Font(bold=True, color='FFFFFFFF')
                            
                            cell_fa_data.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            # Se não tem dados, coloca 0
                            cell_fa_data = ws_porcentagens.cell(row=row_pct, column=col_idx)
                            cell_fa_data.value = 0
                            cell_fa_data.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                            cell_fa_data.font = Font(bold=True, color='FFFFFFFF')
                            cell_fa_data.alignment = Alignment(horizontal='center', vertical='center')
                    
                    row_fa = row_pct
                    row_pct += 1
                    
                    # Linha de TOTAL - soma de todas as faltas (AGORA APÓS FI E FA)
                    cell_total_label = ws_porcentagens.cell(row=row_pct, column=1, value='TOTAL')
                    cell_total_label.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    cell_total_label.font = Font(bold=True)
                    
                    # HC Total (soma de B4 e B5)
                    cell_hc_total = ws_porcentagens.cell(row=row_pct, column=2)
                    cell_hc_total.value = '=B4+B5'
                    cell_hc_total.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                    cell_hc_total.font = Font(bold=True)
                    cell_hc_total.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Soma das faltas por data (linha 9 + linha 11)
                    for dia in range(1, dias_no_mes + 1):
                        col_idx = dia + 1
                        data_obj = datetime.date(ano_dados, mes_dados, dia)
                        
                        # Verifica se é domingo ou feriado
                        eh_domingo = data_obj.weekday() == 6
                        eh_feriado = data_obj in feriados_temp
                        
                        cell_total_data = ws_porcentagens.cell(row=row_pct, column=col_idx)
                        col_letter = get_column_letter(col_idx)
                        
                        # Se é domingo ou feriado, escreve o texto com background preto
                        if eh_feriado:
                            cell_total_data.value = "FERIADO"
                            cell_total_data.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                            cell_total_data.font = Font(bold=True, color='FFFFFFFF')
                        elif eh_domingo:
                            cell_total_data.value = "DOMINGO"
                            cell_total_data.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                            cell_total_data.font = Font(bold=True, color='FFFFFFFF')
                        else:
                            cell_total_data.value = f'={col_letter}9+{col_letter}11'
                            cell_total_data.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                            cell_total_data.font = Font(bold=True)
                        
                        cell_total_data.alignment = Alignment(horizontal='center', vertical='center')
                    
                    row_total_faltas = row_pct
                    row_pct += 1
                    
                    # Linha de %Acumulado - TOTAL / HC Total
                    cell_acum_label = ws_porcentagens.cell(row=row_pct, column=1, value='%Acumulado')
                    cell_acum_label.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                    cell_acum_label.font = Font(bold=True, color='FFFFFFFF')
                    
                    # Célula vazia em B
                    cell_acum_hc = ws_porcentagens.cell(row=row_pct, column=2)
                    cell_acum_hc.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                    
                    # Soma acumulada de faltas / HC do dia respectivo * 100
                    # Cores condicionais: Verde <3%, Amarelo 3-3.5%, Vermelho >3.5%
                    row_acumulado = row_pct
                    for dia in range(1, dias_no_mes + 1):
                        col_idx = dia + 1
                        data_obj = datetime.date(ano_dados, mes_dados, dia)
                        
                        # Verifica se é domingo ou feriado
                        eh_domingo = data_obj.weekday() == 6
                        eh_feriado = data_obj in feriados_temp
                        
                        cell_acum_data = ws_porcentagens.cell(row=row_pct, column=col_idx)
                        col_letter = get_column_letter(col_idx)
                        
                        # Se é domingo ou feriado, escreve o texto com background preto
                        if eh_feriado:
                            cell_acum_data.value = "FERIADO"
                            cell_acum_data.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                            cell_acum_data.font = Font(bold=True, color='FFFFFFFF')
                        elif eh_domingo:
                            cell_acum_data.value = "DOMINGO"
                            cell_acum_data.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                            cell_acum_data.font = Font(bold=True, color='FFFFFFFF')
                        else:
                            # Referencia: célula do TOTAL (row_total_faltas) / HC da data respectiva (mesmo col_letter em row_total_hc) * 100
                            cell_acum_data.value = f'=IFERROR(({col_letter}{row_total_faltas}/{col_letter}{row_total_hc})*100,0)'
                            cell_acum_data.number_format = '0.00"%"'
                            cell_acum_data.fill = PatternFill(start_color='FFF0F0F0', end_color='FFF0F0F0', fill_type='solid')
                            cell_acum_data.font = Font(bold=True)
                        
                        cell_acum_data.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Adiciona regras condicionais para %Acumulado
                    from openpyxl.formatting.rule import CellIsRule
                    # Verde: < 3% (VERDE FORTE)
                    green_fill = PatternFill(start_color='FF00B050', end_color='FF00B050', fill_type='solid')
                    green_font = Font(bold=True, color='FFFFFFFF')
                    green_rule = CellIsRule(operator='lessThan', formula=['3'], fill=green_fill, font=green_font)
                    
                    # Amarelo: >= 3% e <= 3.5% (AMARELO FORTE)
                    yellow_fill = PatternFill(start_color='FFFF9900', end_color='FFFF9900', fill_type='solid')
                    yellow_font = Font(bold=True, color='FFFFFFFF')
                    yellow_rule = CellIsRule(operator='between', formula=['3', '3.5'], fill=yellow_fill, font=yellow_font)
                    
                    # Vermelho: > 3.5% (VERMELHO FORTE)
                    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    red_font = Font(bold=True, color='FFFFFFFF')
                    red_rule = CellIsRule(operator='greaterThan', formula=['3.5'], fill=red_fill, font=red_font)
                    
                    # Aplica as regras ao intervalo de %Acumulado
                    acum_range = f'{get_column_letter(2)}{row_acumulado}:{get_column_letter(len(sorted(mapa_datas.keys()))+1)}{row_acumulado}'
                    ws_porcentagens.conditional_formatting.add(acum_range, green_rule)
                    ws_porcentagens.conditional_formatting.add(acum_range, yellow_rule)
                    ws_porcentagens.conditional_formatting.add(acum_range, red_rule)
                    
                    # Ajusta largura das colunas
                    ws_porcentagens.column_dimensions['A'].width = 25
                    ws_porcentagens.column_dimensions['B'].width = 15
                    for col_idx in range(2, len(sorted(mapa_datas.keys())) + 2):
                        ws_porcentagens.column_dimensions[get_column_letter(col_idx)].width = 12
                    
                    # ===== PORCENTAGENS POR TURNO =====
                    # Cria uma nova aba para porcentagens por turno
                    if 'TURNO' in df_mest.columns:
                        ws_turno = w.book.create_sheet('Porcentagens TURNO')
                        
                        # Linha 1: Título
                        ws_turno.merge_cells('A1:Z1')
                        titulo_turno = ws_turno.cell(row=1, column=1, value='📊 PORCENTAGENS DE ABSENTEÍSMO POR TURNO')
                        titulo_turno.font = Font(bold=True, size=14, color='FFFFFF')
                        titulo_turno.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                        
                        row_turno = 3
                        
                        # Para cada turno (1, 2, 3)
                        for turno_num in [1, 2, 3]:
                            turno_label = f'TURNO {turno_num}'
                            turno_value = f"TURNO {turno_num}"
                            
                            # Título do turno
                            ws_turno.merge_cells(f'A{row_turno}:Z{row_turno}')
                            cell_turno_header = ws_turno.cell(row=row_turno, column=1, value=turno_label)
                            cell_turno_header.font = Font(bold=True, size=12, color='FFFFFF')
                            cell_turno_header.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                            row_turno += 1
                            
                            # ===== M&A / BLOQ =====
                            # Header M&A / BLOQ com datas
                            cell_ma_header = ws_turno.cell(row=row_turno, column=1, value='M&A / BLOQ')
                            cell_ma_header.font = Font(bold=True, color='FFFFFF', size=10)
                            cell_ma_header.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                            
                            for dia in range(1, dias_no_mes + 1):
                                data_formatada = f"{dia:02d}/{mes_dados:02d}"
                                col_idx = dia + 1
                                cell_header_data_ma = ws_turno.cell(row=row_turno, column=col_idx, value=data_formatada)
                                cell_header_data_ma.font = Font(bold=True, color='FFFFFF', size=9)
                                cell_header_data_ma.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                                cell_header_data_ma.alignment = Alignment(horizontal='center', vertical='center')
                            row_turno += 1
                            
                            # FI
                            cell_fi_label = ws_turno.cell(row=row_turno, column=1, value='FI')
                            cell_fi_label.font = Font(bold=True, color='FFFFFFFF')
                            cell_fi_label.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                            turno_text = f"TURNO {turno_num}"
                            for dia in range(1, dias_no_mes + 1):
                                col_idx = dia + 1
                                data_obj = datetime.date(ano_dados, mes_dados, dia)
                                cell_fi = ws_turno.cell(row=row_turno, column=col_idx)
                                
                                # Detecta se é domingo (6) ou feriado
                                eh_domingo = data_obj.weekday() == 6
                                eh_feriado = data_obj in feriados_temp if 'feriados_temp' in locals() else False
                                
                                if eh_feriado:
                                    cell_fi.value = "FERIADO"
                                    cell_fi.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell_fi.font = Font(color='FFFFFFFF', bold=True)
                                elif eh_domingo:
                                    cell_fi.value = "DOMINGO"
                                    cell_fi.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell_fi.font = Font(color='FFFFFFFF', bold=True)
                                elif data_obj in mapa_datas:
                                    col_data = mapa_datas[data_obj]
                                    data_col_idx = list(df_mest_final.columns).index(col_data) + 1
                                    data_col_letter = get_column_letter(data_col_idx)
                                    
                                    cell_fi.value = '=SUMPRODUCT((ISNUMBER(SEARCH("' + turno_text + '";Dados!$' + turno_col_letter + ':$' + turno_col_letter + ')))*(ISNUMBER(SEARCH("PROJETO INTERPRISE - MOVIMENTACAO E ARMAZENAGEM";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("MOVIMENTACAO E ARMAZENAGEM";Dados!$' + area_col_letter + ':$' + area_col_letter + '))*NOT(ISNUMBER(SEARCH("PROJETO INTERPRISE";Dados!$' + area_col_letter + ':$' + area_col_letter + ')))+ISNUMBER(SEARCH("BLOQ";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("CD-RJ | FOB";Dados!$' + area_col_letter + ':$' + area_col_letter + ')))*(Dados!$' + data_col_letter + ':$' + data_col_letter + '="FI"))'
                                    cell_fi.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                                    cell_fi.font = Font(bold=True, color='FFFFFFFF')
                                else:
                                    cell_fi.value = 0
                                    cell_fi.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                                    cell_fi.font = Font(bold=True, color='FFFFFFFF')
                                cell_fi.alignment = Alignment(horizontal='center', vertical='center')
                            row_turno += 1
                            
                            # FA
                            cell_fa_label = ws_turno.cell(row=row_turno, column=1, value='FA')
                            cell_fa_label.font = Font(bold=True, color='FFFFFFFF')
                            cell_fa_label.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                            for dia in range(1, dias_no_mes + 1):
                                col_idx = dia + 1
                                data_obj = datetime.date(ano_dados, mes_dados, dia)
                                cell_fa = ws_turno.cell(row=row_turno, column=col_idx)
                                
                                # Detecta se é domingo (6) ou feriado
                                eh_domingo = data_obj.weekday() == 6
                                eh_feriado = data_obj in feriados_temp if 'feriados_temp' in locals() else False
                                
                                if eh_feriado:
                                    cell_fa.value = "FERIADO"
                                    cell_fa.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell_fa.font = Font(color='FFFFFFFF', bold=True)
                                elif eh_domingo:
                                    cell_fa.value = "DOMINGO"
                                    cell_fa.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell_fa.font = Font(color='FFFFFFFF', bold=True)
                                elif data_obj in mapa_datas:
                                    col_data = mapa_datas[data_obj]
                                    data_col_idx = list(df_mest_final.columns).index(col_data) + 1
                                    data_col_letter = get_column_letter(data_col_idx)
                                    
                                    cell_fa.value = '=SUMPRODUCT((ISNUMBER(SEARCH("' + turno_text + '";Dados!$' + turno_col_letter + ':$' + turno_col_letter + ')))*(ISNUMBER(SEARCH("PROJETO INTERPRISE - MOVIMENTACAO E ARMAZENAGEM";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("MOVIMENTACAO E ARMAZENAGEM";Dados!$' + area_col_letter + ':$' + area_col_letter + '))*NOT(ISNUMBER(SEARCH("PROJETO INTERPRISE";Dados!$' + area_col_letter + ':$' + area_col_letter + ')))+ISNUMBER(SEARCH("BLOQ";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("CD-RJ | FOB";Dados!$' + area_col_letter + ':$' + area_col_letter + ')))*(Dados!$' + data_col_letter + ':$' + data_col_letter + '="FA"))'
                                    cell_fa.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                                    cell_fa.font = Font(bold=True, color='FFFFFFFF')
                                else:
                                    cell_fa.value = 0
                                    cell_fa.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                                    cell_fa.font = Font(bold=True, color='FFFFFFFF')
                                cell_fa.alignment = Alignment(horizontal='center', vertical='center')
                            row_turno += 1
                            
                            # TOTAL M&A
                            cell_total_ma_label = ws_turno.cell(row=row_turno, column=1, value='TOTAL')
                            cell_total_ma_label.font = Font(bold=True, color='FFFFFFFF')
                            cell_total_ma_label.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                            for dia in range(1, dias_no_mes + 1):
                                col_idx = dia + 1
                                data_obj = datetime.date(ano_dados, mes_dados, dia)
                                cell_total_ma = ws_turno.cell(row=row_turno, column=col_idx)
                                
                                # Detecta se é domingo (6) ou feriado
                                eh_domingo = data_obj.weekday() == 6
                                eh_feriado = data_obj in feriados_temp if 'feriados_temp' in locals() else False
                                
                                # Soma FI + FA da linha anterior
                                prev_row_fi = row_turno - 2
                                prev_row_fa = row_turno - 1
                                col_letter = get_column_letter(col_idx)
                                
                                if eh_feriado:
                                    cell_total_ma.value = "FERIADO"
                                elif eh_domingo:
                                    cell_total_ma.value = "DOMINGO"
                                else:
                                    cell_total_ma.value = f'={col_letter}{prev_row_fi}+{col_letter}{prev_row_fa}'
                                
                                if eh_feriado or eh_domingo:
                                    cell_total_ma.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell_total_ma.font = Font(color='FFFFFFFF', bold=True)
                                else:
                                    # TOTAL com verde escuro Profarma
                                    cell_total_ma.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                                    cell_total_ma.font = Font(color='FFFFFFFF', bold=True)
                                cell_total_ma.alignment = Alignment(horizontal='center', vertical='center')
                            row_turno += 2  # Espaço
                            
                            # ===== CRDK / D&E =====
                            # Header CRDK / D&E com datas
                            cell_crdk_header = ws_turno.cell(row=row_turno, column=1, value='CRDK / D&E')
                            cell_crdk_header.font = Font(bold=True, color='FFFFFF', size=10)
                            cell_crdk_header.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                            
                            for dia in range(1, dias_no_mes + 1):
                                data_formatada = f"{dia:02d}/{mes_dados:02d}"
                                col_idx = dia + 1
                                cell_header_data_crdk = ws_turno.cell(row=row_turno, column=col_idx, value=data_formatada)
                                cell_header_data_crdk.font = Font(bold=True, color='FFFFFF', size=9)
                                cell_header_data_crdk.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                                cell_header_data_crdk.alignment = Alignment(horizontal='center', vertical='center')
                            row_turno += 1
                            
                            # FI CRDK
                            cell_fi_crdk_label = ws_turno.cell(row=row_turno, column=1, value='FI')
                            cell_fi_crdk_label.font = Font(bold=True, color='FFFFFFFF')
                            cell_fi_crdk_label.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                            for dia in range(1, dias_no_mes + 1):
                                col_idx = dia + 1
                                data_obj = datetime.date(ano_dados, mes_dados, dia)
                                cell_fi_crdk = ws_turno.cell(row=row_turno, column=col_idx)
                                
                                # Detecta se é domingo (6) ou feriado
                                eh_domingo = data_obj.weekday() == 6
                                eh_feriado = data_obj in feriados_temp if 'feriados_temp' in locals() else False
                                
                                if eh_feriado:
                                    cell_fi_crdk.value = "FERIADO"
                                    cell_fi_crdk.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell_fi_crdk.font = Font(color='FFFFFFFF', bold=True)
                                elif eh_domingo:
                                    cell_fi_crdk.value = "DOMINGO"
                                    cell_fi_crdk.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell_fi_crdk.font = Font(color='FFFFFFFF', bold=True)
                                elif data_obj in mapa_datas:
                                    col_data = mapa_datas[data_obj]
                                    data_col_idx = list(df_mest_final.columns).index(col_data) + 1
                                    data_col_letter = get_column_letter(data_col_idx)
                                    
                                    cell_fi_crdk.value = '=SUMPRODUCT((ISNUMBER(SEARCH("' + turno_text + '";Dados!$' + turno_col_letter + ':$' + turno_col_letter + ')))*(ISNUMBER(SEARCH("CROSSDOCK DISTRIBUICAO E EXPEDICAO";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("CRDK D&E|CD-RJ HB";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("DISTRIBUICAO E EXPEDICAO";Dados!$' + area_col_letter + ':$' + area_col_letter + '))*NOT(ISNUMBER(SEARCH("CROSSDOCK";Dados!$' + area_col_letter + ':$' + area_col_letter + '))))*(Dados!$' + data_col_letter + ':$' + data_col_letter + '="FI"))'
                                    cell_fi_crdk.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                                    cell_fi_crdk.font = Font(bold=True, color='FFFFFFFF')
                                else:
                                    cell_fi_crdk.value = 0
                                    cell_fi_crdk.fill = PatternFill(start_color='FF007864', end_color='FF007864', fill_type='solid')
                                    cell_fi_crdk.font = Font(bold=True, color='FFFFFFFF')
                                cell_fi_crdk.alignment = Alignment(horizontal='center', vertical='center')
                            row_turno += 1
                            
                            # FA CRDK
                            cell_fa_crdk_label = ws_turno.cell(row=row_turno, column=1, value='FA')
                            cell_fa_crdk_label.font = Font(bold=True, color='FFFFFFFF')
                            cell_fa_crdk_label.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                            for dia in range(1, dias_no_mes + 1):
                                col_idx = dia + 1
                                data_obj = datetime.date(ano_dados, mes_dados, dia)
                                cell_fa_crdk = ws_turno.cell(row=row_turno, column=col_idx)
                                
                                # Detecta se é domingo (6) ou feriado
                                eh_domingo = data_obj.weekday() == 6
                                eh_feriado = data_obj in feriados_temp if 'feriados_temp' in locals() else False
                                
                                if eh_feriado:
                                    cell_fa_crdk.value = "FERIADO"
                                    cell_fa_crdk.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell_fa_crdk.font = Font(color='FFFFFFFF', bold=True)
                                elif eh_domingo:
                                    cell_fa_crdk.value = "DOMINGO"
                                    cell_fa_crdk.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                    cell_fa_crdk.font = Font(color='FFFFFFFF', bold=True)
                                elif data_obj in mapa_datas:
                                    col_data = mapa_datas[data_obj]
                                    data_col_idx = list(df_mest_final.columns).index(col_data) + 1
                                    data_col_letter = get_column_letter(data_col_idx)
                                    
                                    cell_fa_crdk.value = '=SUMPRODUCT((ISNUMBER(SEARCH("' + turno_text + '";Dados!$' + turno_col_letter + ':$' + turno_col_letter + ')))*(ISNUMBER(SEARCH("CROSSDOCK DISTRIBUICAO E EXPEDICAO";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("CRDK D&E|CD-RJ HB";Dados!$' + area_col_letter + ':$' + area_col_letter + '))+ISNUMBER(SEARCH("DISTRIBUICAO E EXPEDICAO";Dados!$' + area_col_letter + ':$' + area_col_letter + '))*NOT(ISNUMBER(SEARCH("CROSSDOCK";Dados!$' + area_col_letter + ':$' + area_col_letter + '))))*(Dados!$' + data_col_letter + ':$' + data_col_letter + '="FA"))'
                                    cell_fa_crdk.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                                    cell_fa_crdk.font = Font(bold=True, color='FFFFFFFF')
                                else:
                                    cell_fa_crdk.value = 0
                                    cell_fa_crdk.fill = PatternFill(start_color='FF008C4B', end_color='FF008C4B', fill_type='solid')
                                    cell_fa_crdk.font = Font(bold=True, color='FFFFFFFF')
                                cell_fa_crdk.alignment = Alignment(horizontal='center', vertical='center')
                            row_turno += 1
                            
                            # TOTAL CRDK
                            cell_total_crdk_label = ws_turno.cell(row=row_turno, column=1, value='TOTAL')
                            cell_total_crdk_label.font = Font(bold=True, color='FFFFFFFF')
                            cell_total_crdk_label.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                            for dia in range(1, dias_no_mes + 1):
                                col_idx = dia + 1
                                data_obj = datetime.date(ano_dados, mes_dados, dia)
                                cell_total_crdk = ws_turno.cell(row=row_turno, column=col_idx)
                                
                                # Detecta se é domingo (6) ou feriado
                                eh_domingo = data_obj.weekday() == 6
                                eh_feriado = data_obj in feriados_temp if 'feriados_temp' in locals() else False
                                
                                # Soma FI + FA da linha anterior
                                prev_row_fi = row_turno - 2
                                prev_row_fa = row_turno - 1
                                col_letter = get_column_letter(col_idx)
                                
                                if eh_feriado:
                                    cell_total_crdk.value = "FERIADO"
                                elif eh_domingo:
                                    cell_total_crdk.value = "DOMINGO"
                                else:
                                    cell_total_crdk.value = f'={col_letter}{prev_row_fi}+{col_letter}{prev_row_fa}'
                                
                                # TOTAL com verde escuro Profarma
                                if eh_feriado or eh_domingo:
                                    cell_total_crdk.fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                                else:
                                    cell_total_crdk.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                                cell_total_crdk.font = Font(color='FFFFFFFF', bold=True)
                                cell_total_crdk.alignment = Alignment(horizontal='center', vertical='center')
                            row_turno += 3  # Espaço entre turnos
                        
                        
                        # Ajusta largura das colunas
                        ws_turno.column_dimensions['A'].width = 25
                        for col_idx in range(2, dias_no_mes + 2):
                            ws_turno.column_dimensions[get_column_letter(col_idx)].width = 10
                    
                    # ===== CRIAR GUIA DE GRÁFICOS =====
                    ws_graficos = w.book.create_sheet('Gráficos')
                    
                    # Linha 1: Título
                    ws_graficos.merge_cells('A1:H1')
                    titulo_graficos = ws_graficos.cell(row=1, column=1, value='📊 ANÁLISE GRÁFICA DE ABSENTEÍSMO')
                    titulo_graficos.font = Font(bold=True, size=14, color='FFFFFF')
                    titulo_graficos.fill = PatternFill(start_color='FF0D4F45', end_color='FF0D4F45', fill_type='solid')
                    ws_graficos.row_dimensions[1].height = 25
                    
                    from openpyxl.chart import PieChart, BarChart, Reference
                    from openpyxl.worksheet.datavalidation import DataValidation
                    
                    # ===== SEÇÃO 1: Seletor de Data =====
                    row_selector = 3
                    ws_graficos.cell(row=row_selector, column=1, value='📅 Selecione a Data:').font = Font(bold=True, size=11)
                    
                    # Cria lista de datas para o dropdown - TODOS os dias do mês
                    datas_lista = sorted(mapa_datas.keys())
                    mes_atual = datas_lista[0].month if datas_lista else 1
                    ano_atual = datas_lista[0].year if datas_lista else 2025
                    
                    # Gera lista com todos os dias do mês (1-31)
                    import calendar
                    dias_no_mes = calendar.monthrange(ano_atual, mes_atual)[1]
                    datas_completas = [f"{dia:02d}/{mes_atual:02d}" for dia in range(1, dias_no_mes + 1)]
                    datas_formatadas = ','.join(datas_completas)
                    
                    # Data Validation na célula B3
                    dv = DataValidation(type='list', formula1=f'"{datas_formatadas}"', allow_blank=False)
                    dv.error = 'Por favor, selecione uma data da lista'
                    dv.errorTitle = 'Seleção Inválida'
                    ws_graficos.add_data_validation(dv)
                    
                    # Define valor padrão (primeira data com dados)
                    cell_selector = ws_graficos.cell(row=row_selector, column=2, value=datas_lista[0].strftime('%d/%m'))
                    cell_selector.fill = PatternFill(start_color='FFFFECC8', end_color='FFFFECC8', fill_type='solid')
                    cell_selector.font = Font(bold=True, size=11)
                    cell_selector.number_format = '@'  # Formato de texto para manter como "dd/mm"
                    dv.add(cell_selector)
                    
                    # ===== SEÇÃO 2: Gráficos Dinâmicos =====
                    row_grafico = 6
                    
                    # Células de cálculo ocultas para dados dinâmicos
                    # Coluna J e K para dados de FI/FA
                    # Coluna L e M para dados de setores
                    
                    ws_graficos.column_dimensions['J'].hidden = True
                    ws_graficos.column_dimensions['K'].hidden = True
                    ws_graficos.column_dimensions['L'].hidden = True
                    ws_graficos.column_dimensions['M'].hidden = True
                    
                    # Cria lista de colunas de data no Relatório para MATCH
                    col_letras_datas = []
                    for data_idx, data_obj in enumerate(datas_lista):
                        col_letra = get_column_letter(data_idx + 2)  # Começa na coluna B (coluna 2) na aba Porcentagens
                        col_letras_datas.append((data_obj.strftime('%d/%m'), col_letra))
                    
                    # ===== GRÁFICO 1: Faltas por Tipo (DINÂMICO) =====
                    ws_graficos.cell(row=row_grafico, column=1, value='Faltas por Tipo').font = Font(bold=True, size=11)
                    
                    row_data = row_grafico + 1
                    ws_graficos.cell(row=row_data, column=1, value='Tipo').font = Font(bold=True)
                    ws_graficos.cell(row=row_data, column=2, value='Quantidade').font = Font(bold=True)
                    ws_graficos.cell(row=row_data, column=1).fill = PatternFill(start_color='FFC5D9F1', end_color='FFC5D9F1', fill_type='solid')
                    ws_graficos.cell(row=row_data, column=2).fill = PatternFill(start_color='FFC5D9F1', end_color='FFC5D9F1', fill_type='solid')
                    
                    # Dados FI
                    row_data += 1
                    ws_graficos.cell(row=row_data, column=1, value='FI - Injustificadas').font = Font(bold=True)
                    cell_fi = ws_graficos.cell(row=row_data, column=2)
                    # Fórmula que busca a coluna da data selecionada e retorna FI
                    cell_fi.value = "=IFERROR(INDEX('Porcentagens ABS'!15:15,MATCH(B3,'Porcentagens ABS'!8:8,0)),0)"
                    cell_fi.fill = PatternFill(start_color='FFFFE6E6', end_color='FFFFE6E6', fill_type='solid')
                    
                    # Dados FA
                    row_data += 1
                    ws_graficos.cell(row=row_data, column=1, value='FA - Atestado').font = Font(bold=True)
                    cell_fa = ws_graficos.cell(row=row_data, column=2)
                    # Fórmula que busca a coluna da data selecionada e retorna FA
                    cell_fa.value = "=IFERROR(INDEX('Porcentagens ABS'!16:16,MATCH(B3,'Porcentagens ABS'!8:8,0)),0)"
                    cell_fa.fill = PatternFill(start_color='FFFFECC8', end_color='FFFFECC8', fill_type='solid')
                    
                    row_fi_fa_data = row_data
                    
                    # Cria gráfico de pizza para tipos de faltas
                    pie_chart_1 = PieChart()
                    pie_chart_1.title = 'Faltas por Tipo (Data Selecionada)'
                    pie_chart_1.style = 10
                    labels = Reference(ws_graficos, min_col=1, min_row=row_grafico+2, max_row=row_fi_fa_data)
                    data = Reference(ws_graficos, min_col=2, min_row=row_grafico+1, max_row=row_fi_fa_data)
                    pie_chart_1.add_data(data, titles_from_data=True)
                    pie_chart_1.set_categories(labels)
                    pie_chart_1.height = 10
                    pie_chart_1.width = 13
                    ws_graficos.add_chart(pie_chart_1, 'A10')
                    
                    # ===== GRÁFICO 2: Faltas por Setor (DINÂMICO) =====
                    col_grafico_setor = 5
                    ws_graficos.cell(row=row_grafico, column=col_grafico_setor, value='Faltas por Setor').font = Font(bold=True, size=11)
                    
                    row_data = row_grafico + 1
                    ws_graficos.cell(row=row_data, column=col_grafico_setor, value='Setor').font = Font(bold=True)
                    ws_graficos.cell(row=row_data, column=col_grafico_setor+1, value='Faltas').font = Font(bold=True)
                    ws_graficos.cell(row=row_data, column=col_grafico_setor).fill = PatternFill(start_color='FFC5D9F1', end_color='FFC5D9F1', fill_type='solid')
                    ws_graficos.cell(row=row_data, column=col_grafico_setor+1).fill = PatternFill(start_color='FFC5D9F1', end_color='FFC5D9F1', fill_type='solid')
                    
                    # Dados M&A / BLOQ
                    row_data += 1
                    ws_graficos.cell(row=row_data, column=col_grafico_setor, value='M&A / BLOQ').font = Font(bold=True)
                    cell_ma = ws_graficos.cell(row=row_data, column=col_grafico_setor+1)
                    # Fórmula que busca a coluna da data selecionada e retorna M&A
                    cell_ma.value = "=IFERROR(INDEX('Porcentagens ABS'!9:9,MATCH(B3,'Porcentagens ABS'!8:8,0)),0)"
                    cell_ma.fill = PatternFill(start_color='FFE8F5E0', end_color='FFE8F5E0', fill_type='solid')
                    
                    # Dados CRDK / D&E
                    row_data += 1
                    ws_graficos.cell(row=row_data, column=col_grafico_setor, value='CRDK / D&E').font = Font(bold=True)
                    cell_crdk = ws_graficos.cell(row=row_data, column=col_grafico_setor+1)
                    # Fórmula que busca a coluna da data selecionada e retorna CRDK
                    cell_crdk.value = "=IFERROR(INDEX('Porcentagens ABS'!11:11,MATCH(B3,'Porcentagens ABS'!8:8,0)),0)"
                    cell_crdk.fill = PatternFill(start_color='FFE6F2FF', end_color='FFE6F2FF', fill_type='solid')
                    
                    row_setor_data = row_data
                    
                    # Cria gráfico de pizza para setores
                    pie_chart_2 = PieChart()
                    pie_chart_2.title = 'Faltas por Setor (Data Selecionada)'
                    pie_chart_2.style = 10
                    labels_2 = Reference(ws_graficos, min_col=col_grafico_setor, min_row=row_grafico+2, max_row=row_setor_data)
                    data_2 = Reference(ws_graficos, min_col=col_grafico_setor+1, min_row=row_grafico+1, max_row=row_setor_data)
                    pie_chart_2.add_data(data_2, titles_from_data=True)
                    pie_chart_2.set_categories(labels_2)
                    pie_chart_2.height = 10
                    pie_chart_2.width = 13
                    ws_graficos.add_chart(pie_chart_2, 'F10')
                    
                    # Ajusta largura das colunas
                    ws_graficos.column_dimensions['A'].width = 25
                    ws_graficos.column_dimensions['B'].width = 15
                    ws_graficos.column_dimensions['E'].width = 25
                    ws_graficos.column_dimensions['F'].width = 15
                    
                    # ===== OBTER FERIADOS E MARCAR NA PLANILHA =====
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    status_text.info("📥 Obtendo feriados nacionais...")
                    progress_bar.progress(10)
                    
                    if mapa_datas:
                        ano_feriados = min(mapa_datas.keys()).year
                        feriados = obter_feriados_brasil(ano_feriados)
                        if feriados:
                            status_text.info("🎨 Marcando feriados na planilha...")
                            progress_bar.progress(20)
                            marcar_feriados_na_workbook(w.book, feriados, mapa_datas, MAPA_CORES)
                    
                    # ===== LER DATAFRAME ATUALIZADO DO WORKBOOK (COM FERIADOS MARCADOS) =====
                    status_text.info("📖 Lendo dados marcados...")
                    progress_bar.progress(30)
                    df_mest_com_feriados = ler_dataframe_do_workbook(w.book)
                    
                    # ===== DETECTAR AFASTAMENTOS NO DATAFRAME COM FERIADOS (ignora FERIADO) =====
                    status_text.info("🔍 Detectando afastamentos...")
                    progress_bar.progress(40)
                    afastamentos = detectar_afastamentos_no_dataframe(df_mest_com_feriados, mapa_datas)
                    
                    # ===== MARCAR AFASTAMENTOS NA PLANILHA =====
                    status_text.info("📌 Marcando afastamentos...")
                    progress_bar.progress(50)
                    marcar_afastamentos_na_workbook(w.book, MAPA_CORES, afastamentos, df_mest_com_feriados, mapa_datas)
                    
                    # ===== LER DATAFRAME ATUALIZADO DO WORKBOOK (COM MARCAÇÕES) =====
                    status_text.info("📖 Lendo dados finais...")
                    progress_bar.progress(60)
                    df_mest_marcado = ler_dataframe_do_workbook(w.book)
                    
                    # ===== CRIAR SHEET DE OFENSORES DE ABS (COM DADOS MARCADOS) =====
                    status_text.info("📊 Gerando relatório de ofensores...")
                    progress_bar.progress(70)
                    criar_sheet_ofensores_abs(df_mest_marcado, w, mapa_datas, MAPA_CORES, afastamentos)
                    
                    # ===== CRIAR SHEET DE RANKING DE ABS =====
                    status_text.info("🏆 Gerando ranking de absenteísmo...")
                    progress_bar.progress(72)
                    
                    criar_sheet_ranking_abs(df_mest_marcado, w, MAPA_CORES)
                    
                    # ===== ENRIQUECER RANKING COM DADOS DO CSV =====
                    status_text.info("📊 Capturando dados do CSV de colaboradores...")
                    progress_bar.progress(73)
                    
                    if file_colaboradores is not None:
                        try:
                            # Carrega CSV de colaboradores
                            file_colaboradores.seek(0)
                            if file_colaboradores.name.endswith('.xlsx'):
                                df_colab_para_ranking = pd.read_excel(file_colaboradores)
                            else:
                                # Tenta diferentes encodings e separadores para CSV
                                encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
                                separadores = [',', ';', '\t', '|']
                                df_colab_para_ranking = None
                                
                                for enc in encodings:
                                    for sep in separadores:
                                        try:
                                            file_colaboradores.seek(0)
                                            # Skip primeira linha se for só "Colaboradores"
                                            df_colab_para_ranking = pd.read_csv(file_colaboradores, encoding=enc, sep=sep, skiprows=1)
                                            break
                                        except Exception as e:
                                            continue
                                    if df_colab_para_ranking is not None:
                                        break
                            
                            if df_colab_para_ranking is not None:
                                # Re-gera TOP 10 para enriquecimento
                                colunas_datas = [col for col in df_mest_marcado.columns if col not in ['NOME', 'FUNÇÃO', 'SITUAÇÃO', 'AREA', 'GESTOR', 'SUPERVISOR', 'NOME_LIMPO']]
                                df_ranking_temp = pd.DataFrame({
                                    'NOME': df_mest_marcado['NOME'],
                                    'GESTOR': df_mest_marcado['GESTOR'],
                                    'FUNÇÃO': df_mest_marcado['FUNÇÃO'],
                                    'AREA': df_mest_marcado['AREA'],
                                    'FI': df_mest_marcado[colunas_datas].apply(lambda row: (row == 'FI').sum(), axis=1),
                                    'FA': df_mest_marcado[colunas_datas].apply(lambda row: (row == 'FA').sum(), axis=1),
                                }).copy()
                                df_ranking_temp = df_ranking_temp[df_ranking_temp['NOME'].notna() & (df_ranking_temp['NOME'] != '')]
                                
                                top10_fa_display = df_ranking_temp.nlargest(10, 'FA')
                                top10_fi_display = df_ranking_temp.nlargest(10, 'FI')
                                
                                # Enriquece com dados do CSV
                                top10_fa_display, top10_fi_display = enriquecer_ranking_com_dados_csv(top10_fa_display, top10_fi_display, df_colab_para_ranking)
                                
                                # Recreia o sheet de ranking com dados enriquecidos
                                status_text.info("✅ Atualizando ranking com dados do CSV...")
                                progress_bar.progress(74)
                                
                                # Remove o sheet anterior (se existir)
                                if 'Ranking ABS' in w.book.sheetnames:
                                    del w.book['Ranking ABS']
                                
                                # Cria novo sheet com dados enriquecidos
                                criar_sheet_ranking_abs(df_mest_marcado, w, MAPA_CORES, top10_fa_display, top10_fi_display)
                                
                                status_text.info("✅ Ranking atualizado com sucesso!")
                                progress_bar.progress(75)
                        except Exception as e:
                            st.warning(f"⚠️ Não foi possível enriquecer ranking com CSV: {str(e)}")
                    
                    # ===== COLORIR CÉLULAS INCOMUNS NA PLANILHA DADOS =====
                    status_text.info("🎯 Marcando presença incomum...")
                    progress_bar.progress(75)
                    colorir_celulas_incomuns_dados(w, MAPA_CORES, mapa_datas)
                    
                    # ===== REMOVER BORDAS E MUDAR BACKGROUND PARA BRANCO =====
                    status_text.info("🎨 Finalizando formatação...")
                    progress_bar.progress(80)
                    
                    from openpyxl.styles import Border, Side
                    
                    # Define borda vazia
                    no_border = Border(
                        left=Side(style=None),
                        right=Side(style=None),
                        top=Side(style=None),
                        bottom=Side(style=None)
                    )
                    white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
                    
                    # Aplica a todas as abas EXCETO Gráficos
                    for ws_name in w.book.sheetnames:
                        if ws_name != 'Gráficos':  # Ignora a aba de Gráficos
                            worksheet = w.book[ws_name]
                            for row in worksheet.iter_rows():
                                for cell in row:
                                    cell.border = no_border
                                    # Só muda background se não tiver cor específica atribuída (mantém cores de header e dados)
                                    if cell.fill.start_color.index == '00000000' or cell.fill.start_color.index == 'FFFFFFFF' or cell.fill.start_color.index == '0':
                                        cell.fill = white_fill
                    
                    out.seek(0)
                
                # Gera nome do arquivo no padrão solicitado
                meses_nomes = {
                    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho',
                    7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
                }
                mes_nome = meses_nomes.get(mes, 'Mês')
                nome_arquivo = f"{mes:02d}- Controle de Absenteismo - {mes_nome}.xlsx"
                
                # Cria versão SEM FÓRMULAS (valores apenas - mais leve)
                from openpyxl.utils import get_column_letter
                out_sem_formulas = io.BytesIO()
                
                # Carrega o workbook com fórmulas
                wb_com_formulas = load_workbook(out)
                
                # Cria um novo workbook para a versão sem fórmulas
                wb_sem_formulas = Workbook()
                wb_sem_formulas.remove(wb_sem_formulas.active)  # Remove sheet padrão
                
                # Copia todas as abas convertendo fórmulas em valores
                for sheet_origin in wb_com_formulas.sheetnames:
                    ws_origin = wb_com_formulas[sheet_origin]
                    ws_new = wb_sem_formulas.create_sheet(sheet_origin)
                    
                    # Copia as dimensões
                    for row in ws_origin.iter_rows():
                        for cell in row:
                            new_cell = ws_new[cell.coordinate]
                            
                            # Copia valor (não fórmula)
                            if cell.value is not None:
                                # Se é fórmula, tenta calcular; senão copia o valor
                                if isinstance(cell.value, str) and cell.value.startswith('='):
                                    # Deixa em branco ou copia a fórmula como texto (não executa)
                                    new_cell.value = cell.value
                                else:
                                    new_cell.value = cell.value
                            
                            # Copia formatação
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                    
                    # Copia largura das colunas
                    for col_letter, col_dimension in ws_origin.column_dimensions.items():
                        ws_new.column_dimensions[col_letter].width = col_dimension.width
                    
                    # Copia altura das linhas
                    for row_num, row_dimension in ws_origin.row_dimensions.items():
                        ws_new.row_dimensions[row_num].height = row_dimension.height
                
                # Salva workbook sem fórmulas
                wb_sem_formulas.save(out_sem_formulas)
                out_sem_formulas.seek(0)
                
                # Finaliza barra de progresso
                status_text.success("✅ Processamento concluído com sucesso!")
                progress_bar.progress(100)
                
                st.divider()
                
                # Dois botões de download lado a lado
                col_download1, col_download2 = st.columns(2)
                
                with col_download1:
                    st.download_button(
                        "📊 COM Fórmulas (Mais Pesado)",
                        out.getvalue(),
                        f"COM_FORMULAS_{nome_arquivo}",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_com_formulas"
                    )
                
                with col_download2:
                    st.download_button(
                        "📋 SEM Fórmulas (Mais Leve)",
                        out_sem_formulas.getvalue(),
                        f"SEM_FORMULAS_{nome_arquivo}",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_sem_formulas"
                    )
            except Exception as e:
                st.error(f"❌ Erro durante o processamento: {str(e)}")

# CSS final para garantir full-width em todo o app
st.markdown("""
<style>
    * {
        box-sizing: border-box !important;
    }
    
    [data-testid="stAppViewContainer"] {
        max-width: 100% !important;
    }
    
    [data-testid="stMain"] {
        max-width: 100% !important;
    }
    
    .main > .block-container {
        max-width: 100% !important;
        padding: 1rem !important;
    }
    
    .stTabs [role="tablist"] {
        width: 100% !important;
    }
    
    [data-testid="column"] {
        flex: 1 1 calc(100% - 1rem) !important;
        width: 100% !important;
    }
    
    .st-emotion-cache-9edo8l.e1wguzas2 {
        flex: 1 1 calc(100% - 1rem) !important;
        width: 100% !important;
    }
    
    section[data-testid="stContainer"] {
        width: 100% !important;
    }
    
    div.stDataFrame {
        width: 100% !important;
        max-width: 100% !important;
    }
    
    .dataframe {
        width: 100% !important;
    }
</style>
""", unsafe_allow_html=True)











